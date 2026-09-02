from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib import messages
from django.http import JsonResponse, HttpResponseForbidden
from django.db.models import Count, Q, Avg, Max
from django.db.models.functions import ExtractWeekDay
from datetime import date as date_type
from django.utils import timezone
from datetime import datetime, time, timedelta
import openpyxl

from .models import ImportBatch, EgesRow, DirectorToken, NombreObraSocial
from .forms import ImportarEGESForm
from .services import (
    get_base_estudios_finalizados as _base_estudios_finalizados,
    get_base_rx_sin_informe as _base_rx_sin_informe,
    aplicar_filtros_fecha_modalidad as _aplicar_filtros_fecha_modalidad,
    calcular_kpis as _calcular_kpis,
    agrupar_periodo as _agrupar_periodo,
    verificar_token as _verificar_token,
    procesar_excel_eges as _procesar_excel_eges,
    calcular_sha256_archivo as _calcular_sha256_archivo,
)

# Paleta de colores consistente para todas las vistas y el portal del director
COLORES_MODALIDAD = {
    'TC':    {'border': 'rgb(59, 130, 246)',   'bg': 'rgba(59, 130, 246, 0.15)'},   # Azul
    'RM':    {'border': 'rgb(147, 51, 234)',   'bg': 'rgba(147, 51, 234, 0.15)'},   # Púrpura
    'RX':    {'border': 'rgb(34, 197, 94)',    'bg': 'rgba(34, 197, 94, 0.15)'},    # Verde
    'DX':    {'border': 'rgb(249, 115, 22)',   'bg': 'rgba(249, 115, 22, 0.15)'},   # Naranja
    'MAM':   {'border': 'rgb(236, 72, 153)',   'bg': 'rgba(236, 72, 153, 0.15)'},   # Rosa
    'ECO':   {'border': 'rgb(251, 191, 36)',   'bg': 'rgba(251, 191, 36, 0.15)'},   # Amarillo
    'SERIE': {'border': 'rgb(20, 184, 166)',   'bg': 'rgba(20, 184, 166, 0.15)'},   # Teal
    'OTROS': {'border': 'rgb(107, 114, 128)',  'bg': 'rgba(107, 114, 128, 0.15)'},  # Gris
}


def es_superuser(user):
    """Verifica que el usuario sea superusuario."""
    return user.is_superuser


@login_required
@user_passes_test(es_superuser, login_url='/admin/')
def lista_batches(request):
    """
    Vista principal: lista de todos los lotes importados.
    Solo accesible por superusuarios.
    """
    batches = ImportBatch.objects.all().select_related('usuario')
    
    # Calcular estadísticas generales
    from django.db.models import Sum
    stats = batches.aggregate(
        total_filas=Sum('total_filas'),
        total_finalizados=Sum('total_estudios_finalizados')
    )
    
    context = {
        'batches': batches,
        'total_batches': batches.count(),
        'total_filas': stats['total_filas'] or 0,
        'total_finalizados': stats['total_finalizados'] or 0,
        'ultimo_batch': batches.first(),
        'titulo_pagina': 'Importación EGES',
    }
    return render(request, 'eges_import/lista_batches.html', context)


@login_required
@user_passes_test(es_superuser, login_url='/admin/')
def dashboard_global(request):
    """
    Dashboard global que consolida TODOS los batches importados.
    Vista ejecutiva para el director.
    """
    # Obtener todos los batches
    batches = ImportBatch.objects.all()
    
    # Métricas globales consolidadas
    total_batches = batches.count()
    lotes_posible_tope = batches.filter(total_filas__gte=1000).count()
    ultimo_batch = batches.order_by('-fecha_importacion').first()
    total_filas_global = sum(b.total_filas for b in batches)
    total_estudios_finalizados_global = sum(b.total_estudios_finalizados for b in batches)
    
    # Totales por modalidad (consolidado)
    total_tc_global = sum(b.total_tc for b in batches)
    total_rm_global = sum(b.total_rm for b in batches)
    total_rx_global = sum(b.total_rx for b in batches)
    total_eco_global = sum(b.total_eco for b in batches)
    total_otros_global = sum(b.total_otros for b in batches)
    
    # Obtener rango de fechas de todos los datos
    todas_filas = EgesRow.objects.filter(fecha_turno__isnull=False)
    fecha_min = todas_filas.order_by('fecha_turno').first()
    fecha_max = todas_filas.order_by('-fecha_turno').first()
    
    context = {
        'titulo_pagina': 'Dashboard Global EGES',
        'total_batches': total_batches,
        'lotes_posible_tope': lotes_posible_tope,
        'ultimo_batch': ultimo_batch,
        'total_filas_global': total_filas_global,
        'total_estudios_finalizados_global': total_estudios_finalizados_global,
        'total_tc_global': total_tc_global,
        'total_rm_global': total_rm_global,
        'total_rx_global': total_rx_global,
        'total_eco_global': total_eco_global,
        'total_otros_global': total_otros_global,
        'fecha_min': fecha_min.fecha_turno if fecha_min else None,
        'fecha_max': fecha_max.fecha_turno if fecha_max else None,
        'batches': batches,
    }
    return render(request, 'eges_import/dashboard_global.html', context)


@login_required
@user_passes_test(es_superuser, login_url='/admin/')
def importar_eges(request):
    """
    Vista para subir un archivo Excel EGES.
    Procesa el archivo, crea un batch y las filas correspondientes.
    """
    if request.method == 'POST':
        form = ImportarEGESForm(request.POST, request.FILES)
        
        if form.is_valid():
            archivo = request.FILES['archivo']
            archivo_sha256 = _calcular_sha256_archivo(archivo)

            lote_existente = ImportBatch.objects.filter(
                archivo_sha256=archivo_sha256,
            ).first()
            if lote_existente:
                messages.warning(
                    request,
                    'Este archivo ya fue importado exactamente igual '
                    f'el {timezone.localtime(lote_existente.fecha_importacion):%d/%m/%Y a las %H:%M}. '
                    'No se realizaron cambios.',
                )
                return redirect('eges_import:detalle_batch', batch_id=lote_existente.id)
            
            print(f"[EGES] Iniciando importación: {archivo.name} ({archivo.size} bytes)")
            
            # Crear el batch
            batch = ImportBatch.objects.create(
                usuario=request.user,
                archivo_nombre=archivo.name,
                archivo_sha256=archivo_sha256,
            )
            
            print(f"[EGES] Batch #{batch.id} creado")
            
            try:
                # Procesar el Excel
                print(f"[EGES] Procesando Excel...")
                resultado = _procesar_excel_eges(archivo, batch)
                
                print(f"[EGES] Resultado: {resultado['creadas']} nuevas, {resultado['duplicadas']} duplicadas, {resultado['errores']} errores")

                # Compatibilidad con lotes históricos, que no tienen huella porque
                # el archivo original no se conservaba: si todo el contenido ya
                # existe, descartamos solamente este lote vacío y no alteramos
                # ninguna fila ni lote previo.
                if resultado['creadas'] == 0 and resultado['duplicadas'] > 0:
                    batch.delete()
                    messages.warning(
                        request,
                        'Todos los estudios de este archivo ya estaban cargados. '
                        'No se creó otro lote y no se realizaron cambios.',
                    )
                    return redirect('eges_import:lista_batches')
                
                # Calcular métricas
                batch.calcular_metricas()
                
                print(f"[EGES] Batch #{batch.id} completado exitosamente")
                
                # Mensaje personalizado según si hubo duplicados
                if resultado['duplicadas'] > 0:
                    messages.warning(
                        request,
                        f'⚠️ Archivo importado. Se procesaron {resultado["creadas"]} filas nuevas. '
                        f'Se detectaron {resultado["duplicadas"]} filas duplicadas (ya existían en la base de datos). '
                        f'{resultado["errores"]} filas con errores fueron omitidas.'
                    )
                else:
                    messages.success(
                        request,
                        f'✓ Archivo importado correctamente. Se procesaron {resultado["creadas"]} filas nuevas. '
                        f'{resultado["errores"]} filas con errores fueron omitidas.'
                    )
                
                return redirect('eges_import:detalle_batch', batch_id=batch.id)
                
            except Exception as e:
                # Si algo falla, eliminar el batch
                print(f"[EGES] ERROR: {str(e)}")
                import traceback
                traceback.print_exc()
                batch.delete()
                messages.error(request, f'Error al procesar el archivo: {str(e)}')
                return redirect('eges_import:importar')
    else:
        form = ImportarEGESForm()
    
    context = {
        'form': form,
        'titulo_pagina': 'Importar archivo EGES',
    }
    return render(request, 'eges_import/importar.html', context)


@login_required
@user_passes_test(es_superuser, login_url='/admin/')
def detalle_batch(request, batch_id):
    """
    Vista de resumen de un batch importado.
    Muestra todas las métricas calculadas.
    """
    batch = get_object_or_404(ImportBatch, id=batch_id)
    
    # Obtener algunas filas de ejemplo y datos operativos para auditoría
    filas_ejemplo = batch.filas.all()[:20]
    tipos_atencion = batch.filas.exclude(tipo_atencion__isnull=True).exclude(tipo_atencion='').values(
        'tipo_atencion'
    ).annotate(total=Count('id')).order_by('-total', 'tipo_atencion')
    estados_turno = batch.filas.exclude(estado_turno__isnull=True).exclude(estado_turno='').values(
        'estado_turno'
    ).annotate(total=Count('id')).order_by('-total', 'estado_turno')
    total_doppler = batch.filas.filter(sub_modalidad='DOPPLER').count()
    total_guardia = batch.filas.filter(tipo_atencion__iexact='Guardia').count()
    
    context = {
        'batch': batch,
        'filas_ejemplo': filas_ejemplo,
        'tipos_atencion': tipos_atencion,
        'estados_turno': estados_turno,
        'total_doppler': total_doppler,
        'total_guardia': total_guardia,
        'titulo_pagina': f'Resumen Batch #{batch.id}',
    }
    return render(request, 'eges_import/detalle_batch.html', context)


def procesar_excel_eges(archivo, batch):
    """
    Procesa un archivo Excel EGES y crea las filas en la base de datos.
    Optimizado con bulk_create para mejor rendimiento.
    
    Returns:
        dict: {'creadas': int, 'duplicadas': int, 'errores': int}
    """
    from django.db import transaction
    
    print(f"[EGES] Abriendo workbook...")
    wb = openpyxl.load_workbook(archivo, read_only=True, data_only=True)
    ws = wb.active
    
    print(f"[EGES] Leyendo encabezados...")
    # Leer encabezados (primera fila)
    headers = []
    for cell in ws[1]:
        headers.append(str(cell.value).strip() if cell.value else '')
    
    print(f"[EGES] Encabezados encontrados: {len(headers)} columnas")
    print(f"[EGES] Columnas: {headers[:5]}...")  # Primeras 5
    
    # Mapeo de columnas (ajustar según el Excel real)
    # Buscar índices de las columnas que necesitamos
    def get_col_index(nombre_col, alternativas=[]):
        """Busca una columna por nombre o alternativas. Prefiere coincidencia exacta."""
        # 1. Exacta primero (evita que 'Código Practica' gane sobre 'Practica')
        for idx, header in enumerate(headers):
            if header.strip().lower() == nombre_col.strip().lower():
                return idx
        # 2. Substring / alternativas
        for idx, header in enumerate(headers):
            if nombre_col.lower() in header.lower():
                return idx
            for alt in alternativas:
                if alt.lower() in header.lower():
                    return idx
        return None
    
    idx_nro_turno = get_col_index('Nro. Turno', ['Turno', 'Número'])
    idx_fecha = get_col_index('Fecha Turno', ['Fecha'])
    idx_hora = get_col_index('Hora Turno', ['Hora'])
    idx_centro = get_col_index('Centro de Atención', ['Centro', 'Sucursal'])
    idx_hc = get_col_index('Historia Clínica', ['HC', 'H.C.', 'DNI', 'Dni'])
    idx_nombre = get_col_index('Apellido y Nombre', ['Nombre', 'Paciente'])
    idx_servicio = get_col_index('Servicio', ['Prestación', 'Estudio'])
    idx_equipo = get_col_index('Equipo', ['Modalidad'])
    idx_estado = get_col_index('Estado Turno', ['Estado'])
    idx_practica = get_col_index('Practica', ['Práctica', 'Procedimiento'])
    idx_codigo_practica = get_col_index('Código Practica', ['Cod. Practica', 'Cód Práctica', 'Codigo Practica'])
    idx_obra_social = get_col_index('Obra Social', ['Cobertura', 'Prepaga', 'O.S.'])
    idx_codigo_os = get_col_index('Codigo OS', ['Cód. OS', 'Cod OS', 'Código OS', 'Cod.OS'])
    idx_nombre_os = get_col_index('Nombre OS', ['Denominacion OS', 'Denominación OS', 'Desc. OS', 'Descripcion OS', 'Descripción OS'])
    idx_medico = get_col_index('Medico Informante', ['Médico Informante', 'Médico', 'Profesional', 'Informante'])
    idx_medico_actuante = get_col_index('Medico Actuante', ['Médico Actuante', 'Actuante'])

    if idx_medico is None:
        print("[EGES] Aviso: columna de médico informante no encontrada — se importará como vacío. "
              f"Columnas disponibles: {headers}")
    else:
        print(f"[EGES] Columna médico informante en índice {idx_medico}: '{headers[idx_medico]}'")
    if idx_medico_actuante is not None:
        print(f"[EGES] Columna médico actuante en índice {idx_medico_actuante}: '{headers[idx_medico_actuante]}'")

    if idx_practica is None:
        print(f"[EGES] Aviso: columna 'Practica' no encontrada — clasificación usará solo 'Servicio'.")
    else:
        print(f"[EGES] Columna práctica detectada en índice {idx_practica}: '{headers[idx_practica]}'")

    if idx_obra_social is None:
        print(f"[EGES] Aviso: columna 'Obra Social' no encontrada — se importará como vacío.")

    print(f"[EGES] Índices mapeados. Iniciando lectura de filas...")
    
    # Primero, leer TODAS las filas del Excel y crear objetos en memoria
    filas_a_crear = []
    filas_procesadas = 0
    _nombres_os_cache: dict[str, str] = {}  # codigo → nombre, para poblar NombreObraSocial
    filas_error = 0
    # Mapa de sesión → médico: propagamos el médico de filas de insumos
    # a las filas de práctica de la misma sesión (EGES registra al médico
    # solo en los insumos en estudios de seriografía/histerosalpingografía)
    _session_medico: dict[tuple, str] = {}  # (hc, fecha, hora, centro) → medico
    
    # Leer todas las filas del Excel
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        # Saltar filas vacías
        if not any(row):
            continue
        
        filas_procesadas += 1
        
        # Log cada 500 filas
        if filas_procesadas % 500 == 0:
            print(f"[EGES] Leyendo fila {filas_procesadas}...")
        
        try:
            # Extraer valores
            numero_turno = str(row[idx_nro_turno]) if idx_nro_turno is not None and row[idx_nro_turno] else ''
            
            # Fecha
            fecha_turno = None
            if idx_fecha is not None and row[idx_fecha]:
                fecha_val = row[idx_fecha]
                if isinstance(fecha_val, datetime):
                    fecha_turno = fecha_val.date()
                elif isinstance(fecha_val, str):
                    try:
                        fecha_turno = datetime.strptime(fecha_val, '%d/%m/%Y').date()
                    except:
                        try:
                            fecha_turno = datetime.strptime(fecha_val, '%Y-%m-%d').date()
                        except:
                            pass
            
            # Hora
            hora_turno = None
            if idx_hora is not None and row[idx_hora]:
                hora_val = row[idx_hora]
                if isinstance(hora_val, datetime):
                    hora_turno = hora_val.time()
                elif isinstance(hora_val, time):
                    hora_turno = hora_val
                elif isinstance(hora_val, str):
                    try:
                        hora_turno = datetime.strptime(hora_val, '%H:%M:%S').time()
                    except:
                        try:
                            hora_turno = datetime.strptime(hora_val, '%H:%M').time()
                        except:
                            pass
            
            centro = str(row[idx_centro]) if idx_centro is not None and row[idx_centro] else ''
            hc = str(row[idx_hc]) if idx_hc is not None and row[idx_hc] else ''
            nombre = str(row[idx_nombre]) if idx_nombre is not None and row[idx_nombre] else ''
            servicio = str(row[idx_servicio]) if idx_servicio is not None and row[idx_servicio] else ''
            equipo = str(row[idx_equipo]) if idx_equipo is not None and row[idx_equipo] else ''
            estado = str(row[idx_estado]) if idx_estado is not None and row[idx_estado] else ''
            medico_inf = str(row[idx_medico]).strip() if idx_medico is not None and row[idx_medico] else None
            medico_act = str(row[idx_medico_actuante]).strip() if idx_medico_actuante is not None and row[idx_medico_actuante] else None
            # Fallback: si el informante es vacío o "No Especificado", usar el actuante
            _NO_ESP = {'médico no especificado', 'no especificado', 'sin especificar', 'none', ''}
            def _valido(v):
                return bool(v and v.strip().lower() not in _NO_ESP)
            medico = medico_inf if _valido(medico_inf) else (medico_act if _valido(medico_act) else None)
            practica = str(row[idx_practica]).strip() if idx_practica is not None and row[idx_practica] else None
            codigo_practica = str(row[idx_codigo_practica]).strip() if idx_codigo_practica is not None and row[idx_codigo_practica] else None
            obra_social = str(row[idx_obra_social]).strip() if idx_obra_social is not None and row[idx_obra_social] else None
            codigo_os = str(row[idx_codigo_os]).strip() if idx_codigo_os is not None and row[idx_codigo_os] else None
            nombre_os = str(row[idx_nombre_os]).strip() if idx_nombre_os is not None and row[idx_nombre_os] else None

            # Acumular mapeo código → nombre para auto-poblar NombreObraSocial
            if nombre_os:
                _clave_os = codigo_os or obra_social
                if _clave_os:
                    _nombres_os_cache[_clave_os] = nombre_os

            # Crear objeto en memoria (NO guardarlo todavía)
            fila = EgesRow(
                batch=batch,
                numero_turno=numero_turno,
                fecha_turno=fecha_turno,
                hora_turno=hora_turno,
                centro_atencion=centro,
                historia_clinica=hc,
                apellido_nombre=nombre,
                servicio=servicio,
                equipo=equipo,
                estado_turno=estado,
                practica=practica,
                codigo_practica=codigo_practica,
                obra_social=obra_social,
                codigo_obra_social=codigo_os,
                medico_informante=medico,
            )

            # Clasificar modalidad, sub_modalidad e insumo en memoria
            fila.es_insumo = fila.clasificar_insumo()
            if fila.es_insumo:
                # Aunque descartamos el insumo, guardamos su médico para propagar
                # a las filas de práctica de la misma sesión
                if medico:
                    _session_medico.setdefault((hc, fecha_turno, hora_turno, centro), medico)
                continue  # No guardar insumos en la base de datos
            # Si la fila de práctica no tiene médico, usar el de la misma sesión
            if not fila.medico_informante:
                fila.medico_informante = _session_medico.get(
                    (hc, fecha_turno, hora_turno, centro)
                )
            fila.modalidad = fila.clasificar_modalidad()
            if fila.modalidad == 'ECO':
                fila.sub_modalidad = fila.clasificar_sub_modalidad()

            filas_a_crear.append(fila)
            
        except Exception as e:
            # Registrar error pero continuar con las demás filas
            filas_error += 1
            print(f"[EGES] Error en fila {row_idx}: {str(e)}")
            continue
    
    wb.close()
    print(f"[EGES] Workbook cerrado. Leídas {len(filas_a_crear)} filas válidas, {filas_error} errores.")

    # Auto-poblar tabla de nombres de obras sociales si el Excel traía esa columna
    if _nombres_os_cache:
        for _codigo, _nombre in _nombres_os_cache.items():
            NombreObraSocial.objects.update_or_create(
                codigo=_codigo, defaults={'nombre': _nombre}
            )
        print(f"[EGES] {len(_nombres_os_cache)} registros actualizados en NombreObraSocial.")
    
    # Ahora verificar duplicados y crear en batch
    print(f"[EGES] Verificando duplicados y guardando en base de datos...")
    
    # Obtener todas las combinaciones existentes para detectar duplicados
    combinaciones_existentes = set(
        EgesRow.objects.values_list(
            'historia_clinica', 'fecha_turno', 'hora_turno', 'centro_atencion', 'practica'
        )
    )
    
    filas_nuevas = []
    filas_duplicadas = 0
    
    for fila in filas_a_crear:
        combinacion = (
            fila.historia_clinica,
            fila.fecha_turno,
            fila.hora_turno,
            fila.centro_atencion,
            fila.practica
        )
        
        if combinacion not in combinaciones_existentes:
            filas_nuevas.append(fila)
            combinaciones_existentes.add(combinacion)  # Evitar duplicados dentro del mismo batch
        else:
            filas_duplicadas += 1
    
    # Guardar en batch (mucho más rápido)
    filas_creadas = 0
    if filas_nuevas:
        print(f"[EGES] Insertando {len(filas_nuevas)} filas nuevas en la base de datos...")
        with transaction.atomic():
            # Insertar en lotes de 500 para evitar problemas de memoria
            batch_size = 500
            for i in range(0, len(filas_nuevas), batch_size):
                lote = filas_nuevas[i:i + batch_size]
                EgesRow.objects.bulk_create(lote, ignore_conflicts=True)
                filas_creadas += len(lote)
                print(f"[EGES] Insertadas {filas_creadas}/{len(filas_nuevas)} filas...")
    
    print(f"[EGES] Proceso completado: {filas_creadas} nuevas, {filas_duplicadas} duplicadas, {filas_error} errores")
    
    return {
        'creadas': filas_creadas,
        'duplicadas': filas_duplicadas,
        'errores': filas_error
    }


@login_required
@user_passes_test(es_superuser, login_url='/admin/')
def grafico_batch_data(request, batch_id):
    """
    Devuelve datos JSON para gráficos Chart.js.
    Agrupa estudios finalizados por mes y modalidad.
    """
    batch = get_object_or_404(ImportBatch, id=batch_id)
    
    # Filtrar estudios finalizados: no insumos + estado Informado
    estudios_finalizados = batch.filas.filter(
        es_insumo=False,
        estado_turno__iexact='Informado',
        fecha_turno__isnull=False
    )
    
    # Agrupar en Python para evitar TruncMonth + USE_TZ=True en SQLite
    filas = estudios_finalizados.values('fecha_turno', 'modalidad').annotate(total=Count('id'))
    datos_dict = {}
    for item in filas:
        periodo = date_type(item['fecha_turno'].year, item['fecha_turno'].month, 1)
        key = (periodo.strftime('%Y-%m'), item['modalidad'])
        datos_dict[key] = datos_dict.get(key, 0) + item['total']

    meses = sorted(set(k[0] for k in datos_dict))
    modalidades = ['TC', 'RM', 'RX', 'DX', 'MAM', 'ECO', 'SERIE', 'OTROS']
    nombres = dict(EgesRow.MODALIDAD_CHOICES)

    datasets = []
    for modalidad in modalidades:
        data = [datos_dict.get((m, modalidad), 0) for m in meses]
        if sum(data) > 0:
            datasets.append({
                'label': nombres.get(modalidad, modalidad),
                'data': data,
                'borderColor': COLORES_MODALIDAD[modalidad]['border'],
                'backgroundColor': COLORES_MODALIDAD[modalidad]['bg'],
                'borderWidth': 2,
                'tension': 0.3,
            })

    return JsonResponse({'labels': meses, 'datasets': datasets})


@login_required
@user_passes_test(es_superuser, login_url='/admin/')
def grafico_dia_semana_data(request, batch_id):
    """
    Devuelve datos para gráfico de distribución por día de la semana.
    """
    batch = get_object_or_404(ImportBatch, id=batch_id)
    
    # Filtrar estudios finalizados
    estudios = batch.filas.filter(
        es_insumo=False,
        estado_turno__iexact='Informado',
        fecha_turno__isnull=False
    )
    
    # Contar por día de semana (0=Lunes, 6=Domingo)
    dias_semana = ['Lunes', 'Martes', 'Miércoles', 'Jueves', 'Viernes', 'Sábado', 'Domingo']
    conteo_por_dia = [0] * 7
    
    for estudio in estudios:
        dia_numero = estudio.fecha_turno.weekday()  # 0=Lunes, 6=Domingo
        conteo_por_dia[dia_numero] += 1
    
    return JsonResponse({
        'labels': dias_semana,
        'datasets': [{
            'label': 'Estudios Finalizados',
            'data': conteo_por_dia,
            'backgroundColor': [
                'rgba(59, 130, 246, 0.8)',   # Lunes - Azul
                'rgba(16, 185, 129, 0.8)',   # Martes - Verde
                'rgba(245, 158, 11, 0.8)',   # Miércoles - Amarillo
                'rgba(139, 92, 246, 0.8)',   # Jueves - Púrpura
                'rgba(239, 68, 68, 0.8)',    # Viernes - Rojo
                'rgba(107, 114, 128, 0.8)',  # Sábado - Gris
                'rgba(236, 72, 153, 0.8)',   # Domingo - Rosa
            ],
            'borderColor': [
                'rgb(59, 130, 246)',
                'rgb(16, 185, 129)',
                'rgb(245, 158, 11)',
                'rgb(139, 92, 246)',
                'rgb(239, 68, 68)',
                'rgb(107, 114, 128)',
                'rgb(236, 72, 153)',
            ],
            'borderWidth': 2
        }]
    })


@login_required
@user_passes_test(es_superuser, login_url='/admin/')
def grafico_franja_horaria_data(request, batch_id):
    """
    Devuelve datos para gráfico de distribución por franja horaria.
    """
    batch = get_object_or_404(ImportBatch, id=batch_id)
    
    # Filtrar estudios finalizados con hora
    estudios = batch.filas.filter(
        es_insumo=False,
        estado_turno__iexact='Informado',
        fecha_turno__isnull=False,
        hora_turno__isnull=False
    )
    
    # Definir franjas horarias (cada 2 horas)
    franjas = [
        '00-02', '02-04', '04-06', '06-08', '08-10', '10-12',
        '12-14', '14-16', '16-18', '18-20', '20-22', '22-24'
    ]
    conteo_por_franja = [0] * 12
    
    for estudio in estudios:
        hora = estudio.hora_turno.hour
        franja_idx = hora // 2  # Dividir en franjas de 2 horas
        if 0 <= franja_idx < 12:
            conteo_por_franja[franja_idx] += 1
    
    return JsonResponse({
        'labels': franjas,
        'datasets': [{
            'label': 'Estudios Finalizados',
            'data': conteo_por_franja,
            'backgroundColor': 'rgba(59, 130, 246, 0.6)',
            'borderColor': 'rgb(59, 130, 246)',
            'borderWidth': 2,
            'borderRadius': 5,
        }]
    })


@login_required
@user_passes_test(es_superuser, login_url='/admin/')
def dashboard_global_grafico_data(request):
    """
    Devuelve datos consolidados de TODOS los batches para gráficos.
    Incluye filtros por rango de fechas y modalidades.
    """
    # Obtener parámetros de filtro
    fecha_desde = request.GET.get('fecha_desde', '')
    fecha_hasta = request.GET.get('fecha_hasta', '')
    modalidades_seleccionadas = request.GET.getlist('modalidades[]')
    
    # Base query: todos los estudios finalizados
    estudios = EgesRow.objects.filter(
        es_insumo=False,
        estado_turno__iexact='Informado',
        fecha_turno__isnull=False
    )
    
    # Aplicar filtros de fecha
    if fecha_desde:
        try:
            fecha_desde_obj = datetime.strptime(fecha_desde, '%Y-%m-%d').date()
            estudios = estudios.filter(fecha_turno__gte=fecha_desde_obj)
        except:
            pass
    
    if fecha_hasta:
        try:
            fecha_hasta_obj = datetime.strptime(fecha_hasta, '%Y-%m-%d').date()
            estudios = estudios.filter(fecha_turno__lte=fecha_hasta_obj)
        except:
            pass
    
    # Aplicar filtro de modalidades
    if modalidades_seleccionadas:
        estudios = estudios.filter(modalidad__in=modalidades_seleccionadas)
    
    # Agrupar en Python para evitar TruncMonth + USE_TZ=True en SQLite
    filas = estudios.values('fecha_turno', 'modalidad').annotate(total=Count('id'))
    datos_dict = {}
    for item in filas:
        periodo = date_type(item['fecha_turno'].year, item['fecha_turno'].month, 1)
        key = (periodo.strftime('%Y-%m'), item['modalidad'])
        datos_dict[key] = datos_dict.get(key, 0) + item['total']

    meses = sorted(set(k[0] for k in datos_dict))
    modalidades = ['TC', 'RM', 'RX', 'DX', 'MAM', 'ECO', 'SERIE', 'OTROS']

    datasets = []
    for modalidad in modalidades:
        data = [datos_dict.get((m, modalidad), 0) for m in meses]
        if sum(data) > 0:
            datasets.append({
                'label': modalidad,
                'data': data,
                'borderColor': COLORES_MODALIDAD[modalidad]['border'],
                'backgroundColor': COLORES_MODALIDAD[modalidad]['bg'],
                'borderWidth': 2,
                'tension': 0.3,
            })

    return JsonResponse({'labels': meses, 'datasets': datasets})


@login_required
@user_passes_test(es_superuser, login_url='/admin/')
def dashboard_global_dia_semana_data(request):
    """Datos consolidados por día de semana."""
    fecha_desde = request.GET.get('fecha_desde', '')
    fecha_hasta = request.GET.get('fecha_hasta', '')
    modalidades_seleccionadas = request.GET.getlist('modalidades[]')
    
    estudios = _aplicar_filtros_fecha_modalidad(
        _base_estudios_finalizados(),
        request.GET,
    )
    conteos = estudios.annotate(
        dia_semana=ExtractWeekDay('fecha_turno'),
    ).values('dia_semana').annotate(total=Count('id'))
    
    dias_semana = ['Lunes', 'Martes', 'Miércoles', 'Jueves', 'Viernes', 'Sábado', 'Domingo']
    conteo_por_dia = [0] * 7
    for item in conteos:
        dia_numero = (item['dia_semana'] + 5) % 7
        conteo_por_dia[dia_numero] = item['total']
    
    return JsonResponse({
        'labels': dias_semana,
        'datasets': [{
            'label': 'Estudios Finalizados',
            'data': conteo_por_dia,
            'backgroundColor': [
                'rgba(59, 130, 246, 0.8)',
                'rgba(16, 185, 129, 0.8)',
                'rgba(245, 158, 11, 0.8)',
                'rgba(139, 92, 246, 0.8)',
                'rgba(239, 68, 68, 0.8)',
                'rgba(107, 114, 128, 0.8)',
                'rgba(236, 72, 153, 0.8)',
            ],
            'borderColor': [
                'rgb(59, 130, 246)',
                'rgb(16, 185, 129)',
                'rgb(245, 158, 11)',
                'rgb(139, 92, 246)',
                'rgb(239, 68, 68)',
                'rgb(107, 114, 128)',
                'rgb(236, 72, 153)',
            ],
            'borderWidth': 2
        }]
    })


@login_required
@user_passes_test(es_superuser, login_url='/admin/')
def dashboard_global_franja_horaria_data(request):
    """Datos consolidados por franja horaria."""
    fecha_desde = request.GET.get('fecha_desde', '')
    fecha_hasta = request.GET.get('fecha_hasta', '')
    modalidades_seleccionadas = request.GET.getlist('modalidades[]')
    
    estudios = _aplicar_filtros_fecha_modalidad(
        _base_estudios_finalizados().filter(hora_turno__isnull=False),
        request.GET,
    ).order_by().values_list('hora_turno', flat=True)
    
    franjas = [
        '00-02', '02-04', '04-06', '06-08', '08-10', '10-12',
        '12-14', '14-16', '16-18', '18-20', '20-22', '22-24'
    ]
    conteo_por_franja = [0] * 12
    
    for hora_turno in estudios:
        hora = hora_turno.hour
        franja_idx = hora // 2
        if 0 <= franja_idx < 12:
            conteo_por_franja[franja_idx] += 1
    
    return JsonResponse({
        'labels': franjas,
        'datasets': [{
            'label': 'Estudios Finalizados',
            'data': conteo_por_franja,
            'backgroundColor': 'rgba(59, 130, 246, 0.6)',
            'borderColor': 'rgb(59, 130, 246)',
            'borderWidth': 2,
            'borderRadius': 5,
        }]
    })


# ─────────────────────────────────────────────────────────────────────────────
# Helpers de datos movidos a services.py:
# get_base_estudios_finalizados, get_base_rx_sin_informe,
# aplicar_filtros_fecha_modalidad, calcular_kpis, agrupar_periodo, verificar_token
# ─────────────────────────────────────────────────────────────────────────────


# ─────────────────────────────────────────────────────────────────────────────
# Endpoint: KPIs
# ─────────────────────────────────────────────────────────────────────────────

def _vista_kpis_data(request):
    """Lógica interna de KPIs, reutilizable desde superuser y director."""
    base = _base_estudios_finalizados()
    candidatos = EgesRow.objects.filter(es_insumo=False, fecha_turno__isnull=False)
    sin_informe = _base_rx_sin_informe()
    base = _aplicar_filtros_fecha_modalidad(base, request.GET)
    candidatos = _aplicar_filtros_fecha_modalidad(candidatos, request.GET)
    sin_informe = _aplicar_filtros_fecha_modalidad(sin_informe, request.GET)
    return JsonResponse(_calcular_kpis(base, candidatos, sin_informe))


@login_required
@user_passes_test(es_superuser, login_url='/admin/')
def kpis_data(request):
    """Endpoint de KPIs para el dashboard del superuser."""
    return _vista_kpis_data(request)


# ─────────────────────────────────────────────────────────────────────────────
# Endpoint: Análisis temporal con granularidad día / semana / mes
# ─────────────────────────────────────────────────────────────────────────────

def _vista_analisis_temporal(request):
    """Lógica interna compartida."""
    agrupacion = request.GET.get('agrupacion', 'mes')  # dia | semana | mes

    estudios = _base_estudios_finalizados()
    estudios = _aplicar_filtros_fecha_modalidad(estudios, request.GET)

    # Agrupamos en Python para evitar el error de TruncWeek/TruncDate con
    # USE_TZ=True sobre DateField en SQLite.
    filas = estudios.filter(
        fecha_turno__isnull=False
    ).values('fecha_turno', 'modalidad').annotate(total=Count('id'))

    datos_dict = {}
    for item in filas:
        periodo = _agrupar_periodo(item['fecha_turno'], agrupacion)
        key = (periodo, item['modalidad'])
        datos_dict[key] = datos_dict.get(key, 0) + item['total']

    periodos = sorted(set(p for p, _ in datos_dict))

    if agrupacion == 'dia':
        labels = [p.strftime('%d/%m/%Y') for p in periodos]
    elif agrupacion == 'semana':
        labels = [p.strftime('Sem %d/%m/%Y') for p in periodos]
    else:
        labels = [p.strftime('%m/%Y') for p in periodos]

    modalidades = ['TC', 'RM', 'RX', 'DX', 'MAM', 'ECO', 'SERIE', 'OTROS']
    nombres = dict(EgesRow.MODALIDAD_CHOICES)

    datasets = []
    for mod in modalidades:
        data = [datos_dict.get((p, mod), 0) for p in periodos]
        if sum(data) > 0:
            datasets.append({
                'label': nombres.get(mod, mod),
                'data': data,
                'borderColor': COLORES_MODALIDAD[mod]['border'],
                'backgroundColor': COLORES_MODALIDAD[mod]['bg'],
                'borderWidth': 2,
                'tension': 0.3,
            })

    return JsonResponse({'labels': labels, 'datasets': datasets, 'agrupacion': agrupacion})


@login_required
@user_passes_test(es_superuser, login_url='/admin/')
def analisis_temporal_data(request):
    return _vista_analisis_temporal(request)


# ─────────────────────────────────────────────────────────────────────────────
# Endpoint: Distribución porcentual por modalidad (dona)
# ─────────────────────────────────────────────────────────────────────────────

def _vista_distribucion_modalidad(request):
    estudios = _base_estudios_finalizados()
    estudios = _aplicar_filtros_fecha_modalidad(estudios, request.GET)

    datos = estudios.values('modalidad').annotate(total=Count('id')).order_by('-total')
    nombres = dict(EgesRow.MODALIDAD_CHOICES)

    labels = [nombres.get(d['modalidad'], d['modalidad']) for d in datos]
    valores = [d['total'] for d in datos]
    colores = [COLORES_MODALIDAD.get(d['modalidad'], COLORES_MODALIDAD['OTROS'])['border'] for d in datos]

    return JsonResponse({
        'labels': labels,
        'datasets': [{
            'data': valores,
            'backgroundColor': colores,
            'borderColor': '#fff',
            'borderWidth': 2,
        }]
    })


@login_required
@user_passes_test(es_superuser, login_url='/admin/')
def distribucion_modalidad_data(request):
    return _vista_distribucion_modalidad(request)


# ─────────────────────────────────────────────────────────────────────────────
# Endpoint: Sub-modalidades de ECO
# ─────────────────────────────────────────────────────────────────────────────

COLORES_SUB_ECO = [
    'rgba(251, 191, 36, 0.8)',   # ECO_ABDOMINAL
    'rgba(245, 158, 11, 0.8)',   # ECOCARDIO
    'rgba(16, 185, 129, 0.8)',   # DOPPLER
    'rgba(59, 130, 246, 0.8)',   # ECO_PELVIS
    'rgba(139, 92, 246, 0.8)',   # ECO_MAMA
    'rgba(236, 72, 153, 0.8)',   # ECO_OBSTETRICA
    'rgba(249, 115, 22, 0.8)',   # ECO_TIROIDES
    'rgba(20, 184, 166, 0.8)',   # ECO_NEONATAL
    'rgba(107, 114, 128, 0.8)',  # ECO_PARTES_BLANDAS
    'rgba(99, 102, 241, 0.8)',   # Resto
]


def _vista_sub_modalidades_eco(request):
    estudios = _base_estudios_finalizados().filter(modalidad='ECO')
    estudios = _aplicar_filtros_fecha_modalidad(estudios, request.GET)

    datos = estudios.values('sub_modalidad').annotate(total=Count('id')).order_by('-total')
    nombres = dict(EgesRow.SUB_MODALIDAD_ECO_CHOICES)
    nombres[None] = 'ECO genérica'
    nombres[''] = 'ECO genérica'

    labels = [nombres.get(d['sub_modalidad'], d['sub_modalidad'] or 'ECO genérica') for d in datos]
    valores = [d['total'] for d in datos]
    bg = [COLORES_SUB_ECO[i % len(COLORES_SUB_ECO)] for i in range(len(datos))]

    return JsonResponse({
        'labels': labels,
        'datasets': [{
            'label': 'Ecografías',
            'data': valores,
            'backgroundColor': bg,
            'borderColor': '#fff',
            'borderWidth': 1,
        }]
    })


@login_required
@user_passes_test(es_superuser, login_url='/admin/')
def sub_modalidades_eco_data(request):
    return _vista_sub_modalidades_eco(request)


# ─────────────────────────────────────────────────────────────────────────────
# Endpoint: Productividad por médico
# ─────────────────────────────────────────────────────────────────────────────

def _vista_productividad_medico(request):
    agrupacion = request.GET.get('agrupacion', 'mes')  # semana | mes | total
    top_n = int(request.GET.get('top', 10))

    estudios = _base_estudios_finalizados().exclude(
        Q(medico_informante__isnull=True) | Q(medico_informante='')
    )
    estudios = _aplicar_filtros_fecha_modalidad(estudios, request.GET)

    # Ranking total por médico
    ranking = (
        estudios
        .values('medico_informante')
        .annotate(total=Count('id'))
        .order_by('-total')[:top_n]
    )

    labels = [r['medico_informante'] for r in ranking]
    valores = [r['total'] for r in ranking]

    # Desglose por modalidad para cada médico (top N)
    medicos_top = list(labels)
    modalidades = ['TC', 'RM', 'RX', 'DX', 'MAM', 'ECO', 'SERIE', 'OTROS']
    nombres = dict(EgesRow.MODALIDAD_CHOICES)

    desglose = (
        estudios
        .filter(medico_informante__in=medicos_top)
        .values('medico_informante', 'modalidad')
        .annotate(total=Count('id'))
    )
    desglose_dict = {}
    for d in desglose:
        desglose_dict[(d['medico_informante'], d['modalidad'])] = d['total']

    datasets = []
    for mod in modalidades:
        data = [desglose_dict.get((med, mod), 0) for med in medicos_top]
        if sum(data) > 0:
            datasets.append({
                'label': nombres.get(mod, mod),
                'data': data,
                'backgroundColor': COLORES_MODALIDAD[mod]['bg'].replace('0.15', '0.7'),
                'borderColor': COLORES_MODALIDAD[mod]['border'],
                'borderWidth': 1,
            })

    return JsonResponse({
        'labels': labels,
        'datasets': datasets,
        'totales': valores,
    })


@login_required
@user_passes_test(es_superuser, login_url='/admin/')
def productividad_medico_data(request):
    return _vista_productividad_medico(request)


# ─────────────────────────────────────────────────────────────────────────────
# Endpoint: Ranking Obras Sociales
# ─────────────────────────────────────────────────────────────────────────────

def _vista_obras_sociales_data(request):
    top_n = min(int(request.GET.get('top', 15)), 30)
    estudios = _base_estudios_finalizados()
    estudios = _aplicar_filtros_fecha_modalidad(estudios, request.GET)
    datos = (
        estudios
        .exclude(Q(obra_social__isnull=True) | Q(obra_social=''))
        .values('obra_social')
        .annotate(total=Count('id'))
        .order_by('-total')[:top_n]
    )
    # Lookup código → nombre (en memoria, tabla pequeña)
    lookup_nombres = dict(NombreObraSocial.objects.values_list('codigo', 'nombre'))
    labels = [lookup_nombres.get(d['obra_social'], d['obra_social']) for d in datos]
    valores = [d['total'] for d in datos]
    return JsonResponse({
        'labels': labels,
        'datasets': [{
            'label': 'Estudios',
            'data': valores,
            'backgroundColor': 'rgba(59, 130, 246, 0.7)',
            'borderColor': 'rgba(37, 99, 235, 1)',
            'borderWidth': 1,
        }],
    })


@login_required
@user_passes_test(es_superuser, login_url='/admin/')
def obras_sociales_data(request):
    return _vista_obras_sociales_data(request)


def _vista_practicas_data(request):
    """Ranking de prácticas para detectar rápidamente aumentos de demanda."""
    estudios = _aplicar_filtros_fecha_modalidad(_base_estudios_finalizados(), request.GET)
    datos = (estudios.exclude(Q(practica__isnull=True) | Q(practica=''))
             .values('practica').annotate(total=Count('id')).order_by('-total')[:15])
    return JsonResponse({'labels': [d['practica'] for d in datos],
                        'datasets': [{'label': 'Estudios', 'data': [d['total'] for d in datos],
                                      'backgroundColor': 'rgba(22, 69, 105, .75)'}]})


def practicas_data(request):
    return _vista_practicas_data(request)


def _vista_obras_sociales_evolucion(request):
    base = _aplicar_filtros_fecha_modalidad(_base_estudios_finalizados(), request.GET)
    opciones_qs = base.exclude(Q(obra_social__isnull=True) | Q(obra_social='')).values_list('obra_social', flat=True).distinct()
    estudios = base
    obras_seleccionadas = request.GET.getlist('obras_sociales[]') or request.GET.getlist('obras_sociales')
    if obras_seleccionadas:
        estudios = estudios.filter(obra_social__in=obras_seleccionadas)
    filas = estudios.exclude(Q(obra_social__isnull=True) | Q(obra_social='')).values('fecha_turno', 'obra_social')
    conteo = {}
    totales = {}
    for fila in filas:
        mes = fila['fecha_turno'].strftime('%Y-%m')
        os_name = fila['obra_social']
        conteo[(mes, os_name)] = conteo.get((mes, os_name), 0) + 1
        totales[os_name] = totales.get(os_name, 0) + 1
    nombres = [n for n, _ in sorted(totales.items(), key=lambda x: -x[1])[:8]]
    meses = sorted({m for m, _ in conteo})
    lookup = dict(NombreObraSocial.objects.values_list('codigo', 'nombre'))
    opciones = sorted(opciones_qs)
    return JsonResponse({'labels': meses, 'opciones': [{'codigo': n, 'nombre': lookup.get(n, n)} for n in opciones], 'datasets': [
        {'label': lookup.get(nombre, nombre), 'data': [conteo.get((mes, nombre), 0) for mes in meses]}
        for nombre in nombres
    ]})


def obras_sociales_evolucion_data(request):
    return _vista_obras_sociales_evolucion(request)


def _vista_practica_evolucion(request):
    estudios = _aplicar_filtros_fecha_modalidad(_base_estudios_finalizados(), request.GET)
    termino = request.GET.get('practica', 'punc')
    estudios = estudios.filter(practica__icontains=termino)
    filas = estudios.values('fecha_turno').annotate(total=Count('id')).order_by('fecha_turno')
    por_mes = {}
    for fila in filas:
        mes = fila['fecha_turno'].strftime('%Y-%m')
        por_mes[mes] = por_mes.get(mes, 0) + fila['total']
    meses = sorted(por_mes)
    return JsonResponse({'labels': meses, 'datasets': [{'label': f'Prácticas: {termino}', 'data': [por_mes[m] for m in meses], 'borderColor': '#dc2626', 'backgroundColor': 'rgba(220,38,38,.15)', 'fill': True, 'tension': .25}]})


def practica_evolucion_data(request):
    return _vista_practica_evolucion(request)


# ─────────────────────────────────────────────────────────────────────────────
# Endpoint: Comparativa período actual vs anterior
# ─────────────────────────────────────────────────────────────────────────────

def _vista_comparativa_data(request):
    """Compara KPIs del período seleccionado vs el anterior de igual duración."""
    codigos_bloqueos_tc = {'900446', '900441', '900413', '341317'}

    fecha_desde_str = request.GET.get('fecha_desde', '')
    fecha_hasta_str = request.GET.get('fecha_hasta', '')
    hoy = date_type.today()
    rango_manual = bool(fecha_desde_str and fecha_hasta_str)

    if rango_manual:
        try:
            fd = datetime.strptime(fecha_desde_str, '%Y-%m-%d').date()
            fh = datetime.strptime(fecha_hasta_str, '%Y-%m-%d').date()
        except ValueError:
            rango_manual = False

    if not rango_manual:
        filtros_modalidad = request.GET.copy()
        filtros_modalidad.pop('fecha_desde', None)
        filtros_modalidad.pop('fecha_hasta', None)
        ultima_fecha = _aplicar_filtros_fecha_modalidad(
            _base_estudios_finalizados(),
            filtros_modalidad,
        ).aggregate(ultima=Max('fecha_turno'))['ultima']
        if ultima_fecha:
            fd = ultima_fecha.replace(day=1)
            primer_dia_siguiente = (fd + timedelta(days=32)).replace(day=1)
            fh = primer_dia_siguiente - timedelta(days=1)
        else:
            fd = hoy - timedelta(days=30)
            fh = hoy

    if rango_manual:
        duracion = max((fh - fd).days, 1)
        fd_prev = fd - timedelta(days=duracion + 1)
        fh_prev = fd - timedelta(days=1)
    else:
        fh_prev = fd - timedelta(days=1)
        fd_prev = fh_prev.replace(day=1)

    def kpis_para_rango(desde, hasta):
        params = request.GET.copy()
        params['fecha_desde'] = desde.strftime('%Y-%m-%d')
        params['fecha_hasta'] = hasta.strftime('%Y-%m-%d')
        base = _aplicar_filtros_fecha_modalidad(_base_estudios_finalizados(), params)
        candidatos = _aplicar_filtros_fecha_modalidad(
            EgesRow.objects.filter(es_insumo=False, fecha_turno__isnull=False), params
        )
        sin_informe = _aplicar_filtros_fecha_modalidad(_base_rx_sin_informe(), params)
        return _calcular_kpis(base, candidatos, sin_informe)

    actual = kpis_para_rango(fd, fh)
    anterior = kpis_para_rango(fd_prev, fh_prev)

    def contar_practicas_por_nombre(desde, hasta):
        params = request.GET.copy()
        params['fecha_desde'] = desde.strftime('%Y-%m-%d')
        params['fecha_hasta'] = hasta.strftime('%Y-%m-%d')
        filas = _aplicar_filtros_fecha_modalidad(_base_estudios_finalizados(), params)
        return {
            item['practica'] or 'Sin práctica': item['total']
            for item in filas.values('practica').annotate(total=Count('id'))
        }

    def variacion(valor_actual, valor_anterior, decimales=0):
        absoluta = round(valor_actual - valor_anterior, decimales)
        if absoluta == 0:
            estado = 'SIN_CAMBIOS'
        elif valor_anterior == 0 and valor_actual > 0:
            estado = 'NUEVA'
        elif valor_actual == 0 and valor_anterior > 0:
            estado = 'DESAPARECIO'
        elif absoluta > 0:
            estado = 'AUMENTO'
        else:
            estado = 'DISMINUCION'
        porcentaje = None if valor_anterior == 0 else round(
            ((valor_actual - valor_anterior) / valor_anterior) * 100,
            1,
        )
        return {
            'actual': round(valor_actual, decimales),
            'anterior': round(valor_anterior, decimales),
            'absoluta': absoluta,
            'porcentaje': porcentaje,
            'estado': estado,
        }

    practicas_actuales = contar_practicas_por_nombre(fd, fh)
    practicas_anteriores = contar_practicas_por_nombre(fd_prev, fh_prev)
    nombres_practicas = set(practicas_actuales) | set(practicas_anteriores)
    cambios_practicas = []
    for nombre in nombres_practicas:
        cambio = variacion(
            practicas_actuales.get(nombre, 0),
            practicas_anteriores.get(nombre, 0),
        )
        cambios_practicas.append({'practica': nombre, **cambio})

    aumentos = sorted(
        [item for item in cambios_practicas if item['absoluta'] > 0],
        key=lambda item: (-item['absoluta'], item['practica']),
    )[:5]
    disminuciones = sorted(
        [item for item in cambios_practicas if item['absoluta'] < 0],
        key=lambda item: (item['absoluta'], item['practica']),
    )[:5]

    modalidades = request.GET.getlist('modalidades[]') or request.GET.getlist('modalidades')
    incluye_tc = not modalidades or 'TC' in modalidades

    def contar_bloqueos_tc(desde, hasta):
        params = request.GET.copy()
        params['fecha_desde'] = desde.strftime('%Y-%m-%d')
        params['fecha_hasta'] = hasta.strftime('%Y-%m-%d')
        filas = _aplicar_filtros_fecha_modalidad(
            _base_estudios_finalizados().filter(modalidad='TC'),
            params,
        )
        return sum(
            1 for codigo in filas.values_list('codigo_practica', flat=True)
            if str(codigo or '').split('/')[0].strip() in codigos_bloqueos_tc
        )

    hay_datos_anterior = anterior['practicas_realizadas'] > 0
    comparativa = {
        'hay_datos_anterior': hay_datos_anterior,
        'pacientes': variacion(
            actual['pacientes_atendidos'], anterior['pacientes_atendidos'],
        ),
        'practicas': variacion(
            actual['practicas_realizadas'], anterior['practicas_realizadas'],
        ),
        'practicas_por_paciente': variacion(
            actual['practicas_por_paciente'], anterior['practicas_por_paciente'], 2,
        ),
        'practicas_mayores_aumentos': aumentos,
        'practicas_mayores_disminuciones': disminuciones,
        'bloqueos_tc': (
            variacion(contar_bloqueos_tc(fd, fh), contar_bloqueos_tc(fd_prev, fh_prev))
            if incluye_tc else None
        ),
    }

    def delta(a, b):
        try:
            a, b = float(a), float(b)
        except (TypeError, ValueError):
            return None
        if b == 0:
            return None
        return round(((a - b) / b) * 100, 1)

    return JsonResponse({
        'actual': actual,
        'anterior': anterior,
        'periodo_actual': {'desde': fd.strftime('%d/%m/%Y'), 'hasta': fh.strftime('%d/%m/%Y')},
        'periodo_anterior': {'desde': fd_prev.strftime('%d/%m/%Y'), 'hasta': fh_prev.strftime('%d/%m/%Y')},
        'modo_periodo': 'manual' if rango_manual else 'automatico',
        'delta': {
            'total_finalizados': delta(actual['total_finalizados'], anterior['total_finalizados']),
            'promedio_dia': delta(actual['promedio_dia'], anterior['promedio_dia']),
            'tasa_conversion': delta(actual['tasa_conversion'], anterior['tasa_conversion']),
        },
        'comparativa': comparativa,
    })


@login_required
@user_passes_test(es_superuser, login_url='/admin/')
def comparativa_data(request):
    return _vista_comparativa_data(request)


# ─────────────────────────────────────────────────────────────────────────────
# Endpoint: Exportar Excel
# ─────────────────────────────────────────────────────────────────────────────

def _exportar_excel(request):
    """Genera y descarga un .xlsx con los estudios finalizados filtrados."""
    import io
    from django.http import HttpResponse
    from openpyxl.styles import Font, PatternFill, Alignment

    estudios = _base_estudios_finalizados()
    estudios = _aplicar_filtros_fecha_modalidad(estudios, request.GET)
    filas = estudios.order_by('fecha_turno', 'historia_clinica').values(
        'fecha_turno', 'hora_turno', 'historia_clinica', 'practica',
        'codigo_practica', 'obra_social', 'modalidad', 'sub_modalidad',
        'estado_turno', 'medico_informante', 'equipo',
    )

    nombres_modalidad = dict(EgesRow.MODALIDAD_CHOICES)
    nombres_sub = dict(EgesRow.SUB_MODALIDAD_ECO_CHOICES)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Estudios'

    encabezados = [
        'Fecha', 'Hora', 'Historia Clínica', 'Práctica', 'Código Práctica',
        'Obra Social', 'Modalidad', 'Sub-modalidad ECO', 'Estado',
        'Médico Informante', 'Equipo',
    ]
    ws.append(encabezados)

    header_fill = PatternFill(start_color='1E40AF', end_color='1E40AF', fill_type='solid')
    for cell in ws[1]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    for row in filas:
        ws.append([
            row['fecha_turno'].strftime('%d/%m/%Y') if row['fecha_turno'] else '',
            row['hora_turno'].strftime('%H:%M') if row['hora_turno'] else '',
            row['historia_clinica'] or '',
            row['practica'] or '',
            row['codigo_practica'] or '',
            row['obra_social'] or '',
            nombres_modalidad.get(row['modalidad'], row['modalidad'] or ''),
            nombres_sub.get(row['sub_modalidad'], row['sub_modalidad'] or '') if row['sub_modalidad'] else '',
            row['estado_turno'] or '',
            row['medico_informante'] or '',
            row['equipo'] or '',
        ])

    for col in ws.columns:
        max_len = max((len(str(cell.value or '')) for cell in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 45)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    fecha_str = date_type.today().strftime('%Y%m%d')
    filename = f'eges_estudios_{fecha_str}.xlsx'
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


@login_required
@user_passes_test(es_superuser, login_url='/admin/')
def exportar_excel(request):
    return _exportar_excel(request)


# ─────────────────────────────────────────────────────────────────────────────
# Portal del Director (acceso por token UUID, sin login)
# ─────────────────────────────────────────────────────────────────────────────

def portal_director(request, token):
    """
    Portal de análisis de productividad para el director.
    Acceso por token UUID en la URL, sin requerir login.
    """
    director_token = _verificar_token(token)
    if not director_token:
        return HttpResponseForbidden(
            "<h1>Acceso denegado</h1><p>El enlace no es válido o ha sido desactivado.</p>"
        )

    # Métricas globales rápidas para el contexto inicial
    total_estudios = _base_estudios_finalizados().count()
    fecha_min_row = EgesRow.objects.filter(fecha_turno__isnull=False).order_by('fecha_turno').first()
    fecha_max_row = EgesRow.objects.filter(fecha_turno__isnull=False).order_by('-fecha_turno').first()

    return render(request, 'eges_import/portal_director.html', {
        'token': str(director_token.token),
        'etiqueta': director_token.nombre_etiqueta,
        'total_estudios': total_estudios,
        'fecha_min': fecha_min_row.fecha_turno if fecha_min_row else None,
        'fecha_max': fecha_max_row.fecha_turno if fecha_max_row else None,
    })


def portal_director_kpis(request, token):
    if not _verificar_token(token):
        return HttpResponseForbidden()
    return _vista_kpis_data(request)


def portal_director_analisis_temporal(request, token):
    if not _verificar_token(token):
        return HttpResponseForbidden()
    return _vista_analisis_temporal(request)


def portal_director_distribucion(request, token):
    if not _verificar_token(token):
        return HttpResponseForbidden()
    return _vista_distribucion_modalidad(request)


def portal_director_eco(request, token):
    if not _verificar_token(token):
        return HttpResponseForbidden()
    return _vista_sub_modalidades_eco(request)


def portal_director_medicos(request, token):
    if not _verificar_token(token):
        return HttpResponseForbidden()
    return _vista_productividad_medico(request)


def portal_director_obras_sociales(request, token):
    if not _verificar_token(token):
        return HttpResponseForbidden()
    return _vista_obras_sociales_data(request)


def portal_director_practicas(request, token):
    if not _verificar_token(token):
        return HttpResponseForbidden()
    return _vista_practicas_data(request)


def portal_director_obras_sociales_evolucion(request, token):
    if not _verificar_token(token):
        return HttpResponseForbidden()
    return _vista_obras_sociales_evolucion(request)


def portal_director_practica_evolucion(request, token):
    if not _verificar_token(token):
        return HttpResponseForbidden()
    return _vista_practica_evolucion(request)


def portal_director_comparativa(request, token):
    if not _verificar_token(token):
        return HttpResponseForbidden()
    return _vista_comparativa_data(request)


def portal_director_exportar_excel(request, token):
    if not _verificar_token(token):
        return HttpResponseForbidden()
    return _exportar_excel(request)


def portal_director_franja_horaria(request, token):
    """Franja horaria para el portal del director (reutiliza lógica del dashboard global)."""
    if not _verificar_token(token):
        return HttpResponseForbidden()

    estudios = _base_estudios_finalizados().filter(hora_turno__isnull=False)
    estudios = _aplicar_filtros_fecha_modalidad(estudios, request.GET)

    franjas = [
        '00-02', '02-04', '04-06', '06-08', '08-10', '10-12',
        '12-14', '14-16', '16-18', '18-20', '20-22', '22-24'
    ]
    conteo = [0] * 12
    for e in estudios:
        idx = e.hora_turno.hour // 2
        if 0 <= idx < 12:
            conteo[idx] += 1

    return JsonResponse({
        'labels': franjas,
        'datasets': [{
            'label': 'Estudios Finalizados',
            'data': conteo,
            'backgroundColor': 'rgba(59, 130, 246, 0.6)',
            'borderColor': 'rgb(59, 130, 246)',
            'borderWidth': 2,
            'borderRadius': 5,
        }]
    })


# ─────────────────────────────────────────────────────────────────────────────
# Exportar PDF
# ─────────────────────────────────────────────────────────────────────────────

def _exportar_pdf(request):
    """
    Genera un informe PDF ejecutivo.
    Usa matplotlib para gráficos (alta calidad) + ReportLab Platypus para layout.
    Estructura: portada+KPIs | evolución+torta | ECO+OS | médicos+franja.
    """
    import io
    import matplotlib
    matplotlib.use('Agg')
    import matplotlib.pyplot as plt
    import numpy as np
    from django.http import HttpResponse
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.units import cm
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
    from reportlab.platypus import (
        SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
        HRFlowable, PageBreak, Image, KeepTogether,
    )
    PAGE_W, PAGE_H = landscape(A4)
    USABLE_W = PAGE_W - 3 * cm  # ~784pt con márgenes de 1.5cm c/lado

    # ── Paleta ────────────────────────────────────────────────────────────────
    AZUL          = colors.HexColor('#1e40af')
    AZUL_CLARO    = colors.HexColor('#dbeafe')
    GRIS_CABECERA = colors.HexColor('#1e293b')
    GRIS_FILA     = colors.HexColor('#f8fafc')
    VERDE         = colors.HexColor('#16a34a')
    ROJO          = colors.HexColor('#dc2626')
    MOD_HEX = {
        'TC': '#3b82f6', 'RM': '#9333ea', 'RX': '#22c55e',
        'DX': '#f97316', 'MAM': '#ec4899', 'ECO': '#eab308',
        'SERIE': '#14b8a6', 'OTROS': '#6b7280',
    }

    # ── Estilos ReportLab ─────────────────────────────────────────────────────
    styles = getSampleStyleSheet()
    s_titulo    = ParagraphStyle('sTitulo',  parent=styles['Title'],  fontSize=22,
                                 textColor=AZUL, spaceAfter=4, fontName='Helvetica-Bold', leading=28)
    s_subtitulo = ParagraphStyle('sSubti',   parent=styles['Normal'], fontSize=10,
                                 textColor=colors.HexColor('#64748b'), spaceAfter=4, fontName='Helvetica')
    s_seccion   = ParagraphStyle('sSeccion', parent=styles['Normal'], fontSize=11,
                                 textColor=AZUL, spaceBefore=12, spaceAfter=6, fontName='Helvetica-Bold')
    s_pie       = ParagraphStyle('sPie',     parent=styles['Normal'], fontSize=7,
                                 textColor=colors.HexColor('#9ca3af'), alignment=TA_CENTER)
    s_delta     = ParagraphStyle('sDelta',   parent=styles['Normal'], fontSize=9,
                                 fontName='Helvetica-Bold', alignment=TA_CENTER)

    _TBL_BASE = [
        ('BACKGROUND', (0, 0), (-1, 0), GRIS_CABECERA),
        ('TEXTCOLOR',  (0, 0), (-1, 0), colors.white),
        ('FONTNAME',   (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE',   (0, 0), (-1, -1), 9),
        ('ALIGN',      (1, 0), (-1, -1), 'CENTER'),
        ('ALIGN',      (0, 0), (0, -1), 'LEFT'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, GRIS_FILA]),
        ('BOX',        (0, 0), (-1, -1), 0.5, colors.HexColor('#e2e8f0')),
        ('INNERGRID',  (0, 0), (-1, -1), 0.3, colors.HexColor('#e2e8f0')),
        ('TOPPADDING',    (0, 0), (-1, -1), 5),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
        ('LEFTPADDING',   (0, 0), (-1, -1), 8),
    ]

    def _tbl_style(extra=None):
        return TableStyle(_TBL_BASE + (extra or []))

    # ── Helpers matplotlib → ReportLab Image ─────────────────────────────────
    DPI = 150

    def _fig_to_image(fig, width_pt, height_pt):
        buf = io.BytesIO()
        fig.savefig(buf, format='png', bbox_inches='tight', dpi=DPI,
                    facecolor='white', edgecolor='none')
        plt.close(fig)
        buf.seek(0)
        return Image(buf, width=width_pt, height=height_pt)

    def _empty_image(width_pt, height_pt, msg='Sin datos'):
        fig, ax = plt.subplots(figsize=(width_pt / 72, height_pt / 72))
        ax.text(0.5, 0.5, msg, ha='center', va='center',
                transform=ax.transAxes, color='#9ca3af', fontsize=11)
        ax.axis('off')
        return _fig_to_image(fig, width_pt, height_pt)

    def _chart_barras_agrupadas(periodos_labels, datasets, mod_labels, mod_colors,
                                 width_pt, height_pt):
        """Barras agrupadas por modalidad a lo largo del tiempo."""
        if not periodos_labels or not datasets:
            return _empty_image(width_pt, height_pt)
        n_p = len(periodos_labels)
        n_m = len(datasets)
        fig, ax = plt.subplots(figsize=(width_pt / 72, height_pt / 72))
        x = np.arange(n_p)
        bar_w = min(0.8 / n_m, 0.15)
        for i, (data, label, color) in enumerate(zip(datasets, mod_labels, mod_colors)):
            offset = (i - n_m / 2 + 0.5) * bar_w
            ax.bar(x + offset, data, bar_w, label=label, color=color, alpha=0.88,
                   edgecolor='white', linewidth=0.3)
        ax.set_xticks(x)
        ax.set_xticklabels(periodos_labels, rotation=30 if n_p > 6 else 0,
                           ha='right', fontsize=7)
        ax.yaxis.set_tick_params(labelsize=7)
        ax.set_axisbelow(True)
        ax.yaxis.grid(True, color='#e5e7eb', linewidth=0.5)
        ax.spines[['top', 'right']].set_visible(False)
        ax.legend(fontsize=7, loc='upper left', framealpha=0.7,
                  ncol=min(n_m, 4), handlelength=1, columnspacing=0.8)
        fig.tight_layout(pad=0.5)
        return _fig_to_image(fig, width_pt, height_pt)

    def _chart_donut(labels, valores, hex_colors, width_pt, height_pt):
        """Gráfico donut de distribución por modalidad."""
        if not valores or sum(valores) == 0:
            return _empty_image(width_pt, height_pt)
        fig, ax = plt.subplots(figsize=(width_pt / 72, height_pt / 72))
        wedges, _, autotexts = ax.pie(
            valores, colors=hex_colors,
            autopct='%1.1f%%', pctdistance=0.78,
            wedgeprops=dict(width=0.52, edgecolor='white', linewidth=1.5),
            startangle=90,
        )
        for at in autotexts:
            at.set_fontsize(7)
            at.set_color('white')
            at.set_fontweight('bold')
        ax.legend(wedges, [f'{l} ({v})' for l, v in zip(labels, valores)],
                  loc='lower center', bbox_to_anchor=(0.5, -0.28),
                  ncol=2, fontsize=7, framealpha=0.7, handlelength=1)
        fig.tight_layout(pad=0.2)
        return _fig_to_image(fig, width_pt, height_pt)

    def _chart_barras_h(labels, valores, hex_color, width_pt, height_pt):
        """Barras horizontales con etiquetas de valor."""
        if not valores or sum(valores) == 0:
            return _empty_image(width_pt, height_pt)
        fig, ax = plt.subplots(figsize=(width_pt / 72, height_pt / 72))
        y_pos = np.arange(len(labels))
        bars = ax.barh(y_pos, valores, color=hex_color, alpha=0.85,
                       edgecolor='white', linewidth=0.3, height=0.6)
        ax.set_yticks(y_pos)
        ax.set_yticklabels(labels, fontsize=7)
        ax.xaxis.set_tick_params(labelsize=7)
        ax.set_axisbelow(True)
        ax.xaxis.grid(True, color='#e5e7eb', linewidth=0.5)
        ax.spines[['top', 'right']].set_visible(False)
        max_v = max(valores) if valores else 1
        for bar in bars:
            w = bar.get_width()
            if w > 0:
                ax.text(w + max_v * 0.01, bar.get_y() + bar.get_height() / 2,
                        str(int(w)), va='center', ha='left', fontsize=7, color='#374151')
        fig.tight_layout(pad=0.5)
        return _fig_to_image(fig, width_pt, height_pt)

    # ── Datos ─────────────────────────────────────────────────────────────────
    base = _base_estudios_finalizados()
    candidatos = EgesRow.objects.filter(es_insumo=False, fecha_turno__isnull=False)
    base_f = _aplicar_filtros_fecha_modalidad(base, request.GET)
    candidatos_f = _aplicar_filtros_fecha_modalidad(candidatos, request.GET)
    sin_informe_f = _aplicar_filtros_fecha_modalidad(_base_rx_sin_informe(), request.GET)
    kpis = _calcular_kpis(base_f, candidatos_f, sin_informe_f)

    # Rango de fechas
    fecha_desde_str = request.GET.get('fecha_desde', '')
    fecha_hasta_str = request.GET.get('fecha_hasta', '')
    hoy = date_type.today()
    if fecha_desde_str and fecha_hasta_str:
        try:
            fd = datetime.strptime(fecha_desde_str, '%Y-%m-%d').date()
            fh = datetime.strptime(fecha_hasta_str, '%Y-%m-%d').date()
        except ValueError:
            fd, fh = hoy - timedelta(days=30), hoy
    else:
        fechas_qs = base_f.values_list('fecha_turno', flat=True)
        if fechas_qs.exists():
            fd, fh = min(fechas_qs), max(fechas_qs)
        else:
            fd, fh = hoy - timedelta(days=30), hoy

    duracion = max((fh - fd).days, 1)
    fd_prev = fd - timedelta(days=duracion + 1)
    fh_prev = fd - timedelta(days=1)

    from django.http import QueryDict

    def _kpis_rango(desde, hasta):
        p = QueryDict(f'fecha_desde={desde.strftime("%Y-%m-%d")}&fecha_hasta={hasta.strftime("%Y-%m-%d")}')
        return _calcular_kpis(
            _aplicar_filtros_fecha_modalidad(_base_estudios_finalizados(), p),
            _aplicar_filtros_fecha_modalidad(
                EgesRow.objects.filter(es_insumo=False, fecha_turno__isnull=False), p),
            _aplicar_filtros_fecha_modalidad(_base_rx_sin_informe(), p),
        )

    kpis_prev = _kpis_rango(fd_prev, fh_prev)

    def _delta(a, b):
        try:
            a, b = float(a), float(b)
            if b == 0:
                return '—', colors.HexColor('#6b7280')
            d = (a - b) / b * 100
            return f'{"▲" if d >= 0 else "▼"} {abs(d):.1f}%', (VERDE if d >= 0 else ROJO)
        except (TypeError, ValueError):
            return '—', colors.HexColor('#6b7280')

    # Distribución por modalidad
    dist_datos  = base_f.values('modalidad').annotate(total=Count('id')).order_by('-total')
    dist_labels = [dict(EgesRow.MODALIDAD_CHOICES).get(d['modalidad'], d['modalidad']) for d in dist_datos]
    dist_valores = [d['total'] for d in dist_datos]
    dist_mods   = [d['modalidad'] for d in dist_datos]

    # Sub-modalidades ECO
    sub_datos   = (base_f.filter(modalidad='ECO', sub_modalidad__isnull=False)
                   .values('sub_modalidad').annotate(total=Count('id')).order_by('-total'))
    sub_labels  = [dict(EgesRow.SUB_MODALIDAD_ECO_CHOICES).get(d['sub_modalidad'], d['sub_modalidad']) for d in sub_datos]
    sub_valores = [d['total'] for d in sub_datos]

    # Médicos top 10
    med_datos = (
        base_f.exclude(Q(medico_informante__isnull=True) | Q(medico_informante=''))
        .values('medico_informante').annotate(total=Count('id')).order_by('-total')[:10]
    )

    # Franja horaria
    franjas = ['00-02', '02-04', '04-06', '06-08', '08-10', '10-12',
               '12-14', '14-16', '16-18', '18-20', '20-22', '22-24']
    conteo_franja = [0] * 12
    for e in base_f.filter(hora_turno__isnull=False):
        idx_f = e.hora_turno.hour // 2
        if 0 <= idx_f < 12:
            conteo_franja[idx_f] += 1

    # Obras sociales top 10
    lookup_nombres = dict(NombreObraSocial.objects.values_list('codigo', 'nombre'))
    os_datos   = (base_f.exclude(Q(obra_social__isnull=True) | Q(obra_social=''))
                  .values('obra_social').annotate(total=Count('id')).order_by('-total')[:10])
    os_labels  = [lookup_nombres.get(d['obra_social'], d['obra_social']) for d in os_datos]
    os_valores = [d['total'] for d in os_datos]

    # Evolución mensual
    filas_temp = (base_f.filter(fecha_turno__isnull=False)
                  .values('fecha_turno', 'modalidad').annotate(total=Count('id')))
    temp_dict = {}
    for item in filas_temp:
        p_key = date_type(item['fecha_turno'].year, item['fecha_turno'].month, 1)
        key = (p_key, item['modalidad'])
        temp_dict[key] = temp_dict.get(key, 0) + item['total']
    periodos = sorted(set(p for p, _ in temp_dict))
    temp_labels_list = [p.strftime('%m/%Y') for p in periodos]
    mods_presentes = [m for m in ['TC', 'RM', 'RX', 'DX', 'MAM', 'ECO', 'SERIE', 'OTROS']
                      if sum(temp_dict.get((p, m), 0) for p in periodos) > 0]

    # ── Generar imágenes matplotlib ─────────────────────────────────────────
    # Gráfico de evolución temporal
    if periodos and mods_presentes:
        evol_data = [[temp_dict.get((p, m), 0) for p in periodos] for m in mods_presentes]
        evol_leyendas = [dict(EgesRow.MODALIDAD_CHOICES).get(m, m) for m in mods_presentes]
        img_evol = _chart_barras_agrupadas(
            temp_labels_list, evol_data, evol_leyendas,
            [MOD_HEX[m] for m in mods_presentes],
            width_pt=USABLE_W * 0.62, height_pt=210,
        )
    else:
        img_evol = _empty_image(USABLE_W * 0.62, 210)

    img_donut = _chart_donut(
        dist_labels, dist_valores,
        [MOD_HEX.get(m, '#6b7280') for m in dist_mods],
        width_pt=USABLE_W * 0.36, height_pt=210,
    )
    HALF_W  = USABLE_W / 2 - 4
    ECO_H   = 230
    img_eco = _chart_barras_h(sub_labels[::-1], sub_valores[::-1],
                               '#eab308', HALF_W, ECO_H)
    img_os  = _chart_barras_h([l[:35] for l in os_labels[::-1]], os_valores[::-1],
                               '#3b82f6', HALF_W, ECO_H)
    img_franja = _chart_barras_h(franjas[::-1], conteo_franja[::-1],
                                  '#3b82f6', USABLE_W, 180)

    # ── Construcción del documento ────────────────────────────────────────────
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        leftMargin=1.5 * cm, rightMargin=1.5 * cm,
        topMargin=1.5 * cm, bottomMargin=1.5 * cm,
    )
    story = []

    # ── PÁGINA 1: Portada + KPIs + Comparativa ───────────────────────────────
    story.append(Paragraph('Informe de Estudios Realizados', s_titulo))
    story.append(Paragraph(
        f'Período: {fd.strftime("%d/%m/%Y")} — {fh.strftime("%d/%m/%Y")}  ·  '
        f'Generado el {hoy.strftime("%d/%m/%Y")}  ·  Sanatorio Colegiales',
        s_subtitulo,
    ))
    story.append(HRFlowable(width='100%', thickness=1, color=AZUL, spaceAfter=10))

    story.append(Paragraph('Indicadores Clave', s_seccion))
    kpi_w = USABLE_W / 4
    kpi_data = [
        ['Estudios Finalizados', 'Promedio / Día Hábil', 'Tasa de Conversión', 'Estudios Pendientes'],
        [str(kpis['total_finalizados']), str(kpis['promedio_dia']),
         f"{kpis['tasa_conversion']}%", str(kpis['total_pendientes'])],
        ['estado = Informado', 'días hábiles del período',
         'finalizados / candidatos', 'no informados aún'],
    ]
    kpi_ts = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), GRIS_CABECERA),
        ('TEXTCOLOR',  (0, 0), (-1, 0), colors.white),
        ('FONTNAME',   (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE',   (0, 0), (-1, 0), 8),
        ('ALIGN',      (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN',     (0, 0), (-1, -1), 'MIDDLE'),
        ('FONTNAME',   (0, 1), (-1, 1), 'Helvetica-Bold'),
        ('FONTSIZE',   (0, 1), (-1, 1), 24),
        ('TEXTCOLOR',  (0, 1), (0, 1), AZUL),
        ('TEXTCOLOR',  (1, 1), (1, 1), colors.HexColor('#7c3aed')),
        ('TEXTCOLOR',  (2, 1), (2, 1), VERDE),
        ('TEXTCOLOR',  (3, 1), (3, 1), colors.HexColor('#d97706')),
        ('BACKGROUND', (0, 1), (-1, 1), AZUL_CLARO),
        ('BACKGROUND', (0, 2), (-1, 2), GRIS_FILA),
        ('FONTNAME',   (0, 2), (-1, 2), 'Helvetica'),
        ('FONTSIZE',   (0, 2), (-1, 2), 8),
        ('TEXTCOLOR',  (0, 2), (-1, 2), colors.HexColor('#6b7280')),
        ('BOX',        (0, 0), (-1, -1), 0.5, colors.HexColor('#e2e8f0')),
        ('INNERGRID',  (0, 0), (-1, -1), 0.3, colors.HexColor('#e2e8f0')),
        ('TOPPADDING',    (0, 0), (-1, 0), 8),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('TOPPADDING',    (0, 1), (-1, 1), 14),
        ('BOTTOMPADDING', (0, 1), (-1, 1), 14),
        ('TOPPADDING',    (0, 2), (-1, 2), 6),
        ('BOTTOMPADDING', (0, 2), (-1, 2), 6),
    ])
    story.append(Table(kpi_data, colWidths=[kpi_w] * 4,
                       rowHeights=[24, 46, 22], style=kpi_ts))

    story.append(Spacer(1, 14))
    story.append(Paragraph('Comparativa con Período Anterior', s_seccion))
    story.append(Paragraph(
        f'Período actual: {fd.strftime("%d/%m/%Y")} – {fh.strftime("%d/%m/%Y")}  ·  '
        f'Período anterior: {fd_prev.strftime("%d/%m/%Y")} – {fh_prev.strftime("%d/%m/%Y")}',
        s_subtitulo,
    ))
    metricas_comp = [
        ('Estudios Finalizados',   kpis['total_finalizados'],  kpis_prev['total_finalizados']),
        ('Promedio / Día Hábil',   kpis['promedio_dia'],        kpis_prev['promedio_dia']),
        ('Tasa de Conversión (%)', kpis['tasa_conversion'],     kpis_prev['tasa_conversion']),
        ('Estudios Pendientes',    kpis['total_pendientes'],    kpis_prev['total_pendientes']),
    ]
    comp_rows = [['Métrica', 'Período Actual', 'Período Anterior', 'Variación']]
    for met, act, ant in metricas_comp:
        ds, dc = _delta(act, ant)
        comp_rows.append([met, str(act), str(ant),
                          Paragraph(f'<font color="{dc.hexval()}">{ds}</font>', s_delta)])
    story.append(Table(comp_rows,
                       colWidths=[USABLE_W * f for f in [0.40, 0.20, 0.20, 0.20]],
                       style=_tbl_style()))
    story.append(PageBreak())

    # ── PÁGINA 2: Evolución + Distribución ───────────────────────────────────
    story.append(Paragraph('Evolución Temporal de Estudios Finalizados', s_seccion))
    graf_evol = Table(
        [[img_evol, img_donut]],
        colWidths=[USABLE_W * 0.62 + 8, USABLE_W * 0.36 + 8],
    )
    graf_evol.setStyle(TableStyle([
        ('VALIGN',        (0, 0), (-1, -1), 'TOP'),
        ('LEFTPADDING',   (0, 0), (-1, -1), 0),
        ('RIGHTPADDING',  (0, 0), (-1, -1), 0),
        ('TOPPADDING',    (0, 0), (-1, -1), 0),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 0),
    ]))
    story.append(graf_evol)
    story.append(Spacer(1, 10))
    story.append(Paragraph('Distribución por Modalidad', s_seccion))
    total_global = sum(dist_valores) or 1
    dist_rows = [['Modalidad', 'Estudios', '% del Total']]
    for lbl, val in zip(dist_labels, dist_valores):
        dist_rows.append([lbl, str(val), f'{val / total_global * 100:.1f}%'])
    dist_rows.append(['TOTAL', str(sum(dist_valores)), '100%'])
    story.append(Table(dist_rows,
                       colWidths=[USABLE_W * f for f in [0.50, 0.25, 0.25]],
                       style=_tbl_style([
                           ('FONTNAME',   (0, -1), (-1, -1), 'Helvetica-Bold'),
                           ('BACKGROUND', (0, -1), (-1, -1), AZUL_CLARO),
                       ])))
    story.append(PageBreak())

    # ── PÁGINA 3: ECO Sub-modalidades + Top OS ───────────────────────────────
    story.append(Paragraph('Sub-modalidades de Ecografía  ·  Top 10 Obras Sociales', s_seccion))
    graf_p3 = Table(
        [[img_eco, img_os]],
        colWidths=[HALF_W + 6, HALF_W + 6],
    )
    graf_p3.setStyle(TableStyle([
        ('VALIGN',        (0, 0), (-1, -1), 'TOP'),
        ('LEFTPADDING',   (0, 0), (-1, -1), 0),
        ('RIGHTPADDING',  (0, 0), (-1, -1), 4),
        ('TOPPADDING',    (0, 0), (-1, -1), 0),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 0),
    ]))
    story.append(graf_p3)
    if sub_labels:
        story.append(Spacer(1, 8))
        story.append(Paragraph('Detalle Sub-modalidades', s_seccion))
        total_eco = sum(sub_valores) or 1
        sub_rows = [['Sub-modalidad', 'Estudios', '% sobre total ECO']]
        for lbl, val in zip(sub_labels, sub_valores):
            sub_rows.append([lbl, str(val), f'{val / total_eco * 100:.1f}%'])
        story.append(Table(sub_rows,
                           colWidths=[USABLE_W * f for f in [0.55, 0.20, 0.25]],
                           style=_tbl_style()))
    story.append(PageBreak())

    # ── PÁGINA 4: Médicos + Franja Horaria ───────────────────────────────────
    story.append(Paragraph('Productividad por Médico Informante — Top 10', s_seccion))
    total_med = sum(d['total'] for d in med_datos) or 1
    med_rows = [['#', 'Médico Informante', 'Estudios', '% del Total']]
    for i, d in enumerate(med_datos, 1):
        med_rows.append([
            str(i),
            d['medico_informante'] or 'Sin especificar',
            str(d['total']),
            f'{d["total"] / total_med * 100:.1f}%',
        ])
    story.append(Table(med_rows,
                       colWidths=[USABLE_W * f for f in [0.06, 0.55, 0.20, 0.19]],
                       style=_tbl_style()))
    story.append(Spacer(1, 14))
    story.append(KeepTogether([
        Paragraph('Distribución por Franja Horaria', s_seccion),
        img_franja,
    ]))
    story.append(Spacer(1, 10))
    story.append(HRFlowable(width='100%', thickness=0.5,
                             color=colors.HexColor('#e2e8f0'), spaceAfter=4))
    story.append(Paragraph(
        f'Informe generado el {hoy.strftime("%d/%m/%Y")} · '
        f'Sistema de Gestión de Servicios · '
        f'Basado en estudios con estado "Informado"',
        s_pie,
    ))

    doc.build(story)
    buffer.seek(0)

    fecha_str = hoy.strftime('%Y%m%d')
    filename = f'informe_eges_{fecha_str}.pdf'
    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


@login_required
@user_passes_test(es_superuser, login_url='/admin/')
def exportar_pdf(request):
    return _exportar_pdf(request)


def portal_director_exportar_pdf(request, token):
    if not _verificar_token(token):
        return HttpResponseForbidden()
    return _exportar_pdf(request)
