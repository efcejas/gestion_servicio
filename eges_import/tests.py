from io import BytesIO
from datetime import date, time, timedelta

from django.contrib.auth import get_user_model
from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import TestCase
from openpyxl import Workbook

from .forms import ImportarEGESForm
from .models import EgesRow, ImportBatch
from .services import calcular_sha256_archivo, procesar_excel_eges


def _archivo_eges_xlsx():
    wb = Workbook()
    ws = wb.active
    ws.title = 'Atendidos'
    ws.append([
        'Dni',
        'Numero  Afiliado',
        'Centro',
        'Fecha Turno',
        'Hora Desde',
        'Hora Hasta',
        'Tipo Atencion',
        'Medico Informante',
        'Medico Actuante',
        'Paciente',
        'Codigo OS',
        'Obra Social',
        'Equipo',
        'Estado Turno',
        'Código Practica',
        'Practica',
        'Cantidad',
        'Servicio',
        'Estado Informe',
        'Tipo de turno',
        'Tipo de paciente',
        'Region del informe',
    ])
    ws.append([
        '37933118',
        '2052113',
        'Sanatorio Colegiales',
        '01/05/2026',
        '09:45',
        '10:00',
        'Ambulatorio',
        'Médico No Especificado',
        'TOMALA BERNABE FABIAN HOLGER',
        'NIZ MICAELA YANINA',
        '1967',
        'SANCOR SALUD (C)',
        'P - Ecografo - AFFINITI 30',
        'Informado',
        '180112/0',
        'ECOGRAFIA COMPLETA DE ABDOMEN',
        1,
        'Ecografia',
        '-',
        'Normal',
        'Adulto',
        '-',
    ])
    ws.append([
        '37933118',
        '2052113',
        'Sanatorio Colegiales',
        '01/05/2026',
        '09:45',
        '10:00',
        'Ambulatorio',
        'Médico No Especificado',
        'TOMALA BERNABE FABIAN HOLGER',
        'NIZ MICAELA YANINA',
        '1967',
        'SANCOR SALUD (C)',
        'P - Ecografo - AFFINITI 30',
        'Informado',
        '3005554/0',
        'CAMISOLIN',
        1,
        'Ecografia',
        '-',
        'Normal',
        'Adulto',
        '-',
    ])
    contenido = BytesIO()
    wb.save(contenido)
    contenido.seek(0)
    return contenido.getvalue()


def _archivo_atendidos_xlsx():
    wb = Workbook()
    ws = wb.active
    ws.title = 'Turnos'
    ws.append([
        'Fecha', 'Paciente', 'Historia Clinica', 'Protocolo', 'Centro',
        'Servicio', 'Práctica', 'Médico', 'Obra Social', 'Estado',
        'Informante', 'Equipo', 'Duración', 'Contraste', 'Anestesia',
        'Tipo Turno', 'Técnico', 'Tipo Ingreso', 'Aplicación Origen',
    ])
    filas = [
        ('01/05/2026 09:00', 'PROTO-1', 'ACTUANTE UNO', 'INFORMANTE UNO', 'TECNICO UNO'),
        ('01/05/2026 10:00', 'PROTO-2', 'ACTUANTE DOS', 'Médico No Especificado', 'TECNICO DOS'),
        ('01/05/2026 11:00', 'PROTO-3', 'Médico No Especificado', ', Médico No Especificado', 'TECNICO TRES'),
    ]
    for numero, (fecha, protocolo, actuante, informante, tecnico) in enumerate(filas, start=1):
        ws.append([
            fecha, f'PACIENTE {numero}', str(1000 + numero), protocolo,
            'Sanatorio Colegiales', 'Ecografia',
            'ECOGRAFIA COMPLETA DE ABDOMEN', actuante, 'COBERTURA',
            'Informado', informante, 'Ecografo 1', 15, 'No', 'No',
            'Normal', tecnico, 'Ambulatorio', 'Portal',
        ])
    contenido = BytesIO()
    wb.save(contenido)
    contenido.seek(0)
    return contenido.getvalue()


def _archivo_practicas_mismo_protocolo_xlsx(practicas):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Turnos'
    ws.append([
        'Dni', 'Historia Clinica', 'Protocolo', 'Centro', 'Fecha Turno',
        'Hora Desde', 'Servicio', 'Práctica', 'Código Practica',
        'Estado Turno', 'Cantidad',
    ])
    for practica in practicas:
        ws.append([
            '12345678', '12345678', 'PROTO-1', 'Sanatorio Colegiales',
            '01/05/2026', '09:00', 'Ecografia', practica, '900137/0',
            'Informado', 1,
        ])
    contenido = BytesIO()
    wb.save(contenido)
    contenido.seek(0)
    return contenido.getvalue()


class ImportacionEgesAuditoriaPacsTest(TestCase):
    def setUp(self):
        self.user = get_user_model().objects.create_user(username='admin', password='x')

    def test_importa_campos_operativos_y_descarta_insumos(self):
        archivo = SimpleUploadedFile(
            'Turnos-Mayo-ECO.xlsx',
            _archivo_eges_xlsx(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        batch = ImportBatch.objects.create(usuario=self.user, archivo_nombre=archivo.name)

        resultado = procesar_excel_eges(archivo, batch)

        self.assertEqual(resultado['creadas'], 1)
        self.assertEqual(EgesRow.objects.count(), 1)
        fila = EgesRow.objects.get()
        self.assertEqual(fila.dni_paciente, '37933118')
        self.assertEqual(fila.historia_clinica, '37933118')
        self.assertEqual(fila.numero_afiliado, '2052113')
        self.assertEqual(fila.apellido_nombre, 'NIZ MICAELA YANINA')
        self.assertEqual(fila.tipo_atencion, 'Ambulatorio')
        self.assertEqual(fila.hora_turno.strftime('%H:%M'), '09:45')
        self.assertEqual(fila.hora_hasta.strftime('%H:%M'), '10:00')
        self.assertEqual(fila.codigo_practica, '180112/0')
        self.assertEqual(fila.practica, 'ECOGRAFIA COMPLETA DE ABDOMEN')
        self.assertEqual(fila.medico_informante, 'TOMALA BERNABE FABIAN HOLGER')
        self.assertEqual(fila.medico_actuante, 'TOMALA BERNABE FABIAN HOLGER')
        self.assertEqual(fila.estado_informe, '-')
        self.assertEqual(fila.tipo_turno, 'Normal')
        self.assertEqual(fila.tipo_paciente, 'Adulto')
        self.assertEqual(fila.region_informe, '-')

    def test_aguja_de_biopsia_se_clasifica_como_insumo(self):
        fila = EgesRow(practica='AGUJA DE BIOPSIA', servicio='Ecografia')

        self.assertTrue(fila.clasificar_insumo())

    def test_puncion_biopsia_guiada_se_conserva_como_practica(self):
        fila = EgesRow(
            practica='PUNCION BIOPSIA BAJO CONTROL ECOGRAFICO',
            servicio='Ecografia',
        )

        self.assertFalse(fila.clasificar_insumo())

    def test_formulario_acepta_xlsx_eges_real(self):
        archivo = SimpleUploadedFile(
            'Turnos-Mayo-ECO.xlsx',
            _archivo_eges_xlsx(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )

        form = ImportarEGESForm(files={'archivo': archivo})

        self.assertTrue(form.is_valid(), form.errors)

    def test_reimportar_archivo_identico_no_crea_otro_lote_ni_duplica_filas(self):
        self.user.is_superuser = True
        self.user.save(update_fields=['is_superuser'])
        self.client.force_login(self.user)
        contenido = _archivo_eges_xlsx()

        for nombre in ('Turnos-Mayo-ECO.xlsx', 'copia-renombrada.xlsx'):
            response = self.client.post(
                '/eges/importar/',
                {'archivo': SimpleUploadedFile(
                    nombre,
                    contenido,
                    content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                )},
            )
            self.assertEqual(response.status_code, 302)

        self.assertEqual(ImportBatch.objects.count(), 1)
        self.assertEqual(EgesRow.objects.count(), 1)
        self.assertEqual(
            ImportBatch.objects.get().archivo_sha256,
            calcular_sha256_archivo(BytesIO(contenido)),
        )

    def test_reimportar_datos_de_lote_historico_sin_huella_descarta_lote_vacio(self):
        self.user.is_superuser = True
        self.user.save(update_fields=['is_superuser'])
        self.client.force_login(self.user)
        contenido = _archivo_eges_xlsx()
        historico = ImportBatch.objects.create(
            usuario=self.user,
            archivo_nombre='historico-sin-huella.xlsx',
        )
        procesar_excel_eges(BytesIO(contenido), historico)
        historico.calcular_metricas()

        response = self.client.post(
            '/eges/importar/',
            {'archivo': SimpleUploadedFile(
                'reintento-historico.xlsx',
                contenido,
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            )},
        )

        self.assertRedirects(response, '/eges/')
        self.assertEqual(ImportBatch.objects.count(), 1)
        self.assertEqual(EgesRow.objects.count(), 1)
        self.assertTrue(ImportBatch.objects.filter(pk=historico.pk).exists())

    def test_importa_formato_atendidos_y_resuelve_profesional_desde_tres_columnas(self):
        contenido = _archivo_atendidos_xlsx()
        archivo = SimpleUploadedFile(
            'Atendidos-mayo-ECO.xlsx',
            contenido,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        form = ImportarEGESForm(files={'archivo': archivo})
        self.assertTrue(form.is_valid(), form.errors)
        batch = ImportBatch.objects.create(usuario=self.user, archivo_nombre=archivo.name)

        resultado = procesar_excel_eges(archivo, batch)

        self.assertEqual(resultado['creadas'], 3)
        filas = {fila.protocolo: fila for fila in EgesRow.objects.all()}
        self.assertEqual(filas['PROTO-1'].medico_informante, 'INFORMANTE UNO')
        self.assertEqual(filas['PROTO-2'].medico_informante, 'ACTUANTE DOS')
        self.assertEqual(filas['PROTO-3'].medico_informante, 'TECNICO TRES')
        self.assertEqual(filas['PROTO-1'].medico_actuante, 'ACTUANTE UNO')
        self.assertEqual(filas['PROTO-1'].tecnico, 'TECNICO UNO')
        self.assertEqual(str(filas['PROTO-1'].duracion_minutos), '15.00')
        self.assertEqual(filas['PROTO-1'].aplicacion_origen, 'Portal')
        self.assertEqual(filas['PROTO-1'].fecha_turno.isoformat(), '2026-05-01')
        self.assertEqual(filas['PROTO-1'].hora_turno.strftime('%H:%M'), '09:00')

    def test_mismo_protocolo_con_tres_practicas_conserva_las_tres(self):
        contenido = _archivo_practicas_mismo_protocolo_xlsx([
            'ECOGRAFIA ABDOMINAL',
            'ECOGRAFIA RENAL',
            'ECOGRAFIA VESICAL',
        ])
        batch = ImportBatch.objects.create(usuario=self.user, archivo_nombre='practicas.xlsx')

        resultado = procesar_excel_eges(BytesIO(contenido), batch)

        self.assertEqual(resultado['creadas'], 3)
        self.assertEqual(EgesRow.objects.count(), 3)
        self.assertEqual(
            set(EgesRow.objects.values_list('practica', flat=True)),
            {'ECOGRAFIA ABDOMINAL', 'ECOGRAFIA RENAL', 'ECOGRAFIA VESICAL'},
        )

    def test_misma_practica_con_misma_identidad_real_se_deduplica(self):
        contenido = _archivo_practicas_mismo_protocolo_xlsx(['ECOGRAFIA ABDOMINAL'])
        primero = ImportBatch.objects.create(usuario=self.user, archivo_nombre='primero.xlsx')
        segundo = ImportBatch.objects.create(usuario=self.user, archivo_nombre='segundo.xlsx')

        resultado_primero = procesar_excel_eges(BytesIO(contenido), primero)
        resultado_segundo = procesar_excel_eges(BytesIO(contenido), segundo)

        self.assertEqual(resultado_primero['creadas'], 1)
        self.assertEqual(resultado_segundo['creadas'], 0)
        self.assertEqual(resultado_segundo['duplicadas'], 1)
        self.assertEqual(EgesRow.objects.count(), 1)

    def _crear_estudio_dashboard(self, batch, protocolo, fecha, modalidad, dni=None, practica=None):
        practica = practica or ('TRANSITO DE INTESTINO' if modalidad == 'SERIE' else f'PRACTICA {protocolo}')
        fila = EgesRow.objects.create(
            batch=batch,
            protocolo=protocolo,
            historia_clinica=protocolo,
            dni_paciente=dni,
            fecha_turno=fecha,
            centro_atencion='Sanatorio Colegiales',
            practica=practica,
            estado_turno='Informado',
            modalidad=modalidad,
            es_insumo=False,
            medico_informante='PROFESIONAL PRUEBA',
        )
        fila.modalidad = modalidad
        fila.save(update_fields=['modalidad'])
        return fila

    def test_kpis_separan_pacientes_practicas_y_excluyen_insumos(self):
        self.user.is_superuser = True
        self.user.save(update_fields=['is_superuser'])
        self.client.force_login(self.user)
        batch = ImportBatch.objects.create(usuario=self.user, archivo_nombre='kpis.xlsx')

        for numero, practica in enumerate(('ECO ABDOMEN', 'ECO RENAL', 'ECO VESICAL'), start=1):
            self._crear_estudio_dashboard(
                batch, f'ECO-A-{numero}', date(2026, 5, 10), 'ECO',
                dni='12.345.678', practica=practica,
            )
        self._crear_estudio_dashboard(
            batch, 'ECO-B-1', date(2026, 5, 10), 'ECO',
            dni='23456789', practica='ECO MAMARIA',
        )
        EgesRow.objects.create(
            batch=batch,
            historia_clinica='INSUMO-1',
            dni_paciente='99999999',
            fecha_turno=date(2026, 5, 10),
            centro_atencion='Sanatorio Colegiales',
            practica='GUANTES',
            estado_turno='Informado',
            modalidad='ECO',
            es_insumo=True,
        )

        data = self.client.get('/eges/datos/kpis/', {
            'fecha_desde': '2026-05-01',
            'fecha_hasta': '2026-05-31',
            'modalidades[]': 'ECO',
        }).json()

        self.assertEqual(data['pacientes_atendidos'], 2)
        self.assertEqual(data['practicas_realizadas'], 4)
        self.assertEqual(data['practicas_por_paciente'], 2.0)

    def test_kpis_respetan_modalidad_seleccionada(self):
        self.user.is_superuser = True
        self.user.save(update_fields=['is_superuser'])
        self.client.force_login(self.user)
        batch = ImportBatch.objects.create(usuario=self.user, archivo_nombre='modalidades.xlsx')
        self._crear_estudio_dashboard(batch, 'ECO-1', date(2026, 5, 10), 'ECO', dni='11111111')
        self._crear_estudio_dashboard(batch, 'TC-1', date(2026, 5, 10), 'TC', dni='22222222')

        eco = self.client.get('/eges/datos/kpis/', {'modalidades[]': 'ECO'}).json()
        tc = self.client.get('/eges/datos/kpis/', {'modalidades[]': 'TC'}).json()

        self.assertEqual(eco['pacientes_atendidos'], 1)
        self.assertEqual(eco['practicas_realizadas'], 1)
        self.assertEqual(tc['pacientes_atendidos'], 1)
        self.assertEqual(tc['practicas_realizadas'], 1)

    def _crear_comparativa(self, fecha, modalidad, practicas, pacientes=1, prefijo='P'):
        filas = []
        indice = 0
        for practica, cantidad in practicas.items():
            for _ in range(cantidad):
                dni = f'{prefijo}{indice % pacientes:03d}'
                filas.append(EgesRow(
                    batch=self.batch_comparativa,
                    historia_clinica=dni,
                    dni_paciente=dni,
                    fecha_turno=fecha,
                    centro_atencion='Centro test',
                    practica=practica,
                    estado_turno='Informado',
                    modalidad=modalidad,
                    es_insumo=False,
                ))
                indice += 1
        EgesRow.objects.bulk_create(filas)

    def _preparar_comparativa(self):
        self.user.is_superuser = True
        self.user.save(update_fields=['is_superuser'])
        self.client.force_login(self.user)
        self.batch_comparativa = ImportBatch.objects.create(
            usuario=self.user,
            archivo_nombre='comparativa-tests.xlsx',
        )
        return {
            'fecha_desde': '2026-05-01',
            'fecha_hasta': '2026-05-10',
        }

    def test_comparativa_calcula_pacientes_practicas_y_promedio(self):
        filtros = self._preparar_comparativa()
        self._crear_comparativa(date(2026, 5, 1), 'ECO', {'ECO ABDOMEN': 160}, pacientes=100, prefijo='A')
        self._crear_comparativa(date(2026, 4, 21), 'ECO', {'ECO ABDOMEN': 100}, pacientes=80, prefijo='B')

        data = self.client.get('/eges/datos/comparativa/', filtros).json()['comparativa']

        self.assertEqual(data['pacientes']['actual'], 100)
        self.assertEqual(data['pacientes']['anterior'], 80)
        self.assertEqual(data['pacientes']['absoluta'], 20)
        self.assertEqual(data['practicas']['absoluta'], 60)
        self.assertEqual(data['practicas_por_paciente']['actual'], 1.6)
        self.assertEqual(data['practicas_por_paciente']['anterior'], 1.25)

    def test_comparativa_ordena_aumentos_y_disminuciones_por_diferencia(self):
        filtros = self._preparar_comparativa()
        self._crear_comparativa(date(2026, 5, 1), 'ECO', {'ECO ABDOMEN': 120, 'DOPPLER': 60})
        self._crear_comparativa(date(2026, 4, 21), 'ECO', {'ECO ABDOMEN': 100, 'DOPPLER': 80})
        data = self.client.get('/eges/datos/comparativa/', filtros).json()['comparativa']

        aumentos = {item['practica']: item for item in data['practicas_mayores_aumentos']}
        disminuciones = {item['practica']: item for item in data['practicas_mayores_disminuciones']}
        self.assertEqual(aumentos['ECO ABDOMEN']['absoluta'], 20)
        self.assertEqual(disminuciones['DOPPLER']['absoluta'], -20)

    def test_comparativa_identifica_practica_nueva_sin_division_por_cero(self):
        filtros = self._preparar_comparativa()
        self._crear_comparativa(date(2026, 5, 1), 'ECO', {'ECO NUEVA': 15})
        data = self.client.get('/eges/datos/comparativa/', filtros).json()['comparativa']

        nueva = next(item for item in data['practicas_mayores_aumentos'] if item['practica'] == 'ECO NUEVA')
        self.assertEqual(nueva['estado'], 'NUEVA')
        self.assertIsNone(nueva['porcentaje'])

    def test_comparativa_identifica_practica_desaparecida(self):
        filtros = self._preparar_comparativa()
        self._crear_comparativa(date(2026, 4, 21), 'ECO', {'ECO DESAPARECE': 20})
        data = self.client.get('/eges/datos/comparativa/', filtros).json()['comparativa']

        desaparecida = next(item for item in data['practicas_mayores_disminuciones'] if item['practica'] == 'ECO DESAPARECE')
        self.assertEqual(desaparecida['estado'], 'DESAPARECIO')
        self.assertEqual(desaparecida['absoluta'], -20)

    def test_comparativa_filtro_eco_no_incluye_tc(self):
        filtros = self._preparar_comparativa()
        filtros['modalidades[]'] = 'ECO'
        self._crear_comparativa(date(2026, 5, 1), 'ECO', {'ECO ABDOMEN': 2})
        self._crear_comparativa(date(2026, 5, 1), 'TC', {'TOMOGRAFIA': 20})
        data = self.client.get('/eges/datos/comparativa/', filtros).json()['comparativa']

        self.assertEqual(data['practicas']['actual'], 2)
        self.assertIsNone(data['bloqueos_tc'])

    def test_comparativa_informa_periodo_anterior_sin_actividad(self):
        filtros = self._preparar_comparativa()
        self._crear_comparativa(date(2026, 5, 1), 'ECO', {'ECO ABDOMEN': 2})
        data = self.client.get('/eges/datos/comparativa/', filtros).json()['comparativa']

        self.assertFalse(data['hay_datos_anterior'])
        self.assertEqual(data['practicas']['estado'], 'NUEVA')

    def test_comparativa_automatica_usa_ultimo_mes_disponible(self):
        self._preparar_comparativa()
        self._crear_comparativa(date(2026, 7, 15), 'TC', {'TOMOGRAFIA': 2})
        self._crear_comparativa(date(2026, 6, 15), 'TC', {'TOMOGRAFIA': 1})

        data = self.client.get('/eges/datos/comparativa/', {
            'modalidades[]': 'TC',
        }).json()

        self.assertEqual(data['modo_periodo'], 'automatico')
        self.assertEqual(data['periodo_actual'], {'desde': '01/07/2026', 'hasta': '31/07/2026'})
        self.assertEqual(data['periodo_anterior'], {'desde': '01/06/2026', 'hasta': '30/06/2026'})

    def test_comparativa_automatica_ancla_cada_modalidad_a_su_ultimo_mes(self):
        self._preparar_comparativa()
        self._crear_comparativa(date(2026, 5, 15), 'ECO', {'ECO ABDOMEN': 2})
        self._crear_comparativa(date(2026, 4, 15), 'ECO', {'ECO ABDOMEN': 1})
        self._crear_comparativa(date(2026, 7, 15), 'TC', {'TOMOGRAFIA': 2})
        self._crear_comparativa(date(2026, 6, 15), 'TC', {'TOMOGRAFIA': 1})

        eco = self.client.get('/eges/datos/comparativa/', {'modalidades[]': 'ECO'}).json()
        tc = self.client.get('/eges/datos/comparativa/', {'modalidades[]': 'TC'}).json()

        self.assertEqual(eco['periodo_actual']['desde'], '01/05/2026')
        self.assertEqual(eco['periodo_actual']['hasta'], '31/05/2026')
        self.assertEqual(tc['periodo_actual']['desde'], '01/07/2026')
        self.assertEqual(tc['periodo_actual']['hasta'], '31/07/2026')

    def test_comparativa_manual_conserva_rango_e_intervalo_anterior(self):
        self._preparar_comparativa()
        self._crear_comparativa(date(2026, 5, 10), 'ECO', {'ECO ABDOMEN': 2})
        self._crear_comparativa(date(2026, 4, 29), 'ECO', {'ECO ABDOMEN': 1})

        data = self.client.get('/eges/datos/comparativa/', {
            'fecha_desde': '2026-05-10',
            'fecha_hasta': '2026-05-20',
            'modalidades[]': 'ECO',
        }).json()

        self.assertEqual(data['modo_periodo'], 'manual')
        self.assertEqual(data['periodo_actual'], {'desde': '10/05/2026', 'hasta': '20/05/2026'})
        self.assertEqual(data['periodo_anterior'], {'desde': '29/04/2026', 'hasta': '09/05/2026'})

    def test_evolucion_no_agrega_meses_sin_datos(self):
        self._preparar_comparativa()
        self._crear_comparativa(date(2026, 7, 15), 'TC', {'TOMOGRAFIA': 2})

        data = self.client.get('/eges/datos/analisis-temporal/', {
            'modalidades[]': 'TC',
            'agrupacion': 'mes',
        }).json()

        self.assertEqual(data['labels'], ['07/2026'])

    def test_evolucion_obras_sociales_mantiene_rango_y_meses_sin_actividad(self):
        self._preparar_comparativa()
        for fecha, obra_social in (
            (date(2026, 1, 15), 'COBERTURA A'),
            (date(2026, 3, 15), 'COBERTURA A'),
            (date(2026, 2, 15), 'COBERTURA B'),
        ):
            EgesRow.objects.create(
                batch=self.batch_comparativa,
                historia_clinica=f'{obra_social}-{fecha}',
                fecha_turno=fecha,
                centro_atencion='Centro test',
                practica='ECO TEST',
                obra_social=obra_social,
                estado_turno='Informado',
                modalidad='ECO',
                es_insumo=False,
            )

        data = self.client.get('/eges/datos/obras-sociales/evolucion/', {
            'obras_sociales[]': ['COBERTURA A', 'COBERTURA B'],
        }).json()

        self.assertEqual(data['labels'], ['2026-01', '2026-02', '2026-03'])
        self.assertEqual(data['datasets'][0]['data'], [1, 0, 1])
        self.assertEqual(data['datasets'][1]['data'], [0, 1, 0])

    def test_dia_y_franja_conservan_conteos_tras_agregacion(self):
        self._preparar_comparativa()
        for indice, (fecha, hora) in enumerate([
            (date(2026, 5, 4), time(9, 0)),   # lunes, 08-10
            (date(2026, 5, 5), time(11, 0)),  # martes, 10-12
            (date(2026, 5, 4), time(9, 30)),  # lunes, 08-10
        ]):
            EgesRow.objects.create(
                batch=self.batch_comparativa,
                historia_clinica=f'DIA-{indice}',
                dni_paciente=f'DIA-{indice}',
                fecha_turno=fecha,
                hora_turno=hora,
                centro_atencion='Centro test',
                practica='ECO TEST',
                estado_turno='Informado',
                modalidad='ECO',
                es_insumo=False,
            )

        filtros = {'fecha_desde': '2026-05-01', 'fecha_hasta': '2026-05-31', 'modalidades[]': 'ECO'}
        dias = self.client.get('/eges/estadisticas/grafico-dia-semana/', filtros).json()
        franjas = self.client.get('/eges/estadisticas/grafico-franja-horaria/', filtros).json()

        self.assertEqual(dias['datasets'][0]['data'], [2, 1, 0, 0, 0, 0, 0])
        self.assertEqual(franjas['datasets'][0]['data'][4:6], [2, 1])

    def test_comparativa_respeta_modalidad_seleccionada(self):
        self.user.is_superuser = True
        self.user.save(update_fields=['is_superuser'])
        self.client.force_login(self.user)
        batch = ImportBatch.objects.create(usuario=self.user, archivo_nombre='comparativa.xlsx')
        actual_desde = date.today() - timedelta(days=9)
        actual_hasta = date.today()
        anterior_desde = actual_desde - timedelta(days=10)

        self._crear_estudio_dashboard(batch, 'TC-ACTUAL-1', actual_desde, 'TC')
        self._crear_estudio_dashboard(batch, 'TC-ACTUAL-2', actual_hasta, 'TC')
        self._crear_estudio_dashboard(batch, 'TC-ANTERIOR-1', anterior_desde, 'TC')
        for numero in range(5):
            self._crear_estudio_dashboard(batch, f'RM-ACTUAL-{numero}', actual_desde, 'RM')
            self._crear_estudio_dashboard(batch, f'RM-ANTERIOR-{numero}', anterior_desde, 'RM')

        response = self.client.get('/eges/datos/comparativa/', {
            'fecha_desde': actual_desde.isoformat(),
            'fecha_hasta': actual_hasta.isoformat(),
            'modalidades[]': 'TC',
        })

        self.assertEqual(response.status_code, 200)
        data = response.json()
        self.assertEqual(data['actual']['total_finalizados'], 2)
        self.assertEqual(data['anterior']['total_finalizados'], 1)
        self.assertEqual(data['delta']['total_finalizados'], 100.0)

    def test_tendencia_y_productividad_incluyen_serie(self):
        self.user.is_superuser = True
        self.user.save(update_fields=['is_superuser'])
        self.client.force_login(self.user)
        batch = ImportBatch.objects.create(usuario=self.user, archivo_nombre='serie.xlsx')
        fila_serie = self._crear_estudio_dashboard(batch, 'SERIE-1', date.today(), 'SERIE')
        fila_serie.refresh_from_db()

        self.assertEqual(EgesRow.objects.filter(
            modalidad='SERIE', estado_turno__iexact='Informado', es_insumo=False,
        ).count(), 1, (fila_serie.modalidad, fila_serie.estado_turno, fila_serie.es_insumo))

        temporal = self.client.get('/eges/datos/analisis-temporal/').json()
        productividad = self.client.get('/eges/datos/productividad-medico/').json()

        self.assertIn('Seriografía', [dataset['label'] for dataset in temporal['datasets']])
        self.assertIn('Seriografía', [dataset['label'] for dataset in productividad['datasets']])

    def test_dashboard_advierte_lote_con_posible_tope_eges(self):
        self.user.is_superuser = True
        self.user.save(update_fields=['is_superuser'])
        self.client.force_login(self.user)
        ImportBatch.objects.create(
            usuario=self.user,
            archivo_nombre='Atendidos-mayo-ECO.xls',
            total_filas=1000,
        )

        response = self.client.get('/eges/estadisticas/')

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context['lotes_posible_tope'], 1)
        self.assertContains(response, 'podrían haber alcanzado el límite')
