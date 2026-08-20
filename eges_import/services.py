"""
services.py — Lógica de datos del módulo eges_import.

Separa las queries y transformaciones de datos del ciclo request/response.
Estas funciones no reciben ni retornan objetos HTTP: son reutilizables
desde vistas, management commands y tests sin levantar un servidor.
"""
from datetime import date as date_type, datetime, time, timedelta
from decimal import Decimal, InvalidOperation
import hashlib
import os
import unicodedata

import openpyxl
import xlrd
from django.db import transaction
from django.db.models import Count, Min, Max, Q

from .models import DirectorToken, EgesRow, NombreObraSocial


NO_ESPECIFICADO = {'medico no especificado', 'médico no especificado', 'no especificado', 'sin especificar', 'none', ''}


def calcular_sha256_archivo(archivo):
    """Calcula una huella estable y restablece el cursor del archivo subido."""
    digest = hashlib.sha256()
    archivo.seek(0)
    for bloque in iter(lambda: archivo.read(1024 * 1024), b''):
        digest.update(bloque)
    archivo.seek(0)
    return digest.hexdigest()


def _normalizar_header(valor):
    texto = str(valor or '').strip().lower()
    texto = ''.join(
        c for c in unicodedata.normalize('NFKD', texto)
        if not unicodedata.combining(c)
    )
    return ' '.join(texto.split())


def _valor_valido(valor):
    if not valor:
        return False
    texto = _normalizar_header(str(valor).strip(' ,;.-'))
    return texto not in {_normalizar_header(item) for item in NO_ESPECIFICADO}


def _texto(valor):
    if valor is None:
        return ''
    if isinstance(valor, float) and valor.is_integer():
        return str(int(valor))
    return str(valor).strip()


def _decimal(valor, default='1'):
    if valor in (None, ''):
        return Decimal(default)
    try:
        return Decimal(str(valor).replace(',', '.'))
    except (InvalidOperation, ValueError):
        return Decimal(default)


def _normalizar_dni(valor):
    """Normaliza DNI para contar pacientes sin crear identificadores alternativos."""
    return ''.join(caracter for caracter in str(valor or '') if caracter.isdigit())


def _fecha(valor, datemode=None):
    if not valor:
        return None
    if isinstance(valor, datetime):
        return valor.date()
    if isinstance(valor, date_type):
        return valor
    if isinstance(valor, (int, float)) and datemode is not None:
        return xlrd.xldate.xldate_as_datetime(valor, datemode).date()
    if isinstance(valor, str):
        for fmt in ('%d/%m/%Y %H:%M:%S', '%d/%m/%Y %H:%M', '%d/%m/%Y', '%Y-%m-%d'):
            try:
                return datetime.strptime(valor.strip(), fmt).date()
            except ValueError:
                continue
    return None


def _hora(valor, datemode=None):
    if not valor:
        return None
    if isinstance(valor, datetime):
        return valor.time()
    if isinstance(valor, time):
        return valor
    if isinstance(valor, (int, float)) and datemode is not None:
        return xlrd.xldate.xldate_as_datetime(valor, datemode).time()
    if isinstance(valor, str):
        for fmt in ('%d/%m/%Y %H:%M:%S', '%d/%m/%Y %H:%M', '%H:%M:%S', '%H:%M'):
            try:
                return datetime.strptime(valor.strip(), fmt).time()
            except ValueError:
                continue
    return None


def _abrir_workbook(archivo):
    nombre = getattr(archivo, 'name', '') or ''
    extension = os.path.splitext(nombre.lower())[1]
    if hasattr(archivo, 'seek'):
        archivo.seek(0)

    if extension == '.xls':
        contenido = archivo.read()
        if hasattr(archivo, 'seek'):
            archivo.seek(0)
        wb = xlrd.open_workbook(file_contents=contenido)
        sh = wb.sheet_by_index(0)
        headers = [_texto(sh.cell_value(0, c)) for c in range(sh.ncols)]
        rows = (
            [sh.cell_value(r, c) for c in range(sh.ncols)]
            for r in range(1, sh.nrows)
        )
        return headers, rows, wb.datemode

    wb = openpyxl.load_workbook(archivo, read_only=True, data_only=True)
    ws = wb.active
    headers = [_texto(cell.value) for cell in ws[1]]
    rows = ws.iter_rows(min_row=2, values_only=True)
    return headers, rows, None


def leer_encabezados_eges(archivo):
    headers, _rows, _datemode = _abrir_workbook(archivo)
    return headers


def _indice(headers, principal, alternativas=None):
    alternativas = alternativas or []
    buscados = [_normalizar_header(principal), *[_normalizar_header(a) for a in alternativas]]
    normalizados = [_normalizar_header(h) for h in headers]

    for buscado in buscados:
        for idx, header in enumerate(normalizados):
            if header == buscado:
                return idx
    for buscado in buscados:
        for idx, header in enumerate(normalizados):
            if buscado and buscado in header:
                return idx
    return None


def _celda(row, idx):
    if idx is None or idx >= len(row):
        return None
    return row[idx]


def procesar_excel_eges(archivo, batch):
    """
    Procesa archivos EGES .xls/.xlsx y conserva campos útiles para auditoría PACS.

    Mantiene la regla histórica de no persistir insumos y de deduplicar por
    HC + fecha + hora + centro + práctica. El protocolo no identifica una
    práctica única y no participa de la deduplicación.
    """
    print("[EGES] Abriendo workbook...")
    headers, rows, datemode = _abrir_workbook(archivo)

    print(f"[EGES] Encabezados encontrados: {len(headers)} columnas")
    print(f"[EGES] Columnas: {headers[:8]}...")

    idx_nro_turno = _indice(headers, 'Nro. Turno', ['Turno', 'Número'])
    idx_protocolo = _indice(headers, 'Protocolo')
    idx_dni = _indice(headers, 'Dni', ['DNI', 'Documento'])
    idx_numero_afiliado = _indice(headers, 'Numero Afiliado', ['Número Afiliado', 'Nro Afiliado'])
    idx_fecha = _indice(headers, 'Fecha Turno', ['Fecha'])
    idx_hora = _indice(headers, 'Hora Turno', ['Hora Desde', 'Hora'])
    if idx_hora is None:
        # El reporte "Atendidos" combina fecha y hora en una sola columna.
        idx_hora = idx_fecha
    idx_hora_hasta = _indice(headers, 'Hora Hasta')
    idx_centro = _indice(headers, 'Centro de Atención', ['Centro', 'Sucursal'])
    idx_tipo_atencion = _indice(headers, 'Tipo Atencion', ['Tipo Atención', 'Tipo Ingreso'])
    idx_hc = _indice(headers, 'Historia Clínica', ['HC', 'H.C.'])
    idx_nombre = _indice(headers, 'Apellido y Nombre', ['Paciente', 'Nombre'])
    idx_servicio = _indice(headers, 'Servicio', ['Prestación', 'Estudio'])
    idx_equipo = _indice(headers, 'Equipo', ['Modalidad'])
    idx_estado = _indice(headers, 'Estado Turno', ['Estado'])
    idx_estado_informe = _indice(headers, 'Estado Informe')
    idx_tipo_turno = _indice(headers, 'Tipo de turno', ['Tipo Turno'])
    idx_tipo_paciente = _indice(headers, 'Tipo de paciente', ['Tipo Paciente'])
    idx_region_informe = _indice(headers, 'Region del informe', ['Región del informe'])
    idx_practica = _indice(headers, 'Practica', ['Práctica', 'Procedimiento'])
    idx_codigo_practica = _indice(headers, 'Código Practica', ['Cod. Practica', 'Cód Práctica', 'Codigo Practica'])
    idx_cantidad = _indice(headers, 'Cantidad')
    idx_obra_social = _indice(headers, 'Obra Social', ['Cobertura', 'Prepaga', 'O.S.'])
    idx_codigo_os = _indice(headers, 'Codigo OS', ['Cód. OS', 'Cod OS', 'Código OS', 'Cod.OS'])
    idx_nombre_os = _indice(headers, 'Nombre OS', ['Denominacion OS', 'Denominación OS', 'Desc. OS', 'Descripcion OS', 'Descripción OS'])
    idx_medico = _indice(headers, 'Informante', ['Medico Informante', 'Médico Informante'])
    idx_medico_actuante = _indice(headers, 'Medico Actuante', ['Médico Actuante', 'Actuante', 'Medico', 'Médico'])
    idx_tecnico = _indice(headers, 'Tecnico', ['Técnico'])
    idx_duracion = _indice(headers, 'Duracion', ['Duración'])
    idx_contraste = _indice(headers, 'Contraste')
    idx_anestesia = _indice(headers, 'Anestesia')
    idx_aplicacion_origen = _indice(headers, 'Aplicacion Origen', ['Aplicación Origen'])

    if idx_practica is None:
        print("[EGES] Aviso: columna Practica no encontrada; clasificación usará Servicio.")
    if idx_dni is None and idx_hc is None:
        print("[EGES] Aviso: no se encontró DNI ni Historia Clínica.")

    filas_a_crear = []
    filas_error = 0
    filas_procesadas = 0
    nombres_os_cache = {}
    session_medico = {}

    for row_idx, row in enumerate(rows, start=2):
        if not any(row):
            continue
        filas_procesadas += 1
        if filas_procesadas % 500 == 0:
            print(f"[EGES] Leyendo fila {filas_procesadas}...")

        try:
            dni = _texto(_celda(row, idx_dni))
            hc = _texto(_celda(row, idx_hc)) or dni
            fecha_turno = _fecha(_celda(row, idx_fecha), datemode)
            hora_turno = _hora(_celda(row, idx_hora), datemode)
            centro = _texto(_celda(row, idx_centro))
            medico_inf = _texto(_celda(row, idx_medico)) or None
            medico_act = _texto(_celda(row, idx_medico_actuante)) or None
            tecnico = _texto(_celda(row, idx_tecnico)) or None
            medico = next(
                (valor for valor in (medico_inf, medico_act, tecnico) if _valor_valido(valor)),
                None,
            )
            codigo_os = _texto(_celda(row, idx_codigo_os)) or None
            nombre_os = _texto(_celda(row, idx_nombre_os)) or None
            obra_social = _texto(_celda(row, idx_obra_social)) or None

            if nombre_os:
                clave_os = codigo_os or obra_social
                if clave_os:
                    nombres_os_cache[clave_os] = nombre_os

            fila = EgesRow(
                batch=batch,
                numero_turno=_texto(_celda(row, idx_nro_turno)),
                protocolo=_texto(_celda(row, idx_protocolo)) or None,
                fecha_turno=fecha_turno,
                hora_turno=hora_turno,
                hora_hasta=_hora(_celda(row, idx_hora_hasta), datemode),
                centro_atencion=centro,
                tipo_atencion=_texto(_celda(row, idx_tipo_atencion)) or None,
                dni_paciente=dni or None,
                historia_clinica=hc,
                numero_afiliado=_texto(_celda(row, idx_numero_afiliado)) or None,
                apellido_nombre=_texto(_celda(row, idx_nombre)),
                servicio=_texto(_celda(row, idx_servicio)),
                equipo=_texto(_celda(row, idx_equipo)),
                estado_turno=_texto(_celda(row, idx_estado)),
                estado_informe=_texto(_celda(row, idx_estado_informe)) or None,
                tipo_turno=_texto(_celda(row, idx_tipo_turno)) or None,
                tipo_paciente=_texto(_celda(row, idx_tipo_paciente)) or None,
                region_informe=_texto(_celda(row, idx_region_informe)) or None,
                practica=_texto(_celda(row, idx_practica)) or None,
                codigo_practica=_texto(_celda(row, idx_codigo_practica)) or None,
                cantidad=_decimal(_celda(row, idx_cantidad)),
                obra_social=obra_social,
                codigo_obra_social=codigo_os,
                medico_informante=medico,
                medico_actuante=medico_act,
                tecnico=tecnico,
                duracion_minutos=(
                    _decimal(_celda(row, idx_duracion), default='0')
                    if _celda(row, idx_duracion) not in (None, '') else None
                ),
                contraste_eges=_texto(_celda(row, idx_contraste)) or None,
                anestesia_eges=_texto(_celda(row, idx_anestesia)) or None,
                aplicacion_origen=_texto(_celda(row, idx_aplicacion_origen)) or None,
            )

            fila.es_insumo = fila.clasificar_insumo()
            if fila.es_insumo:
                if medico:
                    session_medico.setdefault((hc, fecha_turno, hora_turno, centro), medico)
                continue

            if not fila.medico_informante:
                fila.medico_informante = session_medico.get((hc, fecha_turno, hora_turno, centro))
            fila.modalidad = fila.clasificar_modalidad()
            if fila.modalidad == 'ECO':
                fila.sub_modalidad = fila.clasificar_sub_modalidad()

            filas_a_crear.append(fila)
        except Exception as exc:
            filas_error += 1
            print(f"[EGES] Error en fila {row_idx}: {exc}")

    print(f"[EGES] Leídas {len(filas_a_crear)} filas válidas, {filas_error} errores.")

    for codigo, nombre in nombres_os_cache.items():
        NombreObraSocial.objects.update_or_create(codigo=codigo, defaults={'nombre': nombre})

    combinaciones_existentes = set(
        EgesRow.objects.values_list(
            'historia_clinica', 'fecha_turno', 'hora_turno', 'centro_atencion', 'practica'
        )
    )
    filas_nuevas = []
    filas_duplicadas = 0

    for fila in filas_a_crear:
        combinacion = (
            fila.historia_clinica,
            fila.fecha_turno,
            fila.hora_turno,
            fila.centro_atencion,
            fila.practica,
        )
        if combinacion not in combinaciones_existentes:
            filas_nuevas.append(fila)
            combinaciones_existentes.add(combinacion)
        else:
            filas_duplicadas += 1

    filas_creadas = 0
    if filas_nuevas:
        with transaction.atomic():
            batch_size = 500
            for i in range(0, len(filas_nuevas), batch_size):
                lote = filas_nuevas[i:i + batch_size]
                EgesRow.objects.bulk_create(lote, ignore_conflicts=True)
                filas_creadas += len(lote)

    print(f"[EGES] Proceso completado: {filas_creadas} nuevas, {filas_duplicadas} duplicadas, {filas_error} errores")
    return {
        'creadas': filas_creadas,
        'duplicadas': filas_duplicadas,
        'errores': filas_error,
    }


# ─────────────────────────────────────────────────────────────────────────────
# QuerySets base (sin filtros de fecha/modalidad)
# ─────────────────────────────────────────────────────────────────────────────

def get_base_estudios_finalizados():
    """QuerySet base: estudios finalizados (no insumos, estado Informado, con fecha)."""
    return EgesRow.objects.filter(
        es_insumo=False,
        estado_turno__iexact='Informado',
        fecha_turno__isnull=False,
    ).order_by()


def get_base_rx_sin_informe():
    """QuerySet base: RX entregados sin informe (cerrados pero sin reporte médico)."""
    return EgesRow.objects.filter(
        es_insumo=False,
        modalidad='RX',
        estado_turno__iexact='Entregado Sin Informe',
        fecha_turno__isnull=False,
    )


# ─────────────────────────────────────────────────────────────────────────────
# Filtros dinámicos
# ─────────────────────────────────────────────────────────────────────────────

def aplicar_filtros_fecha_modalidad(qs, request_params):
    """
    Aplica filtros de fecha y modalidad a un QuerySet.

    Parámetros:
        qs            : QuerySet de EgesRow (cualquier estado)
        request_params: QueryDict o dict-like con las claves:
                        fecha_desde (YYYY-MM-DD), fecha_hasta (YYYY-MM-DD),
                        modalidades o modalidades[] (lista de códigos)

    Retorna el QuerySet filtrado (puede ser vacío, nunca None).
    """
    fecha_desde = request_params.get('fecha_desde', '')
    fecha_hasta = request_params.get('fecha_hasta', '')
    # Acepta tanto ?modalidades[]=TC&modalidades[]=RM como ?modalidades=TC,RM
    mods = request_params.getlist('modalidades[]') or request_params.getlist('modalidades')

    if fecha_desde:
        try:
            qs = qs.filter(fecha_turno__gte=datetime.strptime(fecha_desde, '%Y-%m-%d').date())
        except ValueError:
            pass
    if fecha_hasta:
        try:
            qs = qs.filter(fecha_turno__lte=datetime.strptime(fecha_hasta, '%Y-%m-%d').date())
        except ValueError:
            pass
    if mods:
        qs = qs.filter(modalidad__in=mods)
    return qs


# ─────────────────────────────────────────────────────────────────────────────
# Cálculos de KPIs
# ─────────────────────────────────────────────────────────────────────────────

def calcular_kpis(estudios_finalizados_qs, estudios_candidatos_qs, sin_informe_qs=None):
    """
    Devuelve un dict con los KPIs principales.

    Parámetros:
        estudios_finalizados_qs : filtrado por estado Informado
        estudios_candidatos_qs  : todos los no-insumos (con fecha)
        sin_informe_qs          : RX entregados sin informe (estado cerrado pero sin reporte)

    Retorna:
        dict con claves: total_finalizados, total_candidatos, total_pendientes,
        promedio_dia, tasa_conversion, rx_sin_informe.
    """
    resumen_finalizados = estudios_finalizados_qs.aggregate(
        total=Count('id'),
        fecha_min=Min('fecha_turno'),
        fecha_max=Max('fecha_turno'),
        pacientes_sin_dni=Count(
            'id',
            filter=Q(dni_paciente__isnull=True) | Q(dni_paciente=''),
        ),
    )
    total_finalizados = resumen_finalizados['total']
    dnis = {
        dni_normalizado
        for dni in estudios_finalizados_qs.values_list('dni_paciente', flat=True)
        if (dni_normalizado := _normalizar_dni(dni))
    }
    pacientes_atendidos = len(dnis)
    total_candidatos = estudios_candidatos_qs.count()
    rx_sin_informe = sin_informe_qs.count() if sin_informe_qs is not None else 0
    # Los "Entregado Sin Informe" son estudios cerrados, no pendientes reales
    total_pendientes = max(total_candidatos - total_finalizados - rx_sin_informe, 0)

    # Rango de fechas efectivo
    fecha_min = resumen_finalizados['fecha_min']
    fecha_max = resumen_finalizados['fecha_max']
    if fecha_min and fecha_max:
        dias_periodo = max((fecha_max - fecha_min).days + 1, 1)
        # Dias hábiles aproximados (excluimos domingos)
        dias_habiles = sum(
            1 for i in range(dias_periodo)
            if (fecha_min + timedelta(days=i)).weekday() != 6
        )
        promedio_dia = round(total_finalizados / max(dias_habiles, 1), 1)
    else:
        promedio_dia = 0

    tasa_conversion = round((total_finalizados / total_candidatos * 100), 1) if total_candidatos else 0
    practicas_por_paciente = round(total_finalizados / pacientes_atendidos, 2) if pacientes_atendidos else 0

    return {
        'total_finalizados': total_finalizados,
        'practicas_realizadas': total_finalizados,
        'pacientes_atendidos': pacientes_atendidos,
        'pacientes_sin_dni': resumen_finalizados['pacientes_sin_dni'],
        'practicas_por_paciente': practicas_por_paciente,
        'total_candidatos': total_candidatos,
        'total_pendientes': total_pendientes,
        'promedio_dia': promedio_dia,
        'tasa_conversion': tasa_conversion,
        'rx_sin_informe': rx_sin_informe,
    }


# ─────────────────────────────────────────────────────────────────────────────
# Agrupación temporal
# ─────────────────────────────────────────────────────────────────────────────

def agrupar_periodo(d, agrupacion):
    """
    Dado un objeto date y agrupacion (dia|semana|mes), devuelve el inicio del periodo.
    Usado para el endpoint de análisis temporal con granularidad variable.
    """
    if agrupacion == 'dia':
        return d
    elif agrupacion == 'semana':
        return d - timedelta(days=d.weekday())  # lunes de esa semana
    else:
        return date_type(d.year, d.month, 1)


# ─────────────────────────────────────────────────────────────────────────────
# Control de acceso por token (Portal del Director)
# ─────────────────────────────────────────────────────────────────────────────

def verificar_token(token_str):
    """
    Verifica que el token sea válido y esté activo.

    Retorna el objeto DirectorToken si es válido, o None si no lo es.
    No lanza excepciones: los tokens inválidos o inexistentes retornan None.
    """
    try:
        import uuid as _uuid
        token_uuid = _uuid.UUID(str(token_str))
        token = DirectorToken.objects.get(token=token_uuid, activo=True)
        token.registrar_acceso()
        return token
    except (DirectorToken.DoesNotExist, ValueError):
        return None
