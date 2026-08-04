from django.shortcuts import render, redirect, get_object_or_404
from django.urls import reverse_lazy
from django.views import View
from django.views.generic import ListView, CreateView, TemplateView, UpdateView, DeleteView, DetailView
from django.views.generic.edit import FormView
from django.contrib.messages.views import SuccessMessageMixin
from django.contrib import messages
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.db.models import Sum, Count, Q, Prefetch, Case, When, IntegerField
from django.db import transaction
from django.http import FileResponse, HttpResponse, HttpResponseRedirect
from django.contrib.auth import get_user_model
from decimal import Decimal, InvalidOperation
import io, json
import pandas as pd
from datetime import datetime, date, timedelta
from collections import defaultdict
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.units import inch
from .models import (
    Estudios,
    GrupoTarifario,
    TarifaGrupoTarifario,
    RegistroEstudio,
    RegistroEstudiosPorMedico,
    ROLES_LIQUIDAR_COMO_EXTRA_RESIDENCIA,
    SolicitudRevisionHorarioRegistro,
    HistorialRecalculoSolicitudRevisionHorario,
    HistorialRecalculoTarifaGuardiaPasiva,
    HistorialRecalculoTarifaRegistro,
    PreparacionLiquidacionRRHH,
    RevisionAuditoriaEcoRegistro,
    RevisionCruceEgesRegistro,
    CorreccionPacsRegistro,
    GuardiaPasiva,
    SesionContable,
    ConfiguracionGuardiaPasiva,
    HistorialSesionContable,
)
from .grupo_tarifario_mapping import (
    contextos_disponibles_para_estudio,
    es_estudio_cardiologico,
    es_eco_general_real_estudio,
)
from .permisos import puede_ver_desglose_administrativo
from .services_auditoria import (
    evaluar_gate_consistencia_sesion,
    auditar_residentes_eco_por_sesion,
    resumir_pendientes_auditoria_eco,
)
from .services import (
    ROLES_RESIDENCIA,
    adjuntar_ultima_correccion_pacs,
    clasificar_horario_residencia_por_proxy,
    estudio_aplica_descuento_residencia,
)
from .forms import (
    EstudiosAdminForm,
    SolicitudRevisionHorarioRegistroForm,
    SolicitudRevisionHorarioResolucionForm,
    SolicitudRevisionHorarioAplicarForm,
    SolicitudRevisionHorarioRecalcularAplicacionForm,
    SolicitudRevisionHorarioBulkActionForm,
    RevisionAuditoriaEcoRegistroForm,
    RevisionAuditoriaEcoBulkForm,
    RevisionCruceEgesRegistroForm,
    CorreccionPacsRegistroForm,
    CorreccionPacsAplicadaBulkForm,
    AnulacionRegistroEstudioForm,
    PreparacionLiquidacionRRHHForm,
    RegistroEstudiosPorMedicoCreateViewForm,  # Alias de PracticaForm (compatibilidad)
    PracticaForm,
    GuardiaPasivaForm,
    FiltroMedicoMesForm, 
    FiltroEstudiosPorMedicoForm,
    CargaExcelForm,
    TarifaGrupoTarifarioAdminForm,
)
from .services_rrhh import (
    asunto_default_rrhh,
    calcular_hash_snapshot,
    construir_snapshot_liquidacion_rrhh,
    cuerpo_default_rrhh,
    evaluar_requisito_rrhh_para_facturar,
    proxima_version_preparacion_rrhh,
)
from .services_cierre import construir_checklist_cierre_sesion
from .services_eges import (
    construir_preview_cruce_liquidacion_eges,
    procesar_control_eges_sesion,
    resumir_control_eges_sesion,
    serializar_resultado_control_eges,
)
from eges_import.models import ImportBatch
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font
from django.utils.timezone import now
from openpyxl import Workbook
from django.urls import reverse

# ===== PORTAL ADMINISTRATIVO =====


def _puede_acceder_panel_administrativo(user):
    return user.is_superuser or user.rol in ['administrativo', 'jefe_servicio']


def _puede_accion_masiva_revision_horaria(user):
    return user.is_superuser or user.rol == 'jefe_servicio'


def _enriquecer_checklist_cierre_visual(checklist, sesion):
    """Agrega datos de presentacion sin cambiar la logica del checklist."""
    base_solicitudes_url = reverse('liquidacion:solicitudes_revision_horario_list')
    sesion_pk = sesion.pk
    sesion_rrhh_habilitada = sesion.estado in ['CERRADA', 'FACTURADA', 'PAGADA']

    urls_por_item = {
        'registros_validos': f'#gate-sesion-{sesion_pk}',
        'solicitudes_pendientes': (
            f'{base_solicitudes_url}?sesion={sesion_pk}'
            f'&estado={SolicitudRevisionHorarioRegistro.ESTADO_PENDIENTE}'
        ),
        'aprobadas_sin_aplicar': (
            f'{base_solicitudes_url}?sesion={sesion_pk}'
            f'&estado={SolicitudRevisionHorarioRegistro.ESTADO_APROBADA}'
        ),
        'control_eges': reverse(
            'liquidacion:cruce_eges_liquidacion_preview',
            kwargs={'pk': sesion_pk},
        ),
        'preparacion_rrhh': (
            reverse('liquidacion:preparacion_rrhh_preview', kwargs={'pk': sesion_pk})
            if sesion_rrhh_habilitada
            else None
        ),
        'lista_para_facturar': f'#acciones-sesion-{sesion_pk}',
        'sesion_pagada': f'#historial-sesion-{sesion_pk}',
    }
    labels_cortos = {
        'registros_validos': 'Registros',
        'solicitudes_pendientes': 'Solicitudes',
        'aprobadas_sin_aplicar': 'Aplicaciones',
        'control_eges': 'Control EGES',
        'preparacion_rrhh': 'RRHH',
        'lista_para_facturar': 'Facturación',
        'sesion_pagada': 'Pago',
    }

    ayuda_por_item = {
        'registros_validos': 'Abrir el Gate Administrativo para ver bloqueantes, advertencias y acciones sugeridas.',
        'solicitudes_pendientes': 'Ir a la bandeja de revision horaria filtrada por solicitudes pendientes.',
        'aprobadas_sin_aplicar': 'Ir a solicitudes aprobadas que aun deben aplicarse economicamente.',
        'control_eges': 'Cruzar el periodo con EGES y resolver solo las diferencias encontradas.',
        'preparacion_rrhh': 'Preparar o completar el snapshot RRHH de residencia.',
        'lista_para_facturar': 'Volver a la accion principal cuando los pasos previos esten listos.',
        'sesion_pagada': 'Revisar historial y estado final de pago.',
    }
    resolver_por_item = {
        'registros_validos': 'Resolver hallazgos del Gate Administrativo.',
        'solicitudes_pendientes': 'Revisar y aprobar/rechazar solicitudes pendientes.',
        'aprobadas_sin_aplicar': 'Aplicar solicitudes aprobadas para actualizar horario y monto.',
        'control_eges': 'Consolidar el cruce EGES o resolver sus casos pendientes.',
        'preparacion_rrhh': 'Guardar la preparacion RRHH en estado PREPARADO.',
        'lista_para_facturar': 'Completar pasos previos y ejecutar la transicion correspondiente.',
        'sesion_pagada': 'Confirmar pago cuando la sesion este FACTURADA.',
    }

    for item in checklist.get('items', []):
        key = item.get('key')
        item['url'] = urls_por_item.get(key)
        item['label_corto'] = labels_cortos.get(key, item.get('label'))
        item['detalle_corto'] = ''
        item['ayuda'] = ayuda_por_item.get(key, '')
        item['resolver_texto'] = resolver_por_item.get(key, '')

        if key == 'preparacion_rrhh':
            if item.get('detalle') == 'No requerido':
                item['detalle_corto'] = 'No requerido'
            elif not sesion_rrhh_habilitada:
                item['detalle_corto'] = 'No disponible'
            elif item.get('estado') == 'ok':
                item['detalle_corto'] = 'Preparado'
            elif item.get('estado') == 'advertencia':
                item['detalle_corto'] = 'Borrador'
            else:
                item['detalle_corto'] = 'Preparación pendiente'
        elif item.get('count'):
            item['detalle_corto'] = str(item['count'])

    proximo_paso = checklist.get('proximo_paso')
    if proximo_paso:
        proximo_paso['url'] = urls_por_item.get(proximo_paso.get('key'))
        proximo_paso['resolver_texto'] = resolver_por_item.get(proximo_paso.get('key'), '')

    return checklist


def _enriquecer_auditoria_residentes_eco_visual(auditoria, sesion):
    """Agrega enlaces de inspeccion administrativa sin alterar el resultado de auditoria."""
    params_base = {'mes': sesion.mes, 'año': sesion.año}
    for item in auditoria.get('items', []):
        params = {**params_base, 'medico': item.get('medico_id')}
        item['url_liquidacion'] = f"{reverse('liquidacion:liquidacion_mensual')}?{urlencode(params)}"
        for registro_alerta in item.get('registros_alerta', []):
            registro_id = registro_alerta.get('registro_id')
            if registro_id:
                volver_params = _query_volver_sesion(sesion)
                url = reverse('liquidacion:registroestudios_admin_detalle', kwargs={'pk': registro_id})
                if volver_params:
                    url = f'{url}?{urlencode(volver_params)}'
                registro_alerta['url_registro'] = url
        for alerta in item.get('alertas', []):
            alerta['valor_display'] = (
                f"{alerta['valor']:.0%}"
                if alerta.get('tipo') == 'proporcion_extra'
                else alerta['valor']
            )

    for item in auditoria.get('top_alertas', []):
        params = {**params_base, 'medico': item.get('medico_id')}
        item['url_liquidacion'] = f"{reverse('liquidacion:liquidacion_mensual')}?{urlencode(params)}"

    return auditoria


def _flatten_registros_alerta_auditoria_eco(auditoria, medico_id='', motivo=''):
    registros = []
    for item in auditoria.get('items', []):
        if medico_id and str(item.get('medico_id')) != str(medico_id):
            continue
        for registro_alerta in item.get('registros_alerta', []):
            motivos = registro_alerta.get('motivos', [])
            if motivo and motivo not in motivos:
                continue
            registro = {
                **registro_alerta,
                'medico_id': item.get('medico_id'),
                'medico_nombre': item.get('medico_nombre'),
                'rol': item.get('rol'),
                'severidad': item.get('severidad'),
            }
            registros.append(registro)

    registros.sort(key=lambda registro: registro.get('fecha_carga') or '', reverse=True)
    return registros


def _filtrar_registros_alerta_auditoria_eco(
    registros,
    estado_revision='',
    fecha_desde='',
    fecha_hasta='',
    ajuste_pacs='',
):
    """Filtra la bandeja ECO con datos ya enriquecidos de revision."""
    filtrados = []
    for registro in registros:
        revision = registro.get('revision_auditoria_eco')
        revision_eges = registro.get('revision_cruce_eges')
        estado_actual = revision.estado if revision else (
            revision_eges.estado if revision_eges else 'SIN_REVISAR'
        )
        fecha_informe = registro.get('fecha_informe') or ''

        if estado_revision and estado_actual != estado_revision:
            continue
        if fecha_desde and fecha_informe < fecha_desde:
            continue
        if fecha_hasta and fecha_informe > fecha_hasta:
            continue
        if ajuste_pacs == 'CON_AJUSTE' and not registro.get('correccion_pacs'):
            continue
        if ajuste_pacs == 'SIN_AJUSTE' and registro.get('correccion_pacs'):
            continue

        filtrados.append(registro)
    return filtrados


def _volver_sesion_id(request):
    return (request.GET.get('volver_sesion') or request.POST.get('volver_sesion') or '').strip()


def _volver_sesion_url(request):
    sesion_id = _volver_sesion_id(request)
    if not sesion_id:
        return ''
    return f"{reverse('liquidacion:sesiones_list')}#sesion-card-{sesion_id}"


def _query_volver_sesion(sesion):
    sesion_id = getattr(sesion, 'pk', None)
    return {'volver_sesion': sesion_id} if sesion_id else {}


def _accion_para_issue_cierre(issue, sesion=None):
    tipo = issue.get('tipo')
    registro_id = issue.get('registro_id')
    estudio_id = issue.get('estudio_id')
    grupo_id = issue.get('grupo_id')
    guardia_id = issue.get('guardia_id')
    sesion_id = issue.get('sesion_id') or getattr(sesion, 'pk', None)
    fecha = issue.get('fecha')

    if tipo in {'revision_horaria_pendiente', 'revision_horaria_aprobada_sin_aplicar', 'revision_horaria_rechazada'}:
        params = {}
        if sesion_id:
            params['sesion'] = sesion_id
        if issue.get('estado_solicitud'):
            params['estado'] = issue['estado_solicitud']
        return {
            'label': 'Resolver solicitudes' if issue.get('estado') == 'bloqueante' else 'Revisar solicitudes',
            'url': f"{reverse('liquidacion:solicitudes_revision_horario_list')}?{urlencode(params)}",
        }

    if tipo == 'recalculo_b3' and sesion_id:
        return {
            'label': 'Revisar solicitudes',
            'url': f"{reverse('liquidacion:solicitudes_revision_horario_list')}?{urlencode({'sesion': sesion_id})}",
        }

    if tipo in {'sin_tarifa_vigente_grupo', 'contextual_sin_tarifa'} and grupo_id:
        params = _query_volver_sesion(sesion)
        url = reverse('liquidacion:grupo_tarifario_tarifa_nueva', kwargs={'grupo_pk': grupo_id})
        if params:
            url = f'{url}?{urlencode(params)}'
        return {
            'label': 'Cargar tarifa',
            'url': url,
        }

    if tipo == 'contextual_sin_grupo':
        params = _query_volver_sesion(sesion)
        url = reverse('liquidacion:grupos_tarifarios_list')
        if params:
            url = f'{url}?{urlencode(params)}'
        return {
            'label': 'Revisar grupos tarifarios',
            'url': url,
        }

    if tipo in {'sin_precio_resoluble', 'sin_grupo_con_fallback'} and estudio_id:
        params = _query_volver_sesion(sesion)
        url = reverse('liquidacion:estudios_edit', kwargs={'pk': estudio_id})
        if params:
            url = f'{url}?{urlencode(params)}'
        return {
            'label': 'Revisar estudio',
            'url': url,
        }

    if guardia_id:
        params = _query_volver_sesion(sesion)
        url = reverse('liquidacion:editar_guardia_pasiva', kwargs={'pk': guardia_id})
        if params:
            url = f'{url}?{urlencode(params)}'
        return {
            'label': 'Editar guardia',
            'url': url,
        }

    if registro_id:
        params = _query_volver_sesion(sesion)
        url = reverse('liquidacion:registroestudios_admin_detalle', kwargs={'pk': registro_id})
        if params:
            url = f'{url}?{urlencode(params)}'
        return {
            'label': 'Inspeccionar registro',
            'url': url,
        }

    return None


def _enriquecer_issues_cierre(issues, sesion=None, limite=None):
    enriquecidos = []
    issues_iterables = issues[:limite] if limite else issues
    for issue in issues_iterables:
        item = dict(issue)
        item['accion'] = _accion_para_issue_cierre(item, sesion=sesion)
        enriquecidos.append(item)
    return enriquecidos


def _build_diagnostico_recalculo_b3(solicitud):
    """Diagnostico solo lectura para explicar el recalculo B3."""
    if not solicitud.fecha_aplicacion:
        return None

    registro = solicitud.registro
    horario_aplicado = solicitud.horario_aplicado
    rol_medico = getattr(registro.medico, 'rol', None)
    fecha_referencia = registro.fecha_del_informe
    monto_actual = registro.monto_calculado
    advertencias = []

    horario_original = registro.horario
    try:
        registro.horario = horario_aplicado
        monto_simulado = registro.calcular_monto()
    finally:
        registro.horario = horario_original

    diferencia = monto_simulado - monto_actual
    if diferencia == 0:
        advertencias.append('El recálculo no cambiaría el monto actual.')
    if horario_aplicado != 'INTRA':
        advertencias.append('No aplica descuento porque el horario aplicado no es INTRA.')

    estudios = []
    relaciones = (
        registro.registroestudio_set
        .select_related('estudio__grupo_tarifario')
        .order_by('id')
    )
    for rel in relaciones:
        estudio = rel.estudio
        resultado = estudio_aplica_descuento_residencia(
            estudio,
            rol_medico,
            fecha_referencia,
        )
        grupo = getattr(estudio, 'grupo_tarifario', None)
        advertencia = ''
        if horario_aplicado != 'INTRA':
            advertencia = 'No aplica descuento porque el horario aplicado no es INTRA.'
        elif (
            (estudio.tipo or '').upper() == 'DOP'
            and not resultado['aplica']
            and resultado['fuente'] == 'fallback_legado'
        ):
            advertencia = 'No existe regla activa aplicable para este estudio en la fecha del informe.'

        estudios.append({
            'nombre': estudio.nombre,
            'tipo': estudio.tipo,
            'grupo_tarifario': grupo.codigo if grupo else '',
            'aplica': resultado['aplica'],
            'fuente': resultado['fuente'],
            'regla_id': resultado['regla_id'],
            'motivo': resultado['motivo'],
            'advertencia': advertencia,
        })

    return {
        'horario_aplicado': horario_aplicado,
        'rol_medico': rol_medico,
        'fecha_referencia': fecha_referencia,
        'monto_actual_registro': monto_actual,
        'monto_simulado_con_reglas_vigentes': monto_simulado,
        'diferencia': diferencia,
        'estudios': estudios,
        'advertencias': advertencias,
    }


def _build_resumen_aplicacion_b4(solicitudes):
    """Resumen solo lectura del impacto de aplicar las solicitudes visibles."""
    resumen = {
        'cantidad_aplicables': 0,
        'monto_actual_total': Decimal('0.00'),
        'monto_simulado_total': Decimal('0.00'),
        'diferencia_total': Decimal('0.00'),
        'items': [],
    }

    for solicitud in solicitudes:
        registro = solicitud.registro
        es_aplicable = (
            solicitud.estado == SolicitudRevisionHorarioRegistro.ESTADO_APROBADA
            and solicitud.fecha_aplicacion is None
        )
        impacto = {
            'aplicable': es_aplicable,
            'monto_actual': registro.monto_calculado,
            'monto_simulado': None,
            'diferencia': None,
        }

        if es_aplicable:
            horario_original = registro.horario
            registro.horario = solicitud.horario_solicitado
            try:
                monto_simulado = registro.calcular_monto()
            finally:
                registro.horario = horario_original

            diferencia = monto_simulado - registro.monto_calculado
            impacto.update({
                'monto_simulado': monto_simulado,
                'diferencia': diferencia,
            })
            resumen['cantidad_aplicables'] += 1
            resumen['monto_actual_total'] += registro.monto_calculado
            resumen['monto_simulado_total'] += monto_simulado
            resumen['diferencia_total'] += diferencia
            resumen['items'].append({
                'solicitud': solicitud,
                'monto_actual': registro.monto_calculado,
                'monto_simulado': monto_simulado,
                'diferencia': diferencia,
            })

        solicitud.impacto_aplicacion_b4 = impacto

    return resumen


class PortalLiquidacionInicioView(LoginRequiredMixin, UserPassesTestMixin, TemplateView):
    """Vista de inicio del portal administrativo de liquidación"""
    template_name = 'liquidacion/portal_inicio.html'

    def test_func(self):
        return _puede_acceder_panel_administrativo(self.request.user)

    def handle_no_permission(self):
        messages.error(
            self.request,
            '❌ No tienes permisos para acceder al portal administrativo de liquidación.'
        )
        return redirect('home')


class GruposTarifariosListView(LoginRequiredMixin, UserPassesTestMixin, ListView):
    """Listado administrativo global de grupos tarifarios y su estado de vigencia."""

    model = GrupoTarifario
    template_name = 'liquidacion/grupos_tarifarios_list.html'
    context_object_name = 'grupos'

    def test_func(self):
        return _puede_acceder_panel_administrativo(self.request.user)

    def handle_no_permission(self):
        messages.error(
            self.request,
            '❌ No tienes permisos para acceder a la configuración económica de liquidación.'
        )
        return redirect('home')

    def get_queryset(self):
        return (
            GrupoTarifario.objects
            .annotate(cantidad_estudios=Count('estudios', distinct=True))
            .prefetch_related(
                Prefetch(
                    'tarifas',
                    queryset=TarifaGrupoTarifario.objects.order_by('-vigencia_desde'),
                    to_attr='tarifas_ordenadas',
                )
            )
            .order_by('modalidad', 'codigo')
        )

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        hoy = date.today()
        grupos_data = []

        for grupo in context['grupos']:
            tarifa_vigente = None
            for tarifa in getattr(grupo, 'tarifas_ordenadas', []):
                if tarifa.vigencia_desde <= hoy and (
                    tarifa.vigencia_hasta is None or tarifa.vigencia_hasta >= hoy
                ):
                    tarifa_vigente = tarifa
                    break

            grupos_data.append({
                'grupo': grupo,
                'tarifa_vigente': tarifa_vigente,
                'tiene_tarifa_vigente': tarifa_vigente is not None,
            })

        context['grupos_data'] = grupos_data
        context['volver_sesion_id'] = _volver_sesion_id(self.request)
        context['volver_sesion_url'] = _volver_sesion_url(self.request)
        return context


class TarifasGrupoBulkUpdateView(LoginRequiredMixin, UserPassesTestMixin, TemplateView):
    """Actualizacion masiva de tarifas por vigencia para grupos tarifarios."""

    template_name = 'liquidacion/grupos_tarifarios_tarifas_bulk_update.html'

    def test_func(self):
        return _puede_acceder_panel_administrativo(self.request.user)

    def handle_no_permission(self):
        messages.error(
            self.request,
            '❌ No tienes permisos para actualizar tarifas de grupos tarifarios.'
        )
        return redirect('home')

    def _grupos_queryset(self):
        return (
            GrupoTarifario.objects
            .filter(activo=True)
            .annotate(cantidad_estudios=Count('estudios', distinct=True))
            .prefetch_related(
                Prefetch(
                    'tarifas',
                    queryset=TarifaGrupoTarifario.objects.order_by('-vigencia_desde'),
                    to_attr='tarifas_ordenadas',
                )
            )
            .order_by('modalidad', 'codigo')
        )

    def _tarifa_vigente(self, grupo, fecha_ref):
        for tarifa in getattr(grupo, 'tarifas_ordenadas', []):
            if tarifa.vigencia_desde <= fecha_ref and (
                tarifa.vigencia_hasta is None or tarifa.vigencia_hasta >= fecha_ref
            ):
                return tarifa
        return None

    def _guardia_pasiva_vigente(self, fecha_ref):
        return (
            ConfiguracionGuardiaPasiva.objects
            .filter(vigente_desde__lte=fecha_ref)
            .filter(Q(vigente_hasta__isnull=True) | Q(vigente_hasta__gte=fecha_ref))
            .order_by('-vigente_desde', '-id')
            .first()
        )

    def _parse_decimal(self, value):
        value = (value or '').strip().replace(',', '.')
        if not value:
            return None
        try:
            parsed = Decimal(value)
        except (InvalidOperation, ValueError):
            return None
        return parsed.quantize(Decimal('0.01'))

    def _build_rows(self, vigencia_desde=None, post_data=None):
        fecha_ref = vigencia_desde or date.today()
        rows = []
        for grupo in self._grupos_queryset():
            tarifa_vigente = self._tarifa_vigente(grupo, fecha_ref)
            precio_cober_default = tarifa_vigente.precio_cober if tarifa_vigente else ''
            precio_otras_default = tarifa_vigente.precio_otras_os if tarifa_vigente else ''
            incluir_default = True

            if post_data is not None:
                incluir_default = post_data.get(f'incluir_{grupo.pk}') == 'on'
                precio_cober_default = post_data.get(f'precio_cober_{grupo.pk}', '')
                precio_otras_default = post_data.get(f'precio_otras_os_{grupo.pk}', '')

            rows.append({
                'grupo': grupo,
                'tarifa_vigente': tarifa_vigente,
                'incluir': incluir_default,
                'precio_cober': precio_cober_default,
                'precio_otras_os': precio_otras_default,
            })
        return rows

    def _build_guardia_pasiva_row(self, vigencia_desde=None, post_data=None):
        fecha_ref = vigencia_desde or date.today()
        tarifa_vigente = self._guardia_pasiva_vigente(fecha_ref)
        incluir_default = False
        monto_default = tarifa_vigente.monto_vigente if tarifa_vigente else ''

        if post_data is not None:
            incluir_default = post_data.get('incluir_guardia_pasiva') == 'on'
            monto_default = post_data.get('monto_guardia_pasiva', '')

        return {
            'tarifa_vigente': tarifa_vigente,
            'incluir': incluir_default,
            'monto': monto_default,
        }

    def _validar_preview(self, request):
        errors = []
        warnings = []
        preview = []
        vigencia_raw = request.POST.get('vigencia_desde') or ''
        motivo = (request.POST.get('motivo_actualizacion') or '').strip()

        try:
            vigencia_desde = datetime.strptime(vigencia_raw, '%Y-%m-%d').date()
        except ValueError:
            vigencia_desde = None
            errors.append('Indica una vigencia desde válida.')

        rows = self._build_rows(vigencia_desde=vigencia_desde, post_data=request.POST)
        guardia_pasiva_row = self._build_guardia_pasiva_row(vigencia_desde=vigencia_desde, post_data=request.POST)
        if not vigencia_desde:
            return None, motivo, rows, guardia_pasiva_row, preview, errors, warnings

        seleccionados = [row for row in rows if row['incluir']]
        incluir_guardia_pasiva = guardia_pasiva_row['incluir']
        if not seleccionados and not incluir_guardia_pasiva:
            errors.append('Selecciona al menos un grupo tarifario o la guardia pasiva para actualizar.')

        dia_anterior = vigencia_desde - timedelta(days=1)
        for row in seleccionados:
            grupo = row['grupo']
            precio_cober = self._parse_decimal(row['precio_cober'])
            precio_otras_os = self._parse_decimal(row['precio_otras_os'])

            if precio_cober is None or precio_cober <= 0:
                errors.append(f'{grupo.codigo}: el precio COBER debe ser mayor a 0.')
                continue
            if precio_otras_os is None or precio_otras_os <= 0:
                errors.append(f'{grupo.codigo}: el precio OTRAS OS debe ser mayor a 0.')
                continue

            if TarifaGrupoTarifario.objects.filter(
                grupo_tarifario=grupo,
                vigencia_desde=vigencia_desde,
            ).exists():
                errors.append(f'{grupo.codigo}: ya existe una tarifa con vigencia {vigencia_desde:%d/%m/%Y}.')
                continue

            futuras_solapadas = TarifaGrupoTarifario.objects.filter(
                grupo_tarifario=grupo,
                vigencia_desde__gt=vigencia_desde,
            ).filter(
                Q(vigencia_hasta__isnull=True) | Q(vigencia_hasta__gte=vigencia_desde)
            )
            if futuras_solapadas.exists():
                errors.append(f'{grupo.codigo}: existe una tarifa futura que se solaparía. Revisar detalle del grupo.')
                continue

            tarifa_anterior = (
                TarifaGrupoTarifario.objects
                .filter(grupo_tarifario=grupo, vigencia_desde__lt=vigencia_desde)
                .filter(Q(vigencia_hasta__isnull=True) | Q(vigencia_hasta__gte=vigencia_desde))
                .order_by('-vigencia_desde', '-id')
                .first()
            )
            if not tarifa_anterior:
                warnings.append(f'{grupo.codigo}: no se encontró tarifa anterior vigente para cerrar.')

            sin_cambios = (
                tarifa_anterior
                and tarifa_anterior.precio_cober == precio_cober
                and tarifa_anterior.precio_otras_os == precio_otras_os
            )

            preview.append({
                'tipo': 'grupo',
                'grupo': grupo,
                'tarifa_anterior': tarifa_anterior,
                'precio_cober_nuevo': precio_cober,
                'precio_otras_os_nuevo': precio_otras_os,
                'vigencia_desde': vigencia_desde,
                'vigencia_hasta_anterior': dia_anterior if tarifa_anterior else None,
                'sin_cambios': sin_cambios,
            })

        if incluir_guardia_pasiva:
            monto_guardia = self._parse_decimal(guardia_pasiva_row['monto'])
            if monto_guardia is None or monto_guardia <= 0:
                errors.append('Guardia pasiva: el monto debe ser mayor a 0.')
            elif ConfiguracionGuardiaPasiva.objects.filter(vigente_desde=vigencia_desde).exists():
                errors.append(f'Guardia pasiva: ya existe una tarifa con vigencia {vigencia_desde:%d/%m/%Y}.')
            else:
                futuras_solapadas = (
                    ConfiguracionGuardiaPasiva.objects
                    .filter(vigente_desde__gt=vigencia_desde)
                    .filter(Q(vigente_hasta__isnull=True) | Q(vigente_hasta__gte=vigencia_desde))
                )
                if futuras_solapadas.exists():
                    errors.append('Guardia pasiva: existe una tarifa futura que se solaparia.')
                else:
                    tarifa_anterior = (
                        ConfiguracionGuardiaPasiva.objects
                        .filter(vigente_desde__lt=vigencia_desde)
                        .filter(Q(vigente_hasta__isnull=True) | Q(vigente_hasta__gte=vigencia_desde))
                        .order_by('-vigente_desde', '-id')
                        .first()
                    )
                    if not tarifa_anterior:
                        warnings.append('Guardia pasiva: no se encontro tarifa anterior vigente para cerrar.')

                    preview.append({
                        'tipo': 'guardia_pasiva',
                        'tarifa_anterior': tarifa_anterior,
                        'monto_nuevo': monto_guardia,
                        'vigencia_desde': vigencia_desde,
                        'vigencia_hasta_anterior': dia_anterior if tarifa_anterior else None,
                        'sin_cambios': tarifa_anterior and tarifa_anterior.monto_vigente == monto_guardia,
                    })

        return vigencia_desde, motivo, rows, guardia_pasiva_row, preview, errors, warnings

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        vigencia_desde = date.today().replace(day=1)
        context.update({
            'vigencia_desde': vigencia_desde,
            'motivo_actualizacion': f'Actualización valores {vigencia_desde:%m/%Y}',
            'rows': self._build_rows(vigencia_desde=vigencia_desde),
            'guardia_pasiva_row': self._build_guardia_pasiva_row(vigencia_desde=vigencia_desde),
            'preview': None,
            'errors_preview': [],
            'warnings_preview': [],
        })
        return context

    def post(self, request, *args, **kwargs):
        vigencia_desde, motivo, rows, guardia_pasiva_row, preview, errors, warnings = self._validar_preview(request)
        confirmar = request.POST.get('confirmar') == '1'

        if errors or not confirmar:
            return self.render_to_response({
                **self.get_context_data(),
                'vigencia_desde': vigencia_desde,
                'motivo_actualizacion': motivo,
                'rows': rows,
                'guardia_pasiva_row': guardia_pasiva_row,
                'preview': preview if not errors else None,
                'errors_preview': errors,
                'warnings_preview': warnings,
            })

        with transaction.atomic():
            creadas = 0
            guardias_creadas = 0
            cerradas = 0
            dia_anterior = vigencia_desde - timedelta(days=1)

            for item in preview:
                tarifa_anterior = item['tarifa_anterior']
                if item['tipo'] == 'guardia_pasiva':
                    if tarifa_anterior and (
                        tarifa_anterior.vigente_hasta is None
                        or tarifa_anterior.vigente_hasta >= vigencia_desde
                    ):
                        tarifa_anterior.vigente_hasta = dia_anterior
                        tarifa_anterior.save(update_fields=['vigente_hasta'])
                        cerradas += 1

                    ConfiguracionGuardiaPasiva.objects.create(
                        monto_vigente=item['monto_nuevo'],
                        vigente_desde=vigencia_desde,
                        vigente_hasta=None,
                        motivo_actualizacion=motivo or f'Actualizacion masiva desde {vigencia_desde:%d/%m/%Y}',
                        actualizado_por=request.user,
                    )
                    guardias_creadas += 1
                    continue

                if tarifa_anterior and (
                    tarifa_anterior.vigencia_hasta is None
                    or tarifa_anterior.vigencia_hasta >= vigencia_desde
                ):
                    tarifa_anterior.vigencia_hasta = dia_anterior
                    tarifa_anterior.save(update_fields=['vigencia_hasta'])
                    cerradas += 1

                TarifaGrupoTarifario.objects.create(
                    grupo_tarifario=item['grupo'],
                    vigencia_desde=vigencia_desde,
                    vigencia_hasta=None,
                    precio_cober=item['precio_cober_nuevo'],
                    precio_otras_os=item['precio_otras_os_nuevo'],
                    motivo_actualizacion=motivo or f'Actualización masiva desde {vigencia_desde:%d/%m/%Y}',
                    actualizado_por=request.user,
                )
                creadas += 1

        messages.success(
            request,
            f'Actualización aplicada: {creadas} tarifa(s) nueva(s), {cerradas} vigencia(s) anterior(es) cerrada(s).'
        )
        return redirect('liquidacion:grupos_tarifarios_list')


class GrupoTarifarioDetalleView(LoginRequiredMixin, UserPassesTestMixin, DetailView):
    """Detalle administrativo de un grupo tarifario (solo lectura)."""

    model = GrupoTarifario
    template_name = 'liquidacion/grupo_tarifario_detalle.html'
    context_object_name = 'grupo'

    def test_func(self):
        return _puede_acceder_panel_administrativo(self.request.user)

    def handle_no_permission(self):
        messages.error(
            self.request,
            '❌ No tienes permisos para acceder al detalle de grupos tarifarios.'
        )
        return redirect('home')

    def get_queryset(self):
        return (
            GrupoTarifario.objects
            .prefetch_related(
                Prefetch(
                    'tarifas',
                    queryset=TarifaGrupoTarifario.objects.order_by('-vigencia_desde'),
                    to_attr='tarifas_ordenadas',
                ),
                Prefetch('estudios', queryset=Estudios.objects.order_by('nombre')),
            )
        )

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        hoy = date.today()
        grupo = context['grupo']
        historial_tarifas = list(getattr(grupo, 'tarifas_ordenadas', []))
        tarifa_vigente = None

        for tarifa in historial_tarifas:
            if tarifa.vigencia_desde <= hoy and (
                tarifa.vigencia_hasta is None or tarifa.vigencia_hasta >= hoy
            ):
                tarifa_vigente = tarifa
                break

        estudios_asociados = list(grupo.estudios.all())

        context['tarifa_vigente'] = tarifa_vigente
        context['historial_tarifas'] = historial_tarifas
        context['estudios_asociados'] = estudios_asociados
        context['cantidad_estudios_asociados'] = len(estudios_asociados)
        context['tiene_tarifa_vigente'] = tarifa_vigente is not None
        context['volver_sesion_id'] = _volver_sesion_id(self.request)
        context['volver_sesion_url'] = _volver_sesion_url(self.request)
        return context


class GrupoTarifarioTarifaNuevaView(LoginRequiredMixin, UserPassesTestMixin, SuccessMessageMixin, CreateView):
    """Alta administrativa de nueva tarifa para un grupo tarifario."""

    model = TarifaGrupoTarifario
    form_class = TarifaGrupoTarifarioAdminForm
    template_name = 'liquidacion/grupo_tarifario_tarifa_form.html'
    success_message = 'Tarifa creada correctamente.'

    def test_func(self):
        return _puede_acceder_panel_administrativo(self.request.user)

    def handle_no_permission(self):
        messages.error(
            self.request,
            '❌ No tienes permisos para crear tarifas de grupos tarifarios.'
        )
        return redirect('home')

    def dispatch(self, request, *args, **kwargs):
        self.grupo_tarifario = get_object_or_404(GrupoTarifario, pk=kwargs['grupo_pk'])
        return super().dispatch(request, *args, **kwargs)

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['grupo_tarifario'] = self.grupo_tarifario
        return kwargs

    def form_valid(self, form):
        with transaction.atomic():
            vigencia_desde = form.instance.vigencia_desde
            dia_anterior = vigencia_desde - timedelta(days=1)
            tarifas_cerradas = (
                TarifaGrupoTarifario.objects
                .filter(grupo_tarifario=self.grupo_tarifario, vigencia_desde__lt=vigencia_desde)
                .filter(Q(vigencia_hasta__isnull=True) | Q(vigencia_hasta__gte=vigencia_desde))
                .update(vigencia_hasta=dia_anterior)
            )
            form.instance.grupo_tarifario = self.grupo_tarifario
            form.instance.actualizado_por = self.request.user
            response = super().form_valid(form)

        if tarifas_cerradas:
            messages.info(
                self.request,
                f'Se cerró automáticamente {tarifas_cerradas} tarifa anterior al {dia_anterior:%d/%m/%Y}.',
            )
        return response

    def get_success_url(self):
        volver_sesion = _volver_sesion_url(self.request)
        if volver_sesion:
            return volver_sesion
        return reverse('liquidacion:grupo_tarifario_detalle', kwargs={'pk': self.grupo_tarifario.pk})

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['grupo'] = self.grupo_tarifario
        context['volver_sesion_id'] = _volver_sesion_id(self.request)
        context['volver_sesion_url'] = _volver_sesion_url(self.request)
        return context


class SolicitudRevisionHorarioAdminListView(LoginRequiredMixin, UserPassesTestMixin, ListView):
    """Bandeja administrativa de solicitudes de revision de horario (solo lectura)."""

    model = SolicitudRevisionHorarioRegistro
    template_name = 'liquidacion/solicitudes_revision_horario_list.html'
    context_object_name = 'solicitudes'

    def test_func(self):
        return _puede_acceder_panel_administrativo(self.request.user)

    def handle_no_permission(self):
        messages.error(
            self.request,
            '❌ No tienes permisos para acceder a la bandeja de revisión de horarios.',
        )
        return redirect('home')

    def get_queryset(self):
        queryset = (
            SolicitudRevisionHorarioRegistro.objects
            .select_related(
                'solicitado_por',
                'registro',
                'registro__medico',
                'registro__sesion_contable',
            )
            .prefetch_related(
                Prefetch(
                    'registro__registroestudio_set',
                    queryset=RegistroEstudio.objects.select_related('estudio').order_by('id'),
                )
            )
        )

        estado = (self.request.GET.get('estado') or '').strip()
        medico_id = (self.request.GET.get('medico') or '').strip()
        sesion_id = (self.request.GET.get('sesion') or '').strip()
        pendiente_aplicacion = (self.request.GET.get('pendiente_aplicacion') or '').strip() == '1'

        if pendiente_aplicacion:
            queryset = queryset.filter(
                estado=SolicitudRevisionHorarioRegistro.ESTADO_APROBADA,
                fecha_aplicacion__isnull=True,
            )
        elif estado:
            queryset = queryset.filter(estado=estado)
        if medico_id:
            queryset = queryset.filter(registro__medico_id=medico_id)
        if sesion_id:
            queryset = queryset.filter(registro__sesion_contable_id=sesion_id)

        return queryset.annotate(
            prioridad_estado=Case(
                When(estado=SolicitudRevisionHorarioRegistro.ESTADO_PENDIENTE, then=0),
                default=1,
                output_field=IntegerField(),
            )
        ).order_by('prioridad_estado', '-fecha_solicitud')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        base_qs = SolicitudRevisionHorarioRegistro.objects.select_related(
            'registro__medico',
            'registro__sesion_contable',
        )
        medico_ids = base_qs.values_list('registro__medico_id', flat=True).distinct()
        sesion_ids = base_qs.values_list('registro__sesion_contable_id', flat=True).distinct()

        User = get_user_model()
        context['medicos_filtro'] = User.objects.filter(id__in=medico_ids).order_by('username')
        context['sesiones_filtro'] = SesionContable.objects.filter(id__in=sesion_ids).order_by('-año', '-mes')
        context['estado_choices'] = SolicitudRevisionHorarioRegistro.ESTADO_CHOICES
        context['estado_actual'] = (self.request.GET.get('estado') or '').strip()
        context['medico_actual'] = (self.request.GET.get('medico') or '').strip()
        context['sesion_actual'] = (self.request.GET.get('sesion') or '').strip()
        context['pendiente_aplicacion_actual'] = (self.request.GET.get('pendiente_aplicacion') or '').strip() == '1'
        context['puede_accion_masiva_b4'] = _puede_accion_masiva_revision_horaria(self.request.user)
        context['resumen_aplicacion_b4'] = _build_resumen_aplicacion_b4(context['solicitudes'])
        return context


class SolicitudRevisionHorarioAdminDetailView(LoginRequiredMixin, UserPassesTestMixin, DetailView):
    """Detalle administrativo de una solicitud de revision de horario (solo lectura)."""

    model = SolicitudRevisionHorarioRegistro
    template_name = 'liquidacion/solicitud_revision_horario_detalle.html'
    context_object_name = 'solicitud'

    def test_func(self):
        return _puede_acceder_panel_administrativo(self.request.user)

    def handle_no_permission(self):
        messages.error(
            self.request,
            '❌ No tienes permisos para ver el detalle de revisión de horario.',
        )
        return redirect('home')

    def get_queryset(self):
        return (
            SolicitudRevisionHorarioRegistro.objects
            .select_related(
                'solicitado_por',
                'registro',
                'registro__medico',
                'registro__sesion_contable',
            )
            .prefetch_related(
                Prefetch(
                    'registro__registroestudio_set',
                    queryset=RegistroEstudio.objects.select_related('estudio').order_by('id'),
                )
            )
        )

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['resolucion_form'] = SolicitudRevisionHorarioResolucionForm()
        context['aplicar_form'] = SolicitudRevisionHorarioAplicarForm()
        context['recalcular_aplicacion_form'] = SolicitudRevisionHorarioRecalcularAplicacionForm()
        sesion = self.object.registro.sesion_contable
        context['puede_recalcular_aplicacion'] = (
            self.object.estado == SolicitudRevisionHorarioRegistro.ESTADO_APROBADA
            and bool(self.object.fecha_aplicacion)
            and bool(self.object.horario_aplicado)
            and bool(sesion)
            and sesion.estado in ['ABIERTA', 'REVISION']
        )
        context['diagnostico_recalculo_b3'] = _build_diagnostico_recalculo_b3(self.object)
        return context


class RegistroEstudiosPorMedicoAdminDetailView(LoginRequiredMixin, UserPassesTestMixin, DetailView):
    """Inspeccion administrativa read-only de un registro que bloquea el cierre."""

    model = RegistroEstudiosPorMedico
    template_name = 'liquidacion/registroestudios_admin_detalle.html'
    context_object_name = 'registro'

    def test_func(self):
        return _puede_acceder_panel_administrativo(self.request.user)

    def handle_no_permission(self):
        messages.error(self.request, 'No tienes permisos para inspeccionar registros de liquidacion.')
        return redirect('home')

    def get_queryset(self):
        return (
            RegistroEstudiosPorMedico.objects
            .select_related('medico', 'sesion_contable')
            .prefetch_related(
                Prefetch(
                    'registroestudio_set',
                    queryset=RegistroEstudio.objects.select_related('estudio__grupo_tarifario').order_by('id'),
                )
            )
        )

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['volver_sesion_url'] = _volver_sesion_url(self.request)
        context['anulacion_form'] = AnulacionRegistroEstudioForm()
        context['puede_anular'] = (
            not self.object.anulado
            and (
                not self.object.sesion_contable
                or self.object.sesion_contable.estado not in {'FACTURADA', 'PAGADA'}
            )
        )
        context['solicitudes_revision'] = (
            SolicitudRevisionHorarioRegistro.objects
            .filter(registro=self.object)
            .select_related('solicitado_por', 'revisado_por', 'aplicado_por')
            .order_by('-fecha_solicitud')
        )
        return context


class RegistroEstudiosPorMedicoAnularView(LoginRequiredMixin, UserPassesTestMixin, View):
    """Anula economicamente un registro sin borrar su evidencia historica."""

    def test_func(self):
        return _puede_acceder_panel_administrativo(self.request.user)

    def handle_no_permission(self):
        messages.error(self.request, 'No tienes permisos para anular registros de liquidacion.')
        return redirect('home')

    def post(self, request, *args, **kwargs):
        form = AnulacionRegistroEstudioForm(request.POST)
        redirect_url = request.POST.get('next') or reverse(
            'liquidacion:registroestudios_admin_detalle',
            kwargs={'pk': kwargs['pk']},
        )
        if not form.is_valid():
            messages.error(request, 'Debes indicar el motivo de la anulacion.')
            return redirect(redirect_url)

        resultado = None
        with transaction.atomic():
            registro = RegistroEstudiosPorMedico.objects.select_for_update().get(pk=kwargs['pk'])
            sesion = (
                SesionContable.objects.filter(pk=registro.sesion_contable_id).first()
                if registro.sesion_contable_id
                else None
            )
            if registro.anulado:
                resultado = 'YA_ANULADO'
            elif sesion and sesion.estado in {'FACTURADA', 'PAGADA'}:
                resultado = 'SESION_BLOQUEADA'
            else:
                fecha = now()
                motivo = form.cleaned_data['motivo']
                registro.anulado = True
                registro.fecha_anulacion = fecha
                registro.anulado_por = request.user
                registro.motivo_anulacion = motivo
                registro.modificado_por = request.user
                registro.fecha_modificacion = fecha
                registro.motivo_modificacion = f'Registro anulado administrativamente. {motivo}'
                registro.save(update_fields=[
                    'anulado',
                    'fecha_anulacion',
                    'anulado_por',
                    'motivo_anulacion',
                    'modificado_por',
                    'fecha_modificacion',
                    'motivo_modificacion',
                ])
                resultado = 'ANULADO'

        if resultado == 'ANULADO':
            messages.success(
                request,
                'Registro anulado. Se conserva la trazabilidad y queda excluido de liquidacion y exportaciones.',
            )
        elif resultado == 'YA_ANULADO':
            messages.info(request, 'El registro ya estaba anulado.')
        else:
            messages.error(request, 'No se puede anular un registro de una sesion FACTURADA o PAGADA.')
        return redirect(redirect_url)

class AuditoriaEcoSesionView(LoginRequiredMixin, UserPassesTestMixin, TemplateView):
    """Vista completa read-only de registros ECO sospechosos por sesion."""

    template_name = 'liquidacion/auditoria_eco_sesion.html'

    def test_func(self):
        return _puede_acceder_panel_administrativo(self.request.user)

    def handle_no_permission(self):
        messages.error(self.request, 'No tienes permisos para revisar la auditoria ECO.')
        return redirect('home')

    def dispatch(self, request, *args, **kwargs):
        self.sesion = get_object_or_404(SesionContable, pk=kwargs['pk'])
        return super().dispatch(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        auditoria = _enriquecer_auditoria_residentes_eco_visual(
            auditar_residentes_eco_por_sesion(self.sesion),
            self.sesion,
        )
        medico_actual = (self.request.GET.get('medico') or '').strip()
        motivo_actual = (self.request.GET.get('motivo') or '').strip()
        estado_revision_actual = (self.request.GET.get('estado_revision') or '').strip()
        fecha_desde_actual = (self.request.GET.get('fecha_desde') or '').strip()
        fecha_hasta_actual = (self.request.GET.get('fecha_hasta') or '').strip()
        ajuste_pacs_actual = (self.request.GET.get('ajuste_pacs') or '').strip()
        registros_alerta = _flatten_registros_alerta_auditoria_eco(
            auditoria,
            medico_id=medico_actual,
            motivo=motivo_actual,
        )
        registro_ids = [registro['registro_id'] for registro in registros_alerta]
        revisiones_por_registro = {}
        revisiones = (
            RevisionAuditoriaEcoRegistro.objects
            .filter(sesion_contable=self.sesion, registro_id__in=registro_ids)
            .select_related('revisado_por')
            .order_by('registro_id', '-fecha_revision')
        )
        for revision in revisiones:
            if revision.registro_id not in revisiones_por_registro:
                revisiones_por_registro[revision.registro_id] = revision
        revisiones_eges_por_registro = {}
        revisiones_eges = (
            RevisionCruceEgesRegistro.objects
            .filter(sesion_contable=self.sesion, registro_id__in=registro_ids)
            .select_related('revisado_por', 'batch_eges')
            .order_by('registro_id', '-fecha_revision')
        )
        for revision in revisiones_eges:
            if revision.registro_id not in revisiones_eges_por_registro:
                revisiones_eges_por_registro[revision.registro_id] = revision
        correcciones_por_registro = {}
        correcciones = (
            CorreccionPacsRegistro.objects
            .filter(sesion_contable=self.sesion, registro_id__in=registro_ids)
            .select_related('corregido_por')
            .order_by('registro_id', '-fecha_correccion')
        )
        for correccion in correcciones:
            if correccion.registro_id not in correcciones_por_registro:
                correcciones_por_registro[correccion.registro_id] = correccion
        for registro in registros_alerta:
            revision_auditoria = revisiones_por_registro.get(registro['registro_id'])
            revision_eges = revisiones_eges_por_registro.get(registro['registro_id'])
            registro['revision_auditoria_eco'] = revision_auditoria
            registro['revision_cruce_eges'] = revision_eges
            registro['auditoria_eco_resuelta_por_eges'] = bool(
                not revision_auditoria
                and revision_eges
                and revision_eges.estado in (
                    RevisionCruceEgesRegistro.ESTADO_VALIDADO,
                    RevisionCruceEgesRegistro.ESTADO_DESCARTADO,
                )
            )
            registro['auditoria_eco_requiere_correccion_por_eges'] = bool(
                not revision_auditoria
                and revision_eges
                and revision_eges.estado == RevisionCruceEgesRegistro.ESTADO_REQUIERE_CORRECCION
            )
            registro['correccion_pacs'] = correcciones_por_registro.get(registro['registro_id'])

        registros_alerta = _filtrar_registros_alerta_auditoria_eco(
            registros_alerta,
            estado_revision=estado_revision_actual,
            fecha_desde=fecha_desde_actual,
            fecha_hasta=fecha_hasta_actual,
            ajuste_pacs=ajuste_pacs_actual,
        )

        motivos_disponibles = sorted({
            motivo
            for item in auditoria.get('items', [])
            for registro in item.get('registros_alerta', [])
            for motivo in registro.get('motivos', [])
        })

        context.update({
            'sesion': self.sesion,
            'auditoria': auditoria,
            'registros_alerta': registros_alerta,
            'medico_actual': medico_actual,
            'motivo_actual': motivo_actual,
            'estado_revision_actual': estado_revision_actual,
            'fecha_desde_actual': fecha_desde_actual,
            'fecha_hasta_actual': fecha_hasta_actual,
            'ajuste_pacs_actual': ajuste_pacs_actual,
            'ajuste_pacs_choices': [
                ('CON_AJUSTE', 'Con ajuste PACS'),
                ('SIN_AJUSTE', 'Sin ajuste PACS'),
            ],
            'puede_correccion_masiva_pacs': _puede_accion_masiva_revision_horaria(self.request.user),
            'puede_revision_masiva_eco': _puede_accion_masiva_revision_horaria(self.request.user),
            'estado_revision_choices': [
                ('SIN_REVISAR', 'Sin revisar'),
                (RevisionAuditoriaEcoRegistro.ESTADO_VALIDADO, 'Validado contra PACS'),
                (RevisionAuditoriaEcoRegistro.ESTADO_REQUIERE_CORRECCION, 'Requiere correccion'),
                (RevisionAuditoriaEcoRegistro.ESTADO_DESCARTADO, 'Descartado / no corresponde'),
            ],
            'medicos_alerta': [
                item for item in auditoria.get('items', [])
                if item.get('registros_alerta')
            ],
            'motivos_disponibles': motivos_disponibles,
            'current_path': self.request.get_full_path(),
        })
        return context


class CruceEgesLiquidacionPreviewView(LoginRequiredMixin, UserPassesTestMixin, TemplateView):
    """Preview read-only del cruce entre registros de residencia y un batch EGES."""

    template_name = 'liquidacion/cruce_eges_liquidacion_preview.html'

    def test_func(self):
        return _puede_acceder_panel_administrativo(self.request.user)

    def handle_no_permission(self):
        messages.error(self.request, 'No tienes permisos para revisar el cruce EGES.')
        return redirect('home')

    def dispatch(self, request, *args, **kwargs):
        self.sesion = get_object_or_404(SesionContable, pk=kwargs['pk'])
        return super().dispatch(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        batch_id = (self.request.GET.get('batch') or '').strip()
        batches = ImportBatch.objects.all().order_by('-fecha_importacion')
        batch = None
        preview = None
        if batch_id:
            batch = get_object_or_404(ImportBatch, pk=batch_id)
            preview = construir_preview_cruce_liquidacion_eges(self.sesion, batch, filtros=self.request.GET)
            preview['profesionales'] = _opciones_profesionales_cruce_eges(preview['resultados'])
            preview['resultados_original_total'] = len(preview['resultados'])
            resultados_filtrados = _filtrar_resultados_cruce_eges(preview['resultados'], self.request.GET)
            paginacion = _paginar_resultados_cruce_eges(resultados_filtrados, self.request.GET)
            preview['resultados'] = paginacion['resultados']
            preview['resultados_filtrados_total'] = paginacion['total']
            preview['page_obj'] = paginacion['page_obj']
            preview['paginator'] = paginacion['paginator']

        context.update({
            'sesion': self.sesion,
            'batches': batches,
            'batch_seleccionado': batch,
            'preview': preview,
            'control_eges': resumir_control_eges_sesion(self.sesion),
            'filtros_cruce': {
                'profesional': self.request.GET.get('profesional', ''),
                'estado_cruce': self.request.GET.get('estado_cruce', ''),
                'estado_revision': self.request.GET.get('estado_revision', ''),
                'fecha_desde': self.request.GET.get('fecha_desde', ''),
                'fecha_hasta': self.request.GET.get('fecha_hasta', ''),
                'q': self.request.GET.get('q', ''),
            },
        })
        return context


class CruceEgesProcesarControlView(LoginRequiredMixin, UserPassesTestMixin, View):
    """Consolida el cruce del periodo para usarlo como control operativo."""

    def test_func(self):
        return _puede_acceder_panel_administrativo(self.request.user)

    def handle_no_permission(self):
        messages.error(self.request, 'No tienes permisos para procesar el control EGES.')
        return redirect('home')

    def post(self, request, *args, **kwargs):
        sesion = get_object_or_404(SesionContable, pk=kwargs['pk'])
        batch = get_object_or_404(ImportBatch, pk=request.POST.get('batch'))
        redirect_url = (
            reverse('liquidacion:cruce_eges_liquidacion_preview', kwargs={'pk': sesion.pk})
            + f'?batch={batch.pk}'
        )
        if sesion.estado == 'ABIERTA':
            messages.error(
                request,
                'El control EGES se consolida cuando el periodo ya esta en REVISION o CERRADA.',
            )
            return redirect(redirect_url)
        if sesion.estado in {'FACTURADA', 'PAGADA'}:
            messages.error(request, 'No se puede reprocesar el control de una sesion facturada o pagada.')
            return redirect(redirect_url)

        control = procesar_control_eges_sesion(sesion, batch, request.user)
        pendientes = control.total_advertencias + control.total_manuales
        messages.success(
            request,
            (
                f'Control EGES v{control.version} consolidado: '
                f'{control.total_ok} coincidencia(s) y {pendientes} caso(s) para revisar.'
            ),
        )
        return redirect(redirect_url)


def _snapshot_cruce_eges_item(item):
    return serializar_resultado_control_eges(item)


def _cerrar_revision_eges_por_correccion(sesion, registro, correccion, user):
    ultima_revision = (
        RevisionCruceEgesRegistro.objects
        .filter(sesion_contable=sesion, registro=registro)
        .order_by('-fecha_revision')
        .first()
    )
    if (
        not ultima_revision
        or ultima_revision.estado != RevisionCruceEgesRegistro.ESTADO_REQUIERE_CORRECCION
    ):
        return None

    return RevisionCruceEgesRegistro.objects.create(
        sesion_contable=sesion,
        registro=registro,
        batch_eges=ultima_revision.batch_eges,
        estado=RevisionCruceEgesRegistro.ESTADO_VALIDADO,
        motivos_json=ultima_revision.motivos_json,
        snapshot_json=ultima_revision.snapshot_json,
        observacion=(
            f'Correccion economica #{correccion.pk} aplicada. '
            'Revision EGES cerrada automaticamente.'
        ),
        revisado_por=user,
    )


def _parse_fecha_cruce_eges(valor):
    try:
        return datetime.strptime(valor, '%Y-%m-%d').date() if valor else None
    except ValueError:
        return None


def _filtrar_resultados_cruce_eges(resultados, params):
    profesional = (params.get('profesional') or '').strip()
    estado_cruce = (params.get('estado_cruce') or '').strip()
    estado_revision = (params.get('estado_revision') or '').strip()
    fecha_desde = _parse_fecha_cruce_eges((params.get('fecha_desde') or '').strip())
    fecha_hasta = _parse_fecha_cruce_eges((params.get('fecha_hasta') or '').strip())
    busqueda = (params.get('q') or '').strip().lower()

    filtrados = []
    for item in resultados:
        registro = item['registro']
        revision = item.get('revision_cruce_eges')
        if profesional and str(registro.medico_id) != profesional:
            continue
        if estado_cruce and item['estado'] != estado_cruce:
            continue
        if fecha_desde and registro.fecha_del_informe < fecha_desde:
            continue
        if fecha_hasta and registro.fecha_del_informe > fecha_hasta:
            continue
        if busqueda:
            texto = ' '.join([
                registro.apellido_paciente or '',
                registro.nombre_paciente or '',
                registro.dni_paciente or '',
            ]).lower()
            if busqueda not in texto:
                continue
        if estado_revision == 'SIN_REVISAR' and revision:
            continue
        if estado_revision and estado_revision != 'SIN_REVISAR':
            if not revision or revision.estado != estado_revision:
                continue
        filtrados.append(item)
    return filtrados


def _paginar_resultados_cruce_eges(resultados, params, por_pagina=100):
    paginator = Paginator(resultados, por_pagina)
    page_obj = paginator.get_page(params.get('page') or 1)
    return {
        'resultados': list(page_obj.object_list),
        'page_obj': page_obj,
        'paginator': paginator,
        'total': paginator.count,
    }


def _opciones_profesionales_cruce_eges(resultados):
    opciones = {}
    for item in resultados:
        medico = item['registro'].medico
        opciones[medico.pk] = medico.get_full_name() or medico.username
    return [
        {'id': pk, 'nombre': nombre}
        for pk, nombre in sorted(opciones.items(), key=lambda par: par[1])
    ]


class CruceEgesBulkValidarOkView(LoginRequiredMixin, UserPassesTestMixin, View):
    """Valida automaticamente los OK visibles del cruce EGES, sin impacto economico."""

    def test_func(self):
        return _puede_acceder_panel_administrativo(self.request.user)

    def handle_no_permission(self):
        messages.error(self.request, 'No tienes permisos para validar el cruce EGES.')
        return redirect('home')

    def post(self, request, *args, **kwargs):
        sesion = get_object_or_404(SesionContable, pk=kwargs['pk'])
        batch = get_object_or_404(ImportBatch, pk=request.POST.get('batch'))
        redirect_url = request.POST.get('next') or (
            reverse('liquidacion:cruce_eges_liquidacion_preview', kwargs={'pk': sesion.pk})
            + f'?batch={batch.pk}'
        )
        preview = construir_preview_cruce_liquidacion_eges(sesion, batch, filtros=request.POST)
        visibles = _filtrar_resultados_cruce_eges(preview['resultados'], request.POST)
        visibles = _paginar_resultados_cruce_eges(visibles, request.POST)['resultados']
        creadas = []
        for item in visibles:
            if item['estado'] != 'ok' or item.get('revision_cruce_eges'):
                continue
            creadas.append(RevisionCruceEgesRegistro(
                sesion_contable=sesion,
                registro=item['registro'],
                batch_eges=batch,
                estado=RevisionCruceEgesRegistro.ESTADO_VALIDADO,
                motivos_json=item['motivos'],
                snapshot_json=_snapshot_cruce_eges_item(item),
                observacion='Validado automaticamente desde cruce EGES OK visible.',
                revisado_por=request.user,
            ))
        if creadas:
            RevisionCruceEgesRegistro.objects.bulk_create(creadas)
            messages.success(request, f'Se validaron {len(creadas)} registro(s) OK visibles.')
        else:
            messages.info(request, 'No habia registros OK visibles sin revision para validar.')
        return redirect(redirect_url)


class CruceEgesRegistroResolverView(LoginRequiredMixin, UserPassesTestMixin, View):
    """Marca un resultado del cruce EGES como revisado, sin modificar liquidacion."""

    def test_func(self):
        return _puede_acceder_panel_administrativo(self.request.user)

    def handle_no_permission(self):
        messages.error(self.request, 'No tienes permisos para resolver el cruce EGES.')
        return redirect('home')

    def post(self, request, *args, **kwargs):
        sesion = get_object_or_404(SesionContable, pk=kwargs['pk'])
        registro = get_object_or_404(
            RegistroEstudiosPorMedico.objects.select_related('sesion_contable'),
            pk=kwargs['registro_pk'],
            sesion_contable=sesion,
            anulado=False,
        )
        batch = get_object_or_404(ImportBatch, pk=request.POST.get('batch'))
        redirect_url = request.POST.get('next') or (
            reverse('liquidacion:cruce_eges_liquidacion_preview', kwargs={'pk': sesion.pk})
            + f'?batch={batch.pk}'
        )

        form = RevisionCruceEgesRegistroForm(request.POST)
        if not form.is_valid():
            messages.error(request, 'Debes indicar accion y observacion para resolver el cruce EGES.')
            return redirect(redirect_url)

        preview = construir_preview_cruce_liquidacion_eges(
            sesion,
            batch,
            registro_ids=[registro.pk],
        )
        item = next(
            (resultado for resultado in preview['resultados'] if resultado['registro'].pk == registro.pk),
            None,
        )
        if not item:
            messages.error(request, 'No se encontro el registro dentro del cruce EGES seleccionado.')
            return redirect(redirect_url)

        snapshot = _snapshot_cruce_eges_item(item)

        RevisionCruceEgesRegistro.objects.create(
            sesion_contable=sesion,
            registro=registro,
            batch_eges=batch,
            estado=form.cleaned_data['estado'],
            motivos_json=item['motivos'],
            snapshot_json=snapshot,
            observacion=form.cleaned_data['observacion'],
            revisado_por=request.user,
        )
        messages.success(request, f'Revision EGES registrada para el registro #{registro.pk}.')
        return redirect(redirect_url)


class AuditoriaEcoRegistroResolverView(LoginRequiredMixin, UserPassesTestMixin, View):
    """Marca una alerta ECO como revisada sin modificar el registro."""

    def test_func(self):
        return _puede_acceder_panel_administrativo(self.request.user)

    def handle_no_permission(self):
        messages.error(self.request, 'No tienes permisos para resolver alertas ECO.')
        return redirect('home')

    def post(self, request, *args, **kwargs):
        sesion = get_object_or_404(SesionContable, pk=kwargs['pk'])
        registro = get_object_or_404(
            RegistroEstudiosPorMedico.objects.select_related('sesion_contable'),
            pk=kwargs['registro_pk'],
            sesion_contable=sesion,
            anulado=False,
        )
        form = RevisionAuditoriaEcoRegistroForm(request.POST)
        redirect_url = request.POST.get('next') or reverse('liquidacion:auditoria_eco_sesion', kwargs={'pk': sesion.pk})

        if not form.is_valid():
            messages.error(request, 'Debes indicar estado y observacion para resolver la alerta ECO.')
            return redirect(redirect_url)

        motivos = request.POST.getlist('motivos')
        RevisionAuditoriaEcoRegistro.objects.create(
            sesion_contable=sesion,
            registro=registro,
            estado=form.cleaned_data['estado'],
            motivos_json=motivos,
            observacion=form.cleaned_data['observacion'],
            revisado_por=request.user,
        )
        messages.success(request, f'Revision ECO registrada para el registro #{registro.pk}.')
        return redirect(redirect_url)


class AuditoriaEcoBulkResolverView(LoginRequiredMixin, UserPassesTestMixin, View):
    """Valida o descarta en forma masiva alertas ECO sin impacto economico."""

    def test_func(self):
        return _puede_accion_masiva_revision_horaria(self.request.user)

    def handle_no_permission(self):
        messages.error(
            self.request,
            'Solo superusuarios o jefatura de servicio pueden resolver alertas ECO en forma masiva.',
        )
        return redirect('home')

    def post(self, request, *args, **kwargs):
        sesion = get_object_or_404(SesionContable, pk=kwargs['pk'])
        redirect_url = request.POST.get('next') or reverse('liquidacion:auditoria_eco_sesion', kwargs={'pk': sesion.pk})
        registro_ids = request.POST.getlist('registros')
        registro_choices = [(str(pk), str(pk)) for pk in registro_ids]
        form = RevisionAuditoriaEcoBulkForm(request.POST, registro_choices=registro_choices)

        if not form.is_valid():
            messages.error(request, 'Selecciona registros, accion y observacion para resolver alertas ECO.')
            return redirect(redirect_url)

        seleccionados = [int(pk) for pk in form.cleaned_data['registros']]
        estado = form.cleaned_data['estado']
        observacion = form.cleaned_data['observacion']

        auditoria = auditar_residentes_eco_por_sesion(sesion)
        motivos_por_registro = {}
        for item in auditoria.get('items', []):
            for registro_alerta in item.get('registros_alerta', []):
                registro_id = registro_alerta.get('registro_id')
                if registro_id:
                    motivos_por_registro[registro_id] = registro_alerta.get('motivos', [])

        revisados = set(
            RevisionAuditoriaEcoRegistro.objects
            .filter(sesion_contable=sesion, registro_id__in=seleccionados)
            .values_list('registro_id', flat=True)
        )
        con_correccion = set(
            CorreccionPacsRegistro.objects
            .filter(sesion_contable=sesion, registro_id__in=seleccionados)
            .values_list('registro_id', flat=True)
        )
        registros_validos = list(
            RegistroEstudiosPorMedico.objects
            .filter(pk__in=seleccionados, sesion_contable=sesion, anulado=False)
            .exclude(pk__in=revisados)
            .exclude(pk__in=con_correccion)
            .order_by('pk')
        )

        creadas = [
            RevisionAuditoriaEcoRegistro(
                sesion_contable=sesion,
                registro=registro,
                estado=estado,
                motivos_json=motivos_por_registro.get(registro.pk, []),
                observacion=observacion,
                revisado_por=request.user,
            )
            for registro in registros_validos
        ]
        if creadas:
            RevisionAuditoriaEcoRegistro.objects.bulk_create(creadas)

        omitidos = len(seleccionados) - len(creadas)
        if creadas:
            accion = 'validada(s)' if estado == RevisionAuditoriaEcoRegistro.ESTADO_VALIDADO else 'descartada(s)'
            messages.success(request, f'{len(creadas)} alerta(s) ECO {accion} en forma masiva. Omitidas: {omitidos}.')
        else:
            messages.warning(request, f'No se resolvieron alertas ECO. Omitidas: {omitidos}.')
        return redirect(redirect_url)


class AuditoriaEcoRegistroCorregirView(LoginRequiredMixin, UserPassesTestMixin, View):
    """Aplica un ajuste economico puntual luego de una revision PACS."""

    def test_func(self):
        return _puede_acceder_panel_administrativo(self.request.user)

    def handle_no_permission(self):
        messages.error(self.request, 'No tienes permisos para corregir registros por control PACS.')
        return redirect('home')

    def post(self, request, *args, **kwargs):
        sesion = get_object_or_404(SesionContable, pk=kwargs['pk'])
        form = CorreccionPacsRegistroForm(request.POST)
        redirect_url = request.POST.get('next') or reverse('liquidacion:auditoria_eco_sesion', kwargs={'pk': sesion.pk})

        if sesion.estado in ['FACTURADA', 'PAGADA']:
            messages.error(request, 'No se puede aplicar un ajuste PACS en sesiones FACTURADAS o PAGADAS.')
            return redirect(redirect_url)

        if not form.is_valid():
            messages.error(request, 'Debes indicar horario corregido o monto manual, y una observacion para aplicar el ajuste PACS.')
            return redirect(redirect_url)

        with transaction.atomic():
            registro = get_object_or_404(
                RegistroEstudiosPorMedico.objects.select_for_update(),
                pk=kwargs['registro_pk'],
                sesion_contable=sesion,
                anulado=False,
            )
            revision = (
                RevisionAuditoriaEcoRegistro.objects
                .filter(sesion_contable=sesion, registro=registro)
                .select_related('revisado_por')
                .order_by('-fecha_revision')
                .first()
            )
            if not revision:
                revision_eges = (
                    RevisionCruceEgesRegistro.objects
                    .filter(sesion_contable=sesion, registro=registro)
                    .order_by('-fecha_revision')
                    .first()
                )
                if (
                    revision_eges
                    and revision_eges.estado == RevisionCruceEgesRegistro.ESTADO_REQUIERE_CORRECCION
                ):
                    revision = RevisionAuditoriaEcoRegistro.objects.create(
                        sesion_contable=sesion,
                        registro=registro,
                        estado=RevisionAuditoriaEcoRegistro.ESTADO_REQUIERE_CORRECCION,
                        motivos_json=revision_eges.motivos_json or ['EGES requiere correccion'],
                        observacion=(
                            'Puente automatico desde cruce EGES para aplicar correccion. '
                            f'{revision_eges.observacion}'
                        ),
                        revisado_por=request.user,
                    )

            if not revision or revision.estado != RevisionAuditoriaEcoRegistro.ESTADO_REQUIERE_CORRECCION:
                messages.error(
                    request,
                    'Debe marcarse el registro como "Requiere correccion" antes de aplicar un ajuste PACS.',
                )
                return redirect(redirect_url)

            tipo_correccion = form.cleaned_data['tipo_correccion']
            observacion = form.cleaned_data['observacion']
            monto_anterior = registro.monto_calculado
            horario_anterior = registro.horario
            horario_nuevo = None
            hora_pacs = None

            if tipo_correccion == CorreccionPacsRegistro.TIPO_HORARIO_RECALCULADO:
                horario_nuevo = form.cleaned_data['horario_corregido']
                hora_pacs = form.cleaned_data['hora_pacs']
                registro.horario = horario_nuevo
                monto_nuevo = registro.calcular_monto()
            else:
                monto_nuevo = form.cleaned_data['monto_nuevo']

            if monto_nuevo == monto_anterior and (not horario_nuevo or horario_nuevo == horario_anterior):
                messages.warning(request, 'La correccion coincide con el registro actual. No se aplicaron cambios.')
                return redirect(redirect_url)

            registro.monto_calculado = monto_nuevo
            registro.modificado_por = request.user
            registro.fecha_modificacion = now()
            if horario_nuevo:
                registro.motivo_modificacion = (
                    f'Ajuste por control PACS en auditoria ECO. '
                    f'Horario: {horario_anterior} -> {horario_nuevo}. '
                    f'Hora PACS: {hora_pacs.strftime("%H:%M")}. '
                    f'Monto: ${monto_anterior} -> ${monto_nuevo}. {observacion}'
                )
            else:
                registro.motivo_modificacion = (
                    f'Ajuste por control PACS en auditoria ECO. '
                    f'Monto: ${monto_anterior} -> ${monto_nuevo}. {observacion}'
                )
            update_fields = [
                'monto_calculado',
                'modificado_por',
                'fecha_modificacion',
                'motivo_modificacion',
            ]
            if horario_nuevo:
                update_fields.append('horario')
            registro.save(update_fields=update_fields)
            correccion = CorreccionPacsRegistro.objects.create(
                sesion_contable=sesion,
                registro=registro,
                revision_auditoria_eco=revision,
                tipo_correccion=tipo_correccion,
                horario_anterior=horario_anterior if horario_nuevo else None,
                horario_nuevo=horario_nuevo,
                hora_pacs=hora_pacs,
                monto_anterior=monto_anterior,
                monto_nuevo=monto_nuevo,
                observacion=observacion,
                corregido_por=request.user,
            )
            _cerrar_revision_eges_por_correccion(
                sesion,
                registro,
                correccion,
                request.user,
            )
            RevisionAuditoriaEcoRegistro.objects.create(
                sesion_contable=sesion,
                registro=registro,
                estado=RevisionAuditoriaEcoRegistro.ESTADO_VALIDADO,
                motivos_json=revision.motivos_json,
                observacion=(
                    f'Correccion PACS aplicada. '
                    f'Monto: ${monto_anterior} -> ${monto_nuevo}. {observacion}'
                ),
                revisado_por=request.user,
            )

        messages.success(request, f'Ajuste PACS aplicado al registro #{registro.pk}. Monto: ${monto_anterior} -> ${monto_nuevo}.')
        return redirect(redirect_url)


class AuditoriaEcoCorreccionPacsBulkView(LoginRequiredMixin, UserPassesTestMixin, View):
    """Aplica una nueva correccion auditada sobre ajustes PACS ya existentes."""

    def test_func(self):
        return _puede_accion_masiva_revision_horaria(self.request.user)

    def handle_no_permission(self):
        messages.error(
            self.request,
            'Solo superusuarios o jefatura de servicio pueden corregir ajustes PACS en forma masiva.',
        )
        return redirect('home')

    def post(self, request, *args, **kwargs):
        sesion = get_object_or_404(SesionContable, pk=kwargs['pk'])
        redirect_url = request.POST.get('next') or reverse('liquidacion:auditoria_eco_sesion', kwargs={'pk': sesion.pk})

        if sesion.estado in ['FACTURADA', 'PAGADA']:
            messages.error(request, 'No se pueden corregir ajustes PACS en sesiones FACTURADAS o PAGADAS.')
            return redirect(redirect_url)

        registro_ids = request.POST.getlist('registros')
        registro_choices = [(str(pk), str(pk)) for pk in registro_ids]
        form = CorreccionPacsAplicadaBulkForm(request.POST, registro_choices=registro_choices)
        if not form.is_valid():
            messages.error(request, 'Selecciona registros, horario corregido y observacion para aplicar la correccion masiva.')
            return redirect(redirect_url)

        horario_nuevo = form.cleaned_data['horario_corregido']
        observacion = form.cleaned_data['observacion']
        seleccionados = [int(pk) for pk in form.cleaned_data['registros']]
        aplicados = 0
        omitidos = 0

        with transaction.atomic():
            registros = list(
                RegistroEstudiosPorMedico.objects
                .select_for_update()
                .filter(pk__in=seleccionados, sesion_contable=sesion, anulado=False)
                .order_by('pk')
            )
            registros_por_id = {registro.pk: registro for registro in registros}
            correcciones_previas = (
                CorreccionPacsRegistro.objects
                .filter(sesion_contable=sesion, registro_id__in=registros_por_id)
                .order_by('registro_id', '-fecha_correccion')
            )
            correcciones_por_registro = {}
            for correccion in correcciones_previas:
                correcciones_por_registro.setdefault(correccion.registro_id, correccion)

            revisiones = (
                RevisionAuditoriaEcoRegistro.objects
                .filter(sesion_contable=sesion, registro_id__in=registros_por_id)
                .order_by('registro_id', '-fecha_revision')
            )
            revisiones_por_registro = {}
            for revision in revisiones:
                revisiones_por_registro.setdefault(revision.registro_id, revision)

            for registro_id in seleccionados:
                registro = registros_por_id.get(registro_id)
                correccion_previa = correcciones_por_registro.get(registro_id)
                if not registro or not correccion_previa:
                    omitidos += 1
                    continue

                monto_anterior = registro.monto_calculado
                horario_anterior = registro.horario
                registro.horario = horario_nuevo
                monto_nuevo = registro.calcular_monto()

                if monto_nuevo == monto_anterior and horario_nuevo == horario_anterior:
                    omitidos += 1
                    continue

                registro.monto_calculado = monto_nuevo
                registro.modificado_por = request.user
                registro.fecha_modificacion = now()
                registro.motivo_modificacion = (
                    f'Nueva correccion sobre ajuste PACS aplicado. '
                    f'Horario: {horario_anterior} -> {horario_nuevo}. '
                    f'Monto: ${monto_anterior} -> ${monto_nuevo}. {observacion}'
                )
                registro.save(update_fields=[
                    'horario',
                    'monto_calculado',
                    'modificado_por',
                    'fecha_modificacion',
                    'motivo_modificacion',
                ])

                revision = revisiones_por_registro.get(registro_id)
                nueva_correccion = CorreccionPacsRegistro.objects.create(
                    sesion_contable=sesion,
                    registro=registro,
                    revision_auditoria_eco=revision,
                    tipo_correccion=CorreccionPacsRegistro.TIPO_HORARIO_RECALCULADO,
                    horario_anterior=horario_anterior,
                    horario_nuevo=horario_nuevo,
                    monto_anterior=monto_anterior,
                    monto_nuevo=monto_nuevo,
                    observacion=observacion,
                    corregido_por=request.user,
                )
                RevisionAuditoriaEcoRegistro.objects.create(
                    sesion_contable=sesion,
                    registro=registro,
                    estado=RevisionAuditoriaEcoRegistro.ESTADO_VALIDADO,
                    motivos_json=revision.motivos_json if revision else [],
                    observacion=(
                        f'Nueva correccion sobre ajuste PACS #{nueva_correccion.pk}. '
                        f'Horario: {horario_anterior} -> {horario_nuevo}. '
                        f'Monto: ${monto_anterior} -> ${monto_nuevo}. {observacion}'
                    ),
                    revisado_por=request.user,
                )
                _cerrar_revision_eges_por_correccion(
                    sesion,
                    registro,
                    nueva_correccion,
                    request.user,
                )
                aplicados += 1

        if aplicados:
            messages.success(request, f'Se aplicaron {aplicados} nueva(s) correccion(es) PACS. Omitidos: {omitidos}.')
        else:
            messages.warning(request, f'No se aplicaron correcciones. Omitidos: {omitidos}.')
        return redirect(redirect_url)


class SolicitudRevisionHorarioResolverView(LoginRequiredMixin, UserPassesTestMixin, View):
    """Resolución administrativa B1 (solo POST, sin impacto económico)."""

    def test_func(self):
        return _puede_acceder_panel_administrativo(self.request.user)

    def handle_no_permission(self):
        messages.error(
            self.request,
            '❌ No tienes permisos para resolver solicitudes de revisión de horario.',
        )
        return redirect('home')

    def post(self, request, *args, **kwargs):
        solicitud = get_object_or_404(
            SolicitudRevisionHorarioRegistro.objects.select_related('registro__sesion_contable', 'solicitado_por'),
            pk=kwargs['pk'],
        )

        if solicitud.estado != SolicitudRevisionHorarioRegistro.ESTADO_PENDIENTE:
            messages.error(request, '❌ Solo se puede resolver una solicitud en estado PENDIENTE.')
            return redirect('liquidacion:solicitudes_revision_horario_detalle', pk=solicitud.pk)

        sesion = solicitud.registro.sesion_contable
        if sesion and sesion.estado in ['FACTURADA', 'PAGADA']:
            messages.error(
                request,
                '❌ No se puede resolver la solicitud porque la sesión contable está FACTURADA o PAGADA.',
            )
            return redirect('liquidacion:solicitudes_revision_horario_detalle', pk=solicitud.pk)

        if solicitud.solicitado_por_id == request.user.id:
            messages.error(
                request,
                '❌ No se permite auto-revisión: el solicitante no puede resolver su propia solicitud.',
            )
            return redirect('liquidacion:solicitudes_revision_horario_detalle', pk=solicitud.pk)

        form = SolicitudRevisionHorarioResolucionForm(request.POST)
        if not form.is_valid():
            messages.error(request, '❌ Debes indicar una decisión válida para resolver la solicitud.')
            return redirect('liquidacion:solicitudes_revision_horario_detalle', pk=solicitud.pk)

        decision = form.cleaned_data['decision']
        observacion_revision = form.cleaned_data['observacion_revision']
        estado_nuevo = (
            SolicitudRevisionHorarioRegistro.ESTADO_APROBADA
            if decision == SolicitudRevisionHorarioResolucionForm.DECISION_APROBAR
            else SolicitudRevisionHorarioRegistro.ESTADO_RECHAZADA
        )

        actualizado = SolicitudRevisionHorarioRegistro.objects.filter(
            pk=solicitud.pk,
            estado=SolicitudRevisionHorarioRegistro.ESTADO_PENDIENTE,
        ).update(
            estado=estado_nuevo,
            revisado_por=request.user,
            fecha_revision=now(),
            observacion_revision=observacion_revision,
        )

        if actualizado == 0:
            messages.error(request, '❌ La solicitud ya no está pendiente y no puede resolverse nuevamente.')
            return redirect('liquidacion:solicitudes_revision_horario_detalle', pk=solicitud.pk)

        messages.success(request, f'✅ Solicitud {estado_nuevo.lower()} correctamente.')
        return redirect('liquidacion:solicitudes_revision_horario_detalle', pk=solicitud.pk)


def _aplicar_solicitud_revision_horario_b2(solicitud_pk, user, observacion_aplicacion=''):
    """Aplica economicamente una solicitud aprobada. Retorna dict con ok/motivo."""
    with transaction.atomic():
        try:
            solicitud = SolicitudRevisionHorarioRegistro.objects.select_for_update().get(pk=solicitud_pk)
        except SolicitudRevisionHorarioRegistro.DoesNotExist:
            return {'ok': False, 'solicitud_id': solicitud_pk, 'motivo': 'Solicitud inexistente.'}

        try:
            registro = RegistroEstudiosPorMedico.objects.select_for_update().get(pk=solicitud.registro_id)
        except RegistroEstudiosPorMedico.DoesNotExist:
            return {'ok': False, 'solicitud_id': solicitud.pk, 'motivo': 'Registro asociado inexistente.'}

        if registro.anulado:
            return {'ok': False, 'solicitud_id': solicitud.pk, 'motivo': 'El registro esta anulado.'}

        if solicitud.estado != SolicitudRevisionHorarioRegistro.ESTADO_APROBADA:
            return {'ok': False, 'solicitud_id': solicitud.pk, 'motivo': 'No esta APROBADA.'}

        if solicitud.fecha_aplicacion:
            return {'ok': False, 'solicitud_id': solicitud.pk, 'motivo': 'Ya fue aplicada.'}

        sesion = None
        if registro.sesion_contable_id:
            sesion = SesionContable.objects.get(pk=registro.sesion_contable_id)
        if not sesion or sesion.estado not in ['ABIERTA', 'REVISION']:
            return {
                'ok': False,
                'solicitud_id': solicitud.pk,
                'motivo': 'La sesion no esta ABIERTA o REVISION.',
            }

        horario_anterior = registro.horario
        monto_anterior = registro.monto_calculado
        horario_aplicado = solicitud.horario_solicitado

        registro.horario = horario_aplicado
        monto_aplicado = registro.calcular_monto()
        fecha_aplicacion = now()

        motivo_modificacion = (
            f"Correccion de horario por solicitud de revision #{solicitud.pk} aprobada administrativamente. "
            f"Horario: {horario_anterior} -> {horario_aplicado}. "
            f"Monto: ${monto_anterior} -> ${monto_aplicado}."
        )

        actualizados_registro = RegistroEstudiosPorMedico.objects.filter(pk=registro.pk).update(
            horario=horario_aplicado,
            monto_calculado=monto_aplicado,
            modificado_por=user,
            fecha_modificacion=fecha_aplicacion,
            motivo_modificacion=motivo_modificacion,
        )

        if actualizados_registro == 0:
            return {'ok': False, 'solicitud_id': solicitud.pk, 'motivo': 'No se pudo actualizar el registro.'}

        solicitud.horario_anterior = horario_anterior
        solicitud.horario_aplicado = horario_aplicado
        solicitud.monto_anterior = monto_anterior
        solicitud.monto_aplicado = monto_aplicado
        solicitud.aplicado_por = user
        solicitud.fecha_aplicacion = fecha_aplicacion
        solicitud.observacion_aplicacion = observacion_aplicacion
        solicitud.save(update_fields=[
            'horario_anterior',
            'horario_aplicado',
            'monto_anterior',
            'monto_aplicado',
            'aplicado_por',
            'fecha_aplicacion',
            'observacion_aplicacion',
        ])

    return {'ok': True, 'solicitud_id': solicitud_pk, 'motivo': ''}


class SolicitudRevisionHorarioAplicarView(LoginRequiredMixin, UserPassesTestMixin, View):
    """Aplicación económica B2 sobre solicitud APROBADA."""

    def test_func(self):
        return _puede_acceder_panel_administrativo(self.request.user)

    def handle_no_permission(self):
        messages.error(
            self.request,
            '❌ No tienes permisos para aplicar solicitudes de revisión de horario.',
        )
        return redirect('home')

    def post(self, request, *args, **kwargs):
        solicitud_pk = kwargs['pk']
        form = SolicitudRevisionHorarioAplicarForm(request.POST)
        if not form.is_valid():
            messages.error(request, 'Datos invalidos para aplicar la solicitud.')
            return redirect('liquidacion:solicitudes_revision_horario_detalle', pk=solicitud_pk)

        resultado = _aplicar_solicitud_revision_horario_b2(
            solicitud_pk,
            request.user,
            form.cleaned_data['observacion_aplicacion'],
        )
        if not resultado['ok']:
            messages.error(request, resultado['motivo'])
        else:
            messages.success(request, 'Aplicacion economica realizada correctamente.')
        return redirect('liquidacion:solicitudes_revision_horario_detalle', pk=solicitud_pk)


class SolicitudRevisionHorarioBulkActionView(LoginRequiredMixin, UserPassesTestMixin, View):
    """B4a/B4b: acciones masivas autorizadas sobre solicitudes seleccionadas."""

    def test_func(self):
        return _puede_accion_masiva_revision_horaria(self.request.user)

    def handle_no_permission(self):
        messages.error(
            self.request,
            'Solo superusuarios o jefatura de servicio pueden ejecutar acciones masivas de revision horaria.',
        )
        return redirect('liquidacion:solicitudes_revision_horario_list')

    def post(self, request, *args, **kwargs):
        form = SolicitudRevisionHorarioBulkActionForm(request.POST)
        if not form.is_valid():
            messages.error(request, 'Selecciona solicitudes y una accion masiva valida.')
            return redirect('liquidacion:solicitudes_revision_horario_list')

        solicitud_ids = [int(pk) for pk in form.cleaned_data['solicitudes']]
        accion = form.cleaned_data['accion']
        observacion = form.cleaned_data['observacion']

        if accion == SolicitudRevisionHorarioBulkActionForm.ACCION_APROBAR:
            aprobadas = 0
            omitidas = []
            for solicitud_id in solicitud_ids:
                with transaction.atomic():
                    try:
                        solicitud = SolicitudRevisionHorarioRegistro.objects.select_for_update().get(pk=solicitud_id)
                    except SolicitudRevisionHorarioRegistro.DoesNotExist:
                        omitidas.append(f'#{solicitud_id}: inexistente')
                        continue

                    if solicitud.estado != SolicitudRevisionHorarioRegistro.ESTADO_PENDIENTE:
                        omitidas.append(f'#{solicitud.pk}: no esta PENDIENTE')
                        continue

                    registro = RegistroEstudiosPorMedico.objects.select_related('sesion_contable').get(
                        pk=solicitud.registro_id,
                    )
                    if registro.anulado:
                        omitidas.append(f'#{solicitud.pk}: registro anulado')
                        continue
                    sesion = registro.sesion_contable
                    if sesion and sesion.estado in ['FACTURADA', 'PAGADA']:
                        omitidas.append(f'#{solicitud.pk}: sesion {sesion.estado}')
                        continue

                    if solicitud.solicitado_por_id == request.user.id:
                        omitidas.append(f'#{solicitud.pk}: auto-revision no permitida')
                        continue

                    solicitud.estado = SolicitudRevisionHorarioRegistro.ESTADO_APROBADA
                    solicitud.revisado_por = request.user
                    solicitud.fecha_revision = now()
                    solicitud.observacion_revision = observacion
                    solicitud.save(update_fields=[
                        'estado',
                        'revisado_por',
                        'fecha_revision',
                        'observacion_revision',
                    ])
                    aprobadas += 1

            messages.success(request, f'B4a: {aprobadas} solicitud(es) aprobada(s).')
            if omitidas:
                messages.warning(request, f'Omitidas: {" | ".join(omitidas[:5])}')
            return redirect(
                f"{reverse('liquidacion:solicitudes_revision_horario_list')}"
                f"?estado={SolicitudRevisionHorarioRegistro.ESTADO_APROBADA}&pendiente_aplicacion=1"
            )

        aplicadas = 0
        omitidas = []
        for solicitud_id in solicitud_ids:
            resultado = _aplicar_solicitud_revision_horario_b2(
                solicitud_id,
                request.user,
                observacion,
            )
            if resultado['ok']:
                aplicadas += 1
            else:
                omitidas.append(f"#{solicitud_id}: {resultado['motivo']}")

        messages.success(request, f'B4b: {aplicadas} solicitud(es) aplicada(s) economicamente.')
        if omitidas:
            messages.warning(request, f'Omitidas: {" | ".join(omitidas[:5])}')
        return redirect('liquidacion:solicitudes_revision_horario_list')


class SolicitudRevisionHorarioRecalcularAplicacionView(LoginRequiredMixin, UserPassesTestMixin, View):
    """Recalculo puntual B3 de una solicitud ya aplicada."""

    def test_func(self):
        return _puede_acceder_panel_administrativo(self.request.user)

    def handle_no_permission(self):
        messages.error(
            self.request,
            'No tienes permisos para recalcular solicitudes de revision de horario.',
        )
        return redirect('home')

    def post(self, request, *args, **kwargs):
        solicitud_pk = kwargs['pk']
        form = SolicitudRevisionHorarioRecalcularAplicacionForm(request.POST)
        if not form.is_valid():
            messages.error(request, 'Datos invalidos para recalcular la solicitud.')
            return redirect('liquidacion:solicitudes_revision_horario_detalle', pk=solicitud_pk)

        with transaction.atomic():
            solicitud = get_object_or_404(
                SolicitudRevisionHorarioRegistro.objects.select_for_update(),
                pk=solicitud_pk,
            )
            registro = RegistroEstudiosPorMedico.objects.select_for_update().get(
                pk=solicitud.registro_id,
            )

            if registro.anulado:
                messages.error(request, 'No se puede recalcular una solicitud de un registro anulado.')
                return redirect('liquidacion:solicitudes_revision_horario_detalle', pk=solicitud.pk)

            if solicitud.estado != SolicitudRevisionHorarioRegistro.ESTADO_APROBADA:
                messages.error(request, 'Solo se puede recalcular una solicitud en estado APROBADA.')
                return redirect('liquidacion:solicitudes_revision_horario_detalle', pk=solicitud.pk)

            if not solicitud.fecha_aplicacion:
                messages.error(request, 'Solo se puede recalcular una solicitud ya aplicada.')
                return redirect('liquidacion:solicitudes_revision_horario_detalle', pk=solicitud.pk)

            if not solicitud.horario_aplicado:
                messages.error(request, 'La solicitud aplicada no tiene horario aplicado registrado.')
                return redirect('liquidacion:solicitudes_revision_horario_detalle', pk=solicitud.pk)

            sesion = None
            if registro.sesion_contable_id:
                sesion = SesionContable.objects.get(pk=registro.sesion_contable_id)
            if not sesion or sesion.estado not in ['ABIERTA', 'REVISION']:
                messages.error(
                    request,
                    'Solo se puede recalcular en sesiones ABIERTA o REVISION. En CERRADA, FACTURADA o PAGADA esta bloqueado.',
                )
                return redirect('liquidacion:solicitudes_revision_horario_detalle', pk=solicitud.pk)

            monto_anterior = registro.monto_calculado
            monto_aplicado_anterior = solicitud.monto_aplicado
            horario_aplicado = solicitud.horario_aplicado

            registro.horario = horario_aplicado
            monto_nuevo = registro.calcular_monto()

            if monto_nuevo == monto_anterior:
                messages.info(request, 'El monto recalculado coincide con el monto actual.')
                return redirect('liquidacion:solicitudes_revision_horario_detalle', pk=solicitud.pk)

            fecha_recalculo = now()
            motivo_modificacion = (
                f"Recalculo puntual de solicitud de revisión #{solicitud.pk} con reglas vigentes. "
                f"Monto: ${monto_anterior} → ${monto_nuevo}."
            )

            RegistroEstudiosPorMedico.objects.filter(pk=registro.pk).update(
                horario=horario_aplicado,
                monto_calculado=monto_nuevo,
                modificado_por=request.user,
                fecha_modificacion=fecha_recalculo,
                motivo_modificacion=motivo_modificacion,
            )

            solicitud.monto_aplicado = monto_nuevo
            solicitud.save(update_fields=['monto_aplicado'])

            HistorialRecalculoSolicitudRevisionHorario.objects.create(
                solicitud=solicitud,
                registro=registro,
                recalculado_por=request.user,
                horario_usado=horario_aplicado,
                monto_registro_anterior=monto_anterior,
                monto_aplicado_anterior=monto_aplicado_anterior,
                monto_recalculado=monto_nuevo,
                observacion=form.cleaned_data['observacion'],
                motivo_sistema=motivo_modificacion,
            )

        messages.success(request, 'Recalculo puntual realizado correctamente.')
        return redirect('liquidacion:solicitudes_revision_horario_detalle', pk=solicitud.pk)
# ===== VISTAS REGULARES (Requieren Login) =====
from django.utils.http import urlencode
from django.utils.safestring import mark_safe


def _puede_acceder_panel_medico(user):
    return user.is_superuser or user.es_medico()


LIQUIDAR_COMO_EXTRA_RESIDENCIA_SESSION_KEY = (
    'liquidacion_liquidar_como_extra_residencia_default'
)


def _puede_persistir_extra_residencia_en_sesion(user):
    return getattr(user, 'rol', None) in ROLES_LIQUIDAR_COMO_EXTRA_RESIDENCIA


def _guardar_extra_residencia_en_sesion(request, form):
    if (
        _puede_persistir_extra_residencia_en_sesion(request.user)
        and 'liquidar_como_extra_residencia' in form.cleaned_data
    ):
        request.session[LIQUIDAR_COMO_EXTRA_RESIDENCIA_SESSION_KEY] = bool(
            form.cleaned_data['liquidar_como_extra_residencia']
        )


def _puede_acceder_guardia_pasiva(user):
    return user.rol in ['jefe_residentes', 'instructor_residentes']


def _puede_editar_monto_guardia(user):
    return user.is_superuser or user.rol in ['administrativo', 'jefe_servicio']


def _mes_nombre(mes):
    meses = ['', 'Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio',
             'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre']
    return meses[mes] if 1 <= mes <= 12 else str(mes)

class EstudiosCreateView(LoginRequiredMixin, UserPassesTestMixin, SuccessMessageMixin, CreateView):
    model = Estudios
    form_class = EstudiosAdminForm
    template_name = 'liquidacion/estudios_form.html'
    success_url = reverse_lazy('liquidacion:estudios_list')
    success_message = "El estudio fue registrado exitosamente"

    def test_func(self):
        return _puede_acceder_panel_administrativo(self.request.user)


class EstudiosUpdateView(LoginRequiredMixin, UserPassesTestMixin, SuccessMessageMixin, UpdateView):
    model = Estudios
    form_class = EstudiosAdminForm
    template_name = 'liquidacion/estudios_form.html'
    success_url = reverse_lazy('liquidacion:estudios_list')
    success_message = "El estudio fue actualizado exitosamente"

    def test_func(self):
        return _puede_acceder_panel_administrativo(self.request.user)

    def get_success_url(self):
        volver_sesion = _volver_sesion_url(self.request)
        if volver_sesion:
            return volver_sesion
        return super().get_success_url()

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['volver_sesion_url'] = _volver_sesion_url(self.request)
        return context


class EstudiosListView(LoginRequiredMixin, UserPassesTestMixin, ListView):
    model = Estudios
    template_name = 'liquidacion/estudios_list.html'
    context_object_name = 'estudios'

    def test_func(self):
        return _puede_acceder_panel_administrativo(self.request.user)

    def get_queryset(self):
        queryset = Estudios.objects.select_related('grupo_tarifario').order_by('tipo', 'nombre')
        if self.request.GET.get('sin_grupo') == '1':
            queryset = queryset.filter(grupo_tarifario__isnull=True)
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['filtro_sin_grupo_activo'] = self.request.GET.get('sin_grupo') == '1'
        context['cantidad_sin_grupo'] = Estudios.objects.filter(grupo_tarifario__isnull=True).count()
        return context

class RegistroEstudiosPorMedicoCreateView(LoginRequiredMixin, SuccessMessageMixin, CreateView):
    """
    Vista para registro de prácticas médicas - Liquidación v2.0
    Incluye validaciones de sesión contable y cálculo automático de montos
    """
    model = RegistroEstudiosPorMedico
    form_class = RegistroEstudiosPorMedicoCreateViewForm
    template_name = 'liquidacion/registroestudios_form_tailwind.html'
    success_url = reverse_lazy('liquidacion:registroestudios_nuevo')
    success_message = "✅ Registro guardado exitosamente"

    def dispatch(self, request, *args, **kwargs):
        # Validar que sea médico (por rol, no por grupo)
        if not _puede_acceder_panel_medico(request.user):
            messages.warning(request, "No tienes permiso para acceder a esta sección.")
            return redirect('home')
        return super().dispatch(request, *args, **kwargs)
    
    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user  # Pasar usuario al form para lógica condicional
        if _puede_persistir_extra_residencia_en_sesion(self.request.user):
            session = self.request.session
            if LIQUIDAR_COMO_EXTRA_RESIDENCIA_SESSION_KEY in session:
                initial = kwargs.setdefault('initial', {})
                initial['liquidar_como_extra_residencia'] = bool(
                    session[LIQUIDAR_COMO_EXTRA_RESIDENCIA_SESSION_KEY]
                )
        return kwargs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        user = self.request.user
        today = datetime.today()

        # Obtener o crear sesión contable del mes actual
        sesion, created = SesionContable.objects.get_or_create(
            mes=today.month,
            año=today.year,
            defaults={'estado': 'ABIERTA'}
        )
        context['sesion_contable'] = sesion
        context['puede_registrar'] = sesion.puede_registrar_practicas(user)

        # Tipo de estudio seleccionado (para filtrar estudios con JS)
        tipo_estudio_seleccionado = self.request.POST.get('tipo_estudio', '')
        if not tipo_estudio_seleccionado:
            ultimo_registro = RegistroEstudiosPorMedico.objects.filter(medico=user, anulado=False).order_by('-fecha_registro').first()
            if ultimo_registro and ultimo_registro.estudio.exists():
                tipo_estudio_seleccionado = ultimo_registro.estudio.first().tipo

        context['tipo_estudio_seleccionado'] = tipo_estudio_seleccionado
        
        # Serializar estudios para JS (convertir Decimals a string)
        estudios_data = []
        for estudio in Estudios.objects.filter(activo=True).values(
            'id', 'nombre', 'tipo', 'codigo', 'precio_cober', 'precio_otras_os',
            'precio_unico', 'conteo_regiones_default', 'tiene_contexto_ubicacion'
        ):
            if (not user.is_superuser) and user.rol != 'cardiologo' and es_estudio_cardiologico(
                estudio['tipo'],
                estudio['nombre'],
                estudio['codigo'],
            ):
                continue
            estudio_dict = dict(estudio)
            # Convertir Decimals a string para JSON
            estudio_dict['precio_cober'] = str(estudio_dict['precio_cober'])
            estudio_dict['precio_otras_os'] = str(estudio_dict['precio_otras_os'])
            estudio_dict['es_cardiologico'] = es_estudio_cardiologico(
                estudio_dict['tipo'],
                estudio_dict['nombre'],
                estudio_dict['codigo'],
            )
            estudio_dict['contextos_disponibles'] = contextos_disponibles_para_estudio(
                estudio_dict['tipo'],
                estudio_dict['nombre'],
                estudio_dict['codigo'],
            )
            estudios_data.append(estudio_dict)
        
        context['estudios'] = json.dumps(estudios_data)
        
        # Registros de hoy con prefetch para mostrar desglose y monto persistido
        registros = list(
            RegistroEstudiosPorMedico.objects.filter(
                medico=user,
                sesion_contable=sesion,
                fecha_registro__date=now().date(),
                anulado=False,
            ).prefetch_related('registroestudio_set__estudio').order_by('-fecha_registro')
        )

        puede_admin = puede_ver_desglose_administrativo(user)
        for registro in registros:
            registro.detalle_monto = (
                registro.get_desglose_monto_administrativo()
                if puede_admin
                else registro.get_desglose_monto_simple()
            )

        context['registros'] = registros
        context['puede_ver_desglose_admin'] = puede_admin
        
        # Calcular totales del mes
        total_regiones_mes = sum(reg.cantidad_regiones for reg in registros)
        total_monto_mes = sum(reg.monto_calculado for reg in registros)
        
        context['total_regiones_mes'] = total_regiones_mes
        context['total_monto_mes'] = total_monto_mes
        context['total_practicas_mes'] = len(registros)
        
        # Información del médico
        context['es_staff'] = user.rol in ['medico_staff', 'jefe_servicio', 'cardiologo']
        context['es_cardiologo'] = user.is_superuser or user.rol == 'cardiologo'
        context['trabaja_remoto'] = user.trabaja_remoto

        return context

    @transaction.atomic
    def form_valid(self, form):
        user = self.request.user
        
        # Validar que la sesión permita registrar
        sesion, created = SesionContable.objects.get_or_create(
            mes=form.instance.fecha_del_informe.month,
            año=form.instance.fecha_del_informe.year,
            defaults={'estado': 'ABIERTA'}
        )
        
        if not sesion.puede_registrar_practicas(user):
            messages.error(
                self.request,
                f"❌ La sesión de {_mes_nombre(sesion.mes)} {sesion.año} está en estado "
                f"{sesion.get_estado_display()}. No puedes registrar prácticas."
            )
            return redirect(self.success_url)
        
        # Verificar duplicados recientes (últimos 5 minutos)
        from django.utils import timezone
        from datetime import timedelta
        
        dni_paciente = form.cleaned_data['dni_paciente']
        fecha_informe = form.cleaned_data['fecha_del_informe']
        estudios_seleccionados = form.cleaned_data['estudio']

        # Regla de acceso: solo cardiólogos pueden registrar estudios cardiológicos
        if (not user.is_superuser) and user.rol != 'cardiologo':
            estudios_bloqueados = [
                est.nombre for est in estudios_seleccionados
                if es_estudio_cardiologico(est.tipo, est.nombre, est.codigo)
            ]
            if estudios_bloqueados:
                form.add_error(
                    'estudio',
                    'No tienes permisos para registrar estudios cardiológicos. '
                    f"Detectados: {', '.join(estudios_bloqueados)}"
                )
                return self.form_invalid(form)

        hace_5_minutos = timezone.now() - timedelta(minutes=5)
        
        # Buscar registros recientes del mismo médico, paciente y fecha
        registros_recientes = RegistroEstudiosPorMedico.objects.filter(
            medico=user,
            dni_paciente=dni_paciente,
            fecha_del_informe=fecha_informe,
            fecha_registro__gte=hace_5_minutos,
            anulado=False,
        )
        
        # Verificar si alguno tiene los mismos estudios
        for registro in registros_recientes:
            estudios_registro = set(registro.estudio.values_list('id', flat=True))
            estudios_nuevos = set(est.id for est in estudios_seleccionados)
            if estudios_registro == estudios_nuevos:
                messages.warning(
                    self.request, 
                    f"⚠️ Ya registraste estos mismos estudios hace menos de 5 minutos. "
                    f"Si realmente necesitas crear otro registro, espera unos minutos."
                )
                return redirect(self.success_url)
        
        # Guardar el registro
        self.object = form.save(commit=False)
        self.object.medico = user
        self.object.sesion_contable = sesion
        if user.rol not in ROLES_LIQUIDAR_COMO_EXTRA_RESIDENCIA:
            self.object.liquidar_como_extra_residencia = False
        self.object.save()
        
        # v3.1 - Marzo 2026: NO usar save_m2m() con through model
        # Crear instancias de RegistroEstudio manualmente con cantidades
        
        # Leer cantidades de estudios desde el POST
        cantidades_estudios = {}
        for key in self.request.POST:
            if key.startswith('cantidad_estudio_'):
                estudio_id = int(key.replace('cantidad_estudio_', ''))
                cantidad = int(self.request.POST[key])
                cantidades_estudios[estudio_id] = cantidad
        
        # Importar el modelo intermedio
        from liquidacion.models import RegistroEstudio
        
        # Leer contextos de ubicación desde el POST (para estudios Doppler/ECOCAR)
        contextos_estudios = {}
        for key in self.request.POST:
            if key.startswith('contexto_estudio_'):
                estudio_id = int(key.replace('contexto_estudio_', ''))
                contextos_estudios[estudio_id] = self.request.POST[key]

        # Crear las relaciones en la tabla intermedia con cantidades y contexto
        for estudio in estudios_seleccionados:
            cantidad = cantidades_estudios.get(estudio.id, 1)  # Default = 1 si no está en el dict
            contextos_validos = contextos_disponibles_para_estudio(estudio.tipo, estudio.nombre, estudio.codigo)
            contexto = contextos_estudios.get(estudio.id, contextos_validos[0] if contextos_validos else 'SERVICIO')
            if contexto not in contextos_validos:
                contexto = contextos_validos[0] if contextos_validos else 'SERVICIO'
            RegistroEstudio.objects.create(
                registro=self.object,
                estudio=estudio,
                cantidad=cantidad,
                contexto=contexto,
            )
        
        # Calcular cantidad de regiones con las cantidades especificadas
        total_regiones = 0
        for estudio in estudios_seleccionados:
            cantidad = cantidades_estudios.get(estudio.id, 1)
            total_regiones += (estudio.conteo_regiones_default * cantidad)
        
        self.object.cantidad_regiones = total_regiones

        # Fuente canónica para residencia+ECO general real: clasificación explícita post-M2M.
        if (
            user.rol in ROLES_LIQUIDAR_COMO_EXTRA_RESIDENCIA
            and self.object.liquidar_como_extra_residencia
        ):
            self.object.horario = 'EXTRA'
        elif user.rol in ROLES_RESIDENCIA:
            tiene_eco_general = any(
                es_eco_general_real_estudio(est)
                for est in estudios_seleccionados
            )
            if tiene_eco_general:
                nuevo_horario = clasificar_horario_residencia_por_proxy(
                    rol=user.rol,
                    fecha_registro=self.object.fecha_registro,
                    tiene_eco_general=True,
                    fecha_practica=self.object.fecha_del_informe,
                )
                self.object.horario = nuevo_horario or 'NA'
            else:
                self.object.horario = 'NA'
        
        # v3.1: Calcular monto usando método unificado del modelo (lee cantidades de RegistroEstudio)
        total_monto = self.object.calcular_monto()
        self.object.monto_calculado = total_monto
        
        # También guardar campos de bonus urgencia que vienen del formulario
        self.object.save(update_fields=[
            'cantidad_regiones',
            'horario',
            'liquidar_como_extra_residencia',
            'monto_calculado',
            'paciente_internado',
            'fecha_hora_solicitud',
            'fecha_hora_informe'
        ])
        _guardar_extra_residencia_en_sesion(self.request, form)
        
        # Mostrar desglose del cálculo
        desglose = self.object.get_desglose_monto()
        mensaje_desglose = (
            f"✅ Práctica registrada | "
            f"Estudios: {desglose['estudios']} | "
            f"Regiones: {desglose['regiones']} | "
            f"OS: {desglose['tipo_os']} | "
            f"Horario: {desglose['horario']} | "
            f"Monto: ${desglose['monto_final']}"
        )
        if desglose.get('bonus_urgencia'):
            mensaje_desglose += f" (incluye bonus urgencia {desglose['bonus_urgencia']})"
        
        messages.success(self.request, mensaje_desglose)
        
        return redirect(self.success_url)
    
    def form_invalid(self, form):
        """Mostrar errores de validación al usuario"""
        # Mostrar todos los errores del formulario
        for field, errors in form.errors.items():
            for error in errors:
                if field == '__all__':
                    messages.error(self.request, f"Error: {error}")
                else:
                    field_label = form.fields.get(field).label if field in form.fields else field
                    messages.error(self.request, f"{field_label}: {error}")
        
        return super().form_invalid(form)


# [DEPRECADO - 16 de febrero 2026]
# RegistrarDiaSinPacientesView NO se usa en Colegiales
# Se mantiene comentado por compatibilidad con código legacy
#
# class RegistrarDiaSinPacientesView(LoginRequiredMixin, FormView):
#     template_name = 'liquidacion/registroestudios_form_tailwind.html'
#     form_class = DiaSinPacientesForm
#     success_url = reverse_lazy('liquidacion:registroestudios_nuevo')
# 
#     def form_valid(self, form):
#         fecha = form.cleaned_data['fecha']
#         medico = self.request.user
# 
#         if DiaSinPacientes.objects.filter(medico=medico, fecha=fecha).exists():
#             messages.warning(self.request, f"Ya registraste el día {fecha.strftime('%d/%m/%Y')}.")
#         else:
#             dia = form.save(commit=False)
#             dia.medico = medico
#             dia.save()
#             messages.success(self.request, f"Se registró el día {fecha.strftime('%d/%m/%Y')} como sin pacientes.")
# 
#         return super().form_valid(form)


# ============================================================================
# NUEVAS VISTAS - LIQUIDACIÓN v2.0
# ============================================================================

class RegistrarGuardiaPasivaView(LoginRequiredMixin, SuccessMessageMixin, CreateView):
    """
    Vista para registrar guardias pasivas con valor tomado desde configuración vigente.
    Solo disponible para médicos
    """
    model = GuardiaPasiva
    form_class = GuardiaPasivaForm
    template_name = 'liquidacion/guardia_pasiva_form.html'
    success_url = reverse_lazy('liquidacion:registrar_guardia_pasiva')
    success_message = "✅ Guardia pasiva registrada exitosamente"

    def dispatch(self, request, *args, **kwargs):
        # Médicos y administración pueden registrar guardias
        if not _puede_acceder_guardia_pasiva(request.user):
            messages.warning(request, "No tienes permiso para acceder a esta sección.")
            return redirect('home')
        return super().dispatch(request, *args, **kwargs)

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        user = self.request.user
        today = datetime.today()

        # Obtener sesión contable actual
        sesion, created = SesionContable.objects.get_or_create(
            mes=today.month,
            año=today.year,
            defaults={'estado': 'ABIERTA'}
        )
        context['sesion_contable'] = sesion
        context['puede_registrar'] = sesion.puede_registrar_practicas(user)

        # Guardias recientes del usuario (últimos 3 meses)
        # Esto permite ver guardias de meses anteriores/futuros
        fecha_desde = today.replace(day=1) - timedelta(days=90)
        guardias = GuardiaPasiva.objects.filter(
            medico=user,
            fecha_guardia__gte=fecha_desde
        ).order_by('-fecha_guardia')
        
        context['guardias'] = guardias
        # Contar solo las del mes actual para el badge
        guardias_mes_actual = guardias.filter(sesion_contable=sesion)
        context['total_guardias_mes'] = guardias_mes_actual.count()
        context['total_monto_guardias_mes'] = sum(g.monto for g in guardias_mes_actual)
        context['monto_guardia_vigente'] = ConfiguracionGuardiaPasiva.get_config().monto_vigente

        return context

    def form_valid(self, form):
        user = self.request.user
        fecha_guardia = form.cleaned_data['fecha_guardia']
        
        # Validar que la sesión permita registrar
        sesion, created = SesionContable.objects.get_or_create(
            mes=fecha_guardia.month,
            año=fecha_guardia.year,
            defaults={'estado': 'ABIERTA'}
        )
        
        if not sesion.puede_registrar_practicas(user):
            messages.error(
                self.request,
                f"❌ La sesión de {_mes_nombre(sesion.mes)} {sesion.año} está en estado "
                f"{sesion.get_estado_display()}. No puedes registrar guardias."
            )
            return redirect(self.success_url)
        
        # Verificar que no exista duplicado
        if GuardiaPasiva.objects.filter(medico=user, fecha_guardia=fecha_guardia).exists():
            messages.warning(
                self.request,
                f"⚠️ Ya tienes registrada una guardia para el día {fecha_guardia.strftime('%d/%m/%Y')}."
            )
            return redirect(self.success_url)
        
        # Asignar médico y tomar el monto vigente desde la configuración
        form.instance.medico = user
        
        response = super().form_valid(form)
        
        # Mensaje de éxito detallado
        messages.success(
            self.request,
            f"✅ Guardia pasiva registrada | "
            f"Fecha: {fecha_guardia.strftime('%d/%m/%Y')} | "
            f"Monto: ${form.instance.monto}"
        )
        
        return response


class GuardiaPasivaUpdateView(LoginRequiredMixin, SuccessMessageMixin, UpdateView):
    """
    Vista para editar guardias pasivas
    Solo el médico que creó la guardia puede editarla
    """
    model = GuardiaPasiva
    form_class = GuardiaPasivaForm
    template_name = 'liquidacion/guardia_pasiva_update.html'
    success_url = reverse_lazy('liquidacion:registrar_guardia_pasiva')
    success_message = "✅ Guardia pasiva actualizada correctamente"

    def dispatch(self, request, *args, **kwargs):
        # Médicos y administración pueden editar guardias
        if not _puede_acceder_guardia_pasiva(request.user):
            messages.warning(request, "No tienes permiso para acceder a esta sección.")
            return redirect('home')
        return super().dispatch(request, *args, **kwargs)

    def get_queryset(self):
        # Administración y superusuarios ven todas; el resto solo las propias
        if _puede_editar_monto_guardia(self.request.user):
            return GuardiaPasiva.objects.all()
        return GuardiaPasiva.objects.filter(medico=self.request.user)

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs

    def get_success_url(self):
        volver_sesion = _volver_sesion_url(self.request)
        if volver_sesion:
            return volver_sesion
        return super().get_success_url()

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        guardia = self.object
        
        # Sesión contable de la guardia
        context['sesion_contable'] = guardia.sesion_contable
        context['puede_editar'] = guardia.sesion_contable.puede_registrar_practicas(self.request.user)
        context['monto_guardia_vigente'] = ConfiguracionGuardiaPasiva.get_config().monto_vigente
        context['volver_sesion_url'] = _volver_sesion_url(self.request)

        return context

    def form_valid(self, form):
        guardia = self.object
        user = self.request.user
        fecha_guardia = form.cleaned_data['fecha_guardia']
        
        # Validar que la sesión permita editar
        sesion = guardia.sesion_contable
        if sesion.estado in ['FACTURADA', 'PAGADA']:
            messages.error(
                self.request,
                f"❌ La sesión de {_mes_nombre(sesion.mes)} {sesion.año} está en estado "
                f"{sesion.get_estado_display()}. No se pueden editar guardias en este estado."
            )
            return redirect(self.success_url)

        if not sesion.puede_registrar_practicas(user):
            messages.error(
                self.request,
                f"❌ La sesión de {_mes_nombre(sesion.mes)} {sesion.año} está en estado "
                f"{sesion.get_estado_display()}. No puedes editar guardias."
            )
            return redirect(self.success_url)
        
        # Verificar duplicados (excluyendo esta misma guardia)
        if GuardiaPasiva.objects.filter(
            medico=user, 
            fecha_guardia=fecha_guardia
        ).exclude(pk=guardia.pk).exists():
            messages.warning(
                self.request,
                f"⚠️ Ya tienes registrada otra guardia para el día {fecha_guardia.strftime('%d/%m/%Y')}."
            )
            return redirect(self.success_url)
        
        return super().form_valid(form)


class GuardiaPasivaDeleteView(LoginRequiredMixin, DeleteView):
    """
    Vista para eliminar guardias pasivas
    Solo el médico que creó la guardia puede eliminarla
    """
    model = GuardiaPasiva
    template_name = 'liquidacion/guardia_pasiva_confirm_delete.html'
    success_url = reverse_lazy('liquidacion:registrar_guardia_pasiva')

    def dispatch(self, request, *args, **kwargs):
        # Médicos y administración pueden eliminar guardias
        if not _puede_acceder_guardia_pasiva(request.user):
            messages.warning(request, "No tienes permiso para acceder a esta sección.")
            return redirect('home')
        return super().dispatch(request, *args, **kwargs)

    def get_queryset(self):
        # Administración y superusuarios ven todas; el resto solo las propias
        if _puede_editar_monto_guardia(self.request.user):
            return GuardiaPasiva.objects.all()
        return GuardiaPasiva.objects.filter(medico=self.request.user)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        guardia = self.object
        
        # Sesión contable de la guardia
        context['sesion_contable'] = guardia.sesion_contable
        context['puede_eliminar'] = guardia.sesion_contable.puede_registrar_practicas(self.request.user)
        
        return context

    def post(self, request, *args, **kwargs):
        guardia = self.get_object()
        self.object = guardia

        sesion = guardia.sesion_contable
        if sesion.estado in ['FACTURADA', 'PAGADA']:
            messages.error(
                request,
                f"❌ La sesión de {_mes_nombre(sesion.mes)} {sesion.año} está en estado "
                f"{sesion.get_estado_display()}. No se pueden eliminar guardias en este estado."
            )
            return redirect(self.success_url)

        if not sesion.puede_registrar_practicas(request.user):
            messages.error(
                request,
                f"❌ La sesión de {_mes_nombre(sesion.mes)} {sesion.año} está en estado "
                f"{sesion.get_estado_display()}. No puedes eliminar guardias."
            )
            return redirect(self.success_url)

        fecha = guardia.fecha_guardia.strftime('%d/%m/%Y')
        monto = guardia.monto

        response = super().post(request, *args, **kwargs)
        messages.success(
            request,
            f"🗑️ Guardia eliminada | Fecha: {fecha} | Monto: ${monto}"
        )
        return response


User = get_user_model()

class RegistroEstudiosPorMedicoListView(LoginRequiredMixin, TemplateView):
    """
    Dashboard de prácticas del médico - Liquidación v2.0
    Incluye prácticas, guardias pasivas, totales de montos y sesión contable
    """
    template_name = 'liquidacion/registroestudios_list_tailwind.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # Fecha actual
        fecha_actual = datetime.now()
        mes_actual = fecha_actual.month
        año_actual = fecha_actual.year

        # Inicializar el formulario
        params = self.request.GET.copy()
        if not params.get('mes'):
            params['mes'] = str(mes_actual)
        if not params.get('año'):
            params['año'] = str(año_actual)
        form = FiltroEstudiosPorMedicoForm(params)

        if form.is_valid():
            mes = form.cleaned_data.get('mes') or mes_actual
            año = form.cleaned_data.get('año') or año_actual
        else:
            mes, año = mes_actual, año_actual

        mes = int(mes)

        # Diccionario de nombres de meses
        MESES = {
            1: 'Enero', 2: 'Febrero', 3: 'Marzo', 4: 'Abril',
            5: 'Mayo', 6: 'Junio', 7: 'Julio', 8: 'Agosto',
            9: 'Septiembre', 10: 'Octubre', 11: 'Noviembre', 12: 'Diciembre'
        }

        # Obtener o crear sesión contable
        sesion, created = SesionContable.objects.get_or_create(
            mes=mes,
            año=año,
            defaults={'estado': 'ABIERTA'}
        )
        context['sesion_contable'] = sesion
        context['puede_registrar'] = sesion.puede_registrar_practicas(self.request.user)
        context['puede_registrar_guardia_pasiva'] = _puede_acceder_guardia_pasiva(self.request.user)

        # Pasar los valores base al contexto
        context['form'] = form
        context['mes'] = MESES.get(mes, 'Desconocido')
        context['año'] = año
        context['mes_num'] = mes
        context['año_num'] = año

        # CAMBIO v3.3: Filtrar por fecha_del_informe (cuando se hizo) en lugar de sesión_contable (cuando se registró)
        # Esto permite ver estudios atrasados en el mes que realmente se realizaron
        registros = RegistroEstudiosPorMedico.objects.filter(
            medico=self.request.user,
            fecha_del_informe__year=año,
            fecha_del_informe__month=mes
        ).prefetch_related('estudio', 'registroestudio_set__estudio')

        # Obtener guardias pasivas del mes (por fecha de guardia, no sesión)
        guardias = GuardiaPasiva.objects.filter(
            medico=self.request.user,
            fecha_guardia__year=año,
            fecha_guardia__month=mes
        ).order_by('-fecha_guardia')

        # Obtener parámetros de ordenamiento y filtros
        orden = self.request.GET.get('orden', 'fecha_desc')
        filtro_rapido = self.request.GET.get('filtro_rapido', '')
        busqueda = self.request.GET.get('busqueda', '').strip()
        modalidades_validas = [valor for valor, _ in Estudios.TIPO_ESTUDIO_CHOICES]
        modalidades_seleccionadas = [
            modalidad
            for modalidad in self.request.GET.getlist('modalidad')
            if modalidad in modalidades_validas
        ]
        base_query_params = self.request.GET.copy()
        if not base_query_params.get('mes'):
            base_query_params['mes'] = str(mes)
        if not base_query_params.get('año'):
            base_query_params['año'] = str(año)
        if 'tipo_estudio' in base_query_params:
            del base_query_params['tipo_estudio']

        # Aplicar búsqueda por paciente si existe
        if busqueda:
            registros = registros.filter(
                Q(nombre_paciente__icontains=busqueda) | 
                Q(apellido_paciente__icontains=busqueda) |
                Q(dni_paciente__icontains=busqueda)
            )

        # Aplicar filtros rápidos
        if filtro_rapido == 'hoy':
            from datetime import date
            registros = registros.filter(fecha_registro__date=date.today())

        # Filtro backend por modalidad (si el usuario selecciona checkboxes)
        if modalidades_seleccionadas:
            registros = registros.filter(estudio__tipo__in=modalidades_seleccionadas).distinct()

        # Aplicar ordenamiento
        if orden == 'fecha_asc':
            registros = registros.order_by('fecha_del_informe', 'fecha_registro')
        elif orden == 'fecha_desc':
            registros = registros.order_by('-fecha_del_informe', '-fecha_registro')
        elif orden == 'paciente_asc':
            registros = registros.order_by('apellido_paciente', 'nombre_paciente')
        elif orden == 'paciente_desc':
            registros = registros.order_by('-apellido_paciente', '-nombre_paciente')

        # Si el filtro rápido es 'hoy', ajustar visualmente mes/año
        if filtro_rapido == 'hoy':
            hoy = datetime.now()
            context['mes'] = MESES.get(hoy.month, 'Desconocido')
            context['año'] = hoy.year
            context['mes_num'] = hoy.month
            context['año_num'] = hoy.year

        # Etiqueta descriptiva del filtro rápido
        filtro_labels = {
            '': '',
            'hoy': 'Solo hoy',
        }
        context['filtro_rapido'] = filtro_rapido
        context['filtro_rapido_label'] = filtro_labels.get(filtro_rapido, '')
        context['modalidades_disponibles'] = [
            {'value': valor, 'label': etiqueta}
            for valor, etiqueta in Estudios.TIPO_ESTUDIO_CHOICES
        ]
        context['modalidades_seleccionadas'] = modalidades_seleccionadas
        context['modalidades_query'] = urlencode([
            ('modalidad', modalidad)
            for modalidad in modalidades_seleccionadas
        ])
        modalidades_toggle = []
        for valor, etiqueta in Estudios.TIPO_ESTUDIO_CHOICES:
            params_toggle = base_query_params.copy()
            actuales = [
                modalidad
                for modalidad in params_toggle.getlist('modalidad')
                if modalidad in modalidades_validas
            ]
            if valor in actuales:
                actualizadas = [modalidad for modalidad in actuales if modalidad != valor]
            else:
                actualizadas = actuales + [valor]
            params_toggle.setlist('modalidad', actualizadas)
            if not actualizadas and 'modalidad' in params_toggle:
                del params_toggle['modalidad']
            modalidades_toggle.append({
                'value': valor,
                'label': etiqueta,
                'active': valor in modalidades_seleccionadas,
                'query': params_toggle.urlencode(),
            })
        params_clear_modalidades = base_query_params.copy()
        if 'modalidad' in params_clear_modalidades:
            del params_clear_modalidades['modalidad']
        context['modalidades_toggle'] = modalidades_toggle
        context['modalidades_clear_query'] = params_clear_modalidades.urlencode()

        # Lista unificada para tabla principal
        registros_tabla = list(registros.distinct())

        # Los anulados siguen visibles, pero no integran cantidades ni montos.
        registros_vigentes = [reg for reg in registros_tabla if not reg.anulado]
        registros_eco = [
            reg for reg in registros_vigentes
            if any(rel.estudio.tipo == 'ECO' for rel in reg.registroestudio_set.all())
        ]
        registros_otros = [
            reg for reg in registros_vigentes
            if not any(rel.estudio.tipo == 'ECO' for rel in reg.registroestudio_set.all())
        ]
        # Adjuntar desglose calculado por backend para mostrar en la tabla
        puede_admin = puede_ver_desglose_administrativo(self.request.user)
        registros_ids = [registro.id for registro in registros_tabla]
        pendientes_ids = set(
            SolicitudRevisionHorarioRegistro.objects.filter(
                registro_id__in=registros_ids,
                estado=SolicitudRevisionHorarioRegistro.ESTADO_PENDIENTE,
            ).values_list('registro_id', flat=True)
        )

        revisiones_qs = (
            SolicitudRevisionHorarioRegistro.objects
            .filter(registro_id__in=registros_ids)
            .select_related('solicitado_por', 'revisado_por', 'aplicado_por')
            .order_by('registro_id', '-fecha_solicitud')
        )
        ultima_revision_por_registro = {}
        for revision in revisiones_qs:
            if revision.registro_id not in ultima_revision_por_registro:
                ultima_revision_por_registro[revision.registro_id] = revision

        correcciones_pacs_por_registro = {}
        correcciones_pacs_qs = (
            CorreccionPacsRegistro.objects
            .filter(registro_id__in=registros_ids)
            .select_related('corregido_por')
            .order_by('registro_id', '-fecha_correccion')
        )
        for correccion in correcciones_pacs_qs:
            if correccion.registro_id not in correcciones_pacs_por_registro:
                correcciones_pacs_por_registro[correccion.registro_id] = correccion

        for registro in registros_tabla:
            registro.detalle_monto = (
                registro.get_desglose_monto_administrativo()
                if puede_admin
                else registro.get_desglose_monto_simple()
            )
            sesion_bloqueada_revision = bool(
                registro.sesion_contable
                and registro.sesion_contable.estado in ['FACTURADA', 'PAGADA']
            )
            registro.tiene_revision_pendiente = registro.id in pendientes_ids
            registro.puede_solicitar_revision = (
                (not registro.anulado)
                and (not sesion_bloqueada_revision)
                and (not registro.tiene_revision_pendiente)
            )

            revision = ultima_revision_por_registro.get(registro.id)
            registro.revision_info = revision
            registro.tiene_revision = revision is not None
            registro.revision_badge_label = ''
            registro.revision_badge_classes = ''
            registro.correccion_pacs_info = correcciones_pacs_por_registro.get(registro.id)
            registro.tiene_correccion_pacs = registro.correccion_pacs_info is not None

            if revision:
                if revision.fecha_aplicacion:
                    registro.revision_badge_label = 'Corrección aplicada'
                    registro.revision_badge_classes = 'bg-emerald-100 text-emerald-800 border border-emerald-200'
                elif revision.estado == SolicitudRevisionHorarioRegistro.ESTADO_PENDIENTE:
                    registro.revision_badge_label = 'Revisión pendiente'
                    registro.revision_badge_classes = 'bg-amber-100 text-amber-800 border border-amber-200'
                elif revision.estado == SolicitudRevisionHorarioRegistro.ESTADO_APROBADA:
                    registro.revision_badge_label = 'Revisión aprobada'
                    registro.revision_badge_classes = 'bg-blue-100 text-blue-800 border border-blue-200'
                elif revision.estado == SolicitudRevisionHorarioRegistro.ESTADO_RECHAZADA:
                    registro.revision_badge_label = 'Revisión rechazada'
                    registro.revision_badge_classes = 'bg-rose-100 text-rose-800 border border-rose-200'

        context['puede_ver_desglose_admin'] = puede_admin

        # Agregar contexto para los controles
        context['orden'] = orden
        context['busqueda'] = busqueda

        # ========== CÁLCULOS v2.0 - Usando monto_calculado ==========
        
        # Totales de prácticas de Ecografías
        total_regiones_eco = sum(reg.cantidad_regiones for reg in registros_eco)
        total_monto_eco = sum(reg.monto_calculado for reg in registros_eco)
        
        # Totales de prácticas Otros estudios
        total_regiones_otros = sum(reg.cantidad_regiones for reg in registros_otros)
        total_monto_otros = sum(reg.monto_calculado for reg in registros_otros)
        
        # Totales de guardias pasivas
        total_guardias = guardias.count()
        total_monto_guardias = sum(g.monto for g in guardias)
        
        # Totales generales del mes
        total_practicas = len(registros_vigentes)
        total_regiones_general = total_regiones_eco + total_regiones_otros
        total_monto_practicas = total_monto_eco + total_monto_otros
        total_general = total_monto_practicas + total_monto_guardias

        # Agregar todo al contexto
        context['registros_eco'] = registros_eco
        context['total_regiones_eco'] = total_regiones_eco
        context['total_monto_eco'] = total_monto_eco
        
        context['registros_otros'] = registros_otros
        context['total_regiones_otros'] = total_regiones_otros
        context['total_monto_otros'] = total_monto_otros
        context['registros_tabla'] = registros_tabla
        
        context['guardias'] = guardias
        context['total_guardias'] = total_guardias
        context['total_monto_guardias'] = total_monto_guardias
        
        context['total_practicas'] = total_practicas
        context['total_regiones_general'] = total_regiones_general
        context['total_monto_practicas'] = total_monto_practicas
        context['total_general'] = total_general
        context['focus_registro'] = self.request.GET.get('focus_registro', '').strip()
        context['mostrar_info_horario_descuento'] = self.request.user.rol not in [
            'medico_staff',
            'jefe_servicio',
            'cardiologo',
        ]
        context['mostrar_info_bonus_urgencia'] = bool(self.request.user.trabaja_remoto)
        export_params = self.request.GET.copy()
        if not export_params.get('mes'):
            export_params['mes'] = str(mes)
        if not export_params.get('año'):
            export_params['año'] = str(año)
        context['exportar_mis_registros_url'] = (
            f"{reverse('liquidacion:exportar_excel_mis_registros')}?{export_params.urlencode()}"
        )
        
        return context


def _get_periodo_registros_personales(request):
    fecha_actual = datetime.now()
    mes_actual = fecha_actual.month
    año_actual = fecha_actual.year

    params = request.GET.copy()
    if not params.get('mes'):
        params['mes'] = str(mes_actual)
    if not params.get('año'):
        params['año'] = str(año_actual)

    form = FiltroEstudiosPorMedicoForm(params)
    if form.is_valid():
        mes = int(form.cleaned_data.get('mes') or mes_actual)
        año = int(form.cleaned_data.get('año') or año_actual)
    else:
        mes, año = mes_actual, año_actual

    return mes, año


def _get_registros_personales_filtrados(request):
    mes, año = _get_periodo_registros_personales(request)
    registros = (
        RegistroEstudiosPorMedico.objects
        .filter(
            medico=request.user,
            fecha_del_informe__year=año,
            fecha_del_informe__month=mes,
            anulado=False,
        )
        .select_related('sesion_contable')
        .prefetch_related('registroestudio_set__estudio__grupo_tarifario')
    )

    orden = request.GET.get('orden', 'fecha_desc')
    filtro_rapido = request.GET.get('filtro_rapido', '')
    busqueda = request.GET.get('busqueda', '').strip()
    modalidades_validas = [valor for valor, _ in Estudios.TIPO_ESTUDIO_CHOICES]
    modalidades_seleccionadas = [
        modalidad
        for modalidad in request.GET.getlist('modalidad')
        if modalidad in modalidades_validas
    ]

    if busqueda:
        registros = registros.filter(
            Q(nombre_paciente__icontains=busqueda)
            | Q(apellido_paciente__icontains=busqueda)
            | Q(dni_paciente__icontains=busqueda)
        )

    if filtro_rapido == 'hoy':
        registros = registros.filter(fecha_registro__date=date.today())

    if modalidades_seleccionadas:
        registros = registros.filter(estudio__tipo__in=modalidades_seleccionadas).distinct()

    if orden == 'fecha_asc':
        registros = registros.order_by('fecha_del_informe', 'fecha_registro')
    elif orden == 'fecha_desc':
        registros = registros.order_by('-fecha_del_informe', '-fecha_registro')
    elif orden == 'paciente_asc':
        registros = registros.order_by('apellido_paciente', 'nombre_paciente')
    elif orden == 'paciente_desc':
        registros = registros.order_by('-apellido_paciente', '-nombre_paciente')

    guardias = GuardiaPasiva.objects.filter(
        medico=request.user,
        fecha_guardia__year=año,
        fecha_guardia__month=mes,
    ).order_by('-fecha_guardia')

    return mes, año, registros.distinct(), guardias


@login_required
def exportar_excel_mis_registros(request):
    mes, año, registros, guardias = _get_registros_personales_filtrados(request)
    registros = list(registros)
    guardias = list(guardias)

    wb = Workbook()
    ws = wb.active
    ws.title = 'Practicas'
    ws.append([
        'Fecha informe',
        'Fecha carga',
        'Paciente',
        'DNI',
        'Obra social',
        'Horario',
        'Estudios',
        'Modalidades',
        'Cantidad regiones',
        'Monto calculado',
        'Sesion contable',
        'Estado sesion',
    ])

    for registro in registros:
        relaciones = list(registro.registroestudio_set.all())
        estudios = []
        modalidades = []
        for rel in relaciones:
            estudio = rel.estudio
            cantidad = f' x{rel.cantidad}' if rel.cantidad and rel.cantidad > 1 else ''
            contexto = f' ({rel.get_contexto_display()})' if rel.contexto else ''
            estudios.append(f'{estudio.nombre}{contexto}{cantidad}')
            modalidades.append(estudio.get_tipo_display())

        sesion = registro.sesion_contable
        ws.append([
            registro.fecha_del_informe.strftime('%d/%m/%Y') if registro.fecha_del_informe else '',
            registro.fecha_registro.strftime('%d/%m/%Y %H:%M') if registro.fecha_registro else '',
            f'{registro.apellido_paciente}, {registro.nombre_paciente}',
            registro.dni_paciente,
            registro.get_tipo_obra_social_display(),
            registro.get_horario_display(),
            '; '.join(estudios),
            '; '.join(dict.fromkeys(modalidades)),
            registro.cantidad_regiones,
            float(registro.monto_calculado or 0),
            f'{sesion.mes}/{sesion.año}' if sesion else '',
            sesion.get_estado_display() if sesion else '',
        ])

    primera_fila_practicas = 2
    ultima_fila_practicas = ws.max_row
    tiene_practicas = ultima_fila_practicas >= primera_fila_practicas
    fila_total_practicas = ultima_fila_practicas + 1
    ws.append([
        '',
        '',
        '',
        '',
        '',
        '',
        '',
        'Totales',
        f'=SUM(I{primera_fila_practicas}:I{ultima_fila_practicas})' if tiene_practicas else 0,
        f'=SUM(J{primera_fila_practicas}:J{ultima_fila_practicas})' if tiene_practicas else 0,
        '',
        '',
    ])

    ws_guardias = wb.create_sheet('Guardias')
    ws_guardias.append(['Fecha guardia', 'Tipo', 'Monto', 'Observaciones'])
    for guardia in guardias:
        ws_guardias.append([
            guardia.fecha_guardia.strftime('%d/%m/%Y') if guardia.fecha_guardia else '',
            guardia.get_tipo_guardia_display(),
            float(guardia.monto or 0),
            guardia.observaciones,
        ])

    primera_fila_guardias = 2
    ultima_fila_guardias = ws_guardias.max_row
    tiene_guardias = ultima_fila_guardias >= primera_fila_guardias
    fila_total_guardias = ultima_fila_guardias + 1
    ws_guardias.append([
        '',
        'Totales',
        f'=SUM(C{primera_fila_guardias}:C{ultima_fila_guardias})' if tiene_guardias else 0,
        '',
    ])

    ws_resumen = wb.create_sheet('Resumen')
    total_practicas = sum((registro.monto_calculado or Decimal('0.00')) for registro in registros)
    total_guardias = sum((guardia.monto or Decimal('0.00')) for guardia in guardias)
    ws_resumen.append(['Profesional', request.user.get_full_name() or request.user.username])
    ws_resumen.append(['Periodo', f'{mes}/{año}'])
    ws_resumen.append([
        'Cantidad practicas',
        f'=COUNTA(Practicas!C{primera_fila_practicas}:C{ultima_fila_practicas})' if tiene_practicas else 0,
    ])
    ws_resumen.append(['Regiones practicas', f'=Practicas!I{fila_total_practicas}'])
    ws_resumen.append(['Monto practicas', f'=Practicas!J{fila_total_practicas}'])
    ws_resumen.append([
        'Cantidad guardias',
        f'=COUNTA(Guardias!A{primera_fila_guardias}:A{ultima_fila_guardias})' if tiene_guardias else 0,
    ])
    ws_resumen.append(['Monto guardias', f'=Guardias!C{fila_total_guardias}'])
    ws_resumen.append(['Total', '=B5+B7'])
    ws_resumen.append(['Monto practicas persistido', float(total_practicas)])
    ws_resumen.append(['Monto guardias persistido', float(total_guardias)])
    ws_resumen.append(['Total persistido', float(total_practicas + total_guardias)])

    for sheet in wb.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical='top', wrap_text=True)
        for column_cells in sheet.columns:
            column_letter = get_column_letter(column_cells[0].column)
            sheet.column_dimensions[column_letter].width = min(
                max(len(str(cell.value or '')) for cell in column_cells) + 2,
                45,
            )
        for cell in sheet[1]:
            cell.font = Font(bold=True)
        if sheet.title == 'Practicas':
            for cell in sheet[fila_total_practicas]:
                cell.font = Font(bold=True)
            for row in sheet.iter_rows(min_row=2, min_col=9, max_col=10):
                for cell in row:
                    cell.number_format = '$#,##0.00' if cell.column == 10 else '#,##0'
        elif sheet.title == 'Guardias':
            for cell in sheet[fila_total_guardias]:
                cell.font = Font(bold=True)
            for row in sheet.iter_rows(min_row=2, min_col=3, max_col=3):
                for cell in row:
                    cell.number_format = '$#,##0.00'
        elif sheet.title == 'Resumen':
            for cell in sheet['A']:
                cell.font = Font(bold=True)
            for row_number in [5, 7, 8, 9, 10, 11]:
                sheet[f'B{row_number}'].number_format = '$#,##0.00'
            for row_number in [3, 4, 6]:
                sheet[f'B{row_number}'].number_format = '#,##0'

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    username = ''.join(char for char in request.user.username if char.isalnum() or char in ('-', '_'))
    response = HttpResponse(
        buffer.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="mis_registros_{username}_{mes}_{año}.xlsx"'
    return response


class RegistroEstudiosPorMedicoUpdateView(LoginRequiredMixin, UpdateView):
    model = RegistroEstudiosPorMedico
    form_class = RegistroEstudiosPorMedicoCreateViewForm
    template_name = 'liquidacion/registroestudios_update_tailwind_v2.html'

    def dispatch(self, request, *args, **kwargs):
        if not _puede_acceder_panel_medico(request.user):
            messages.warning(request, "No tienes permiso para acceder a esta sección.")
            return redirect('home')
        return super().dispatch(request, *args, **kwargs)

    def get_queryset(self):
        # Filtra los registros que pertenecen al usuario logueado
        return RegistroEstudiosPorMedico.objects.filter(medico=self.request.user, anulado=False)

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        registro = self.object
        user = self.request.user
        sesion = registro.sesion_contable

        # Serializar estudios para JS (con todos los datos necesarios)
        estudios_data = []
        for estudio in Estudios.objects.filter(activo=True).values(
            'id', 'nombre', 'tipo', 'codigo', 'precio_cober', 'precio_otras_os',
            'precio_unico', 'conteo_regiones_default', 'tiene_contexto_ubicacion'
        ):
            if (not user.is_superuser) and user.rol != 'cardiologo' and es_estudio_cardiologico(
                estudio['tipo'],
                estudio['nombre'],
                estudio['codigo'],
            ):
                continue
            estudio_dict = dict(estudio)
            # Convertir Decimals a string para JSON
            estudio_dict['precio_cober'] = str(estudio_dict['precio_cober'])
            estudio_dict['precio_otras_os'] = str(estudio_dict['precio_otras_os'])
            estudio_dict['es_cardiologico'] = es_estudio_cardiologico(
                estudio_dict['tipo'],
                estudio_dict['nombre'],
                estudio_dict['codigo'],
            )
            estudio_dict['contextos_disponibles'] = contextos_disponibles_para_estudio(
                estudio_dict['tipo'],
                estudio_dict['nombre'],
                estudio_dict['codigo'],
            )
            estudios_data.append(estudio_dict)

        context['estudios'] = json.dumps(estudios_data)
        context['estudios_json'] = json.dumps(estudios_data)  # Alias para compatibilidad

        # Estudios y tipo preseleccionado
        if registro and registro.estudio.exists():
            context['tipo_estudio_seleccionado'] = registro.estudio.first().tipo
            context['estudios_seleccionados'] = list(registro.estudio.values_list('id', flat=True))

            # Cargar cantidades y contextos desde tabla intermedia RegistroEstudio
            from liquidacion.models import RegistroEstudio
            cantidades_dict = {}
            contextos_dict = {}
            for rel in RegistroEstudio.objects.filter(registro=registro).select_related('estudio'):
                cantidades_dict[rel.estudio_id] = rel.cantidad
                contextos_dict[rel.estudio_id] = rel.contexto
            context['cantidades_estudios'] = json.dumps(cantidades_dict)
            context['contextos_estudios'] = json.dumps(contextos_dict)
        else:
            context['tipo_estudio_seleccionado'] = ''
            context['estudios_seleccionados'] = []
            context['cantidades_estudios'] = '{}'
            context['contextos_estudios'] = '{}'
        
        # Información del médico para lógica condicional
        context['trabaja_remoto'] = self.request.user.trabaja_remoto
        context['es_staff'] = self.request.user.rol in ['medico_staff', 'jefe_servicio', 'cardiologo']
        context['es_cardiologo'] = self.request.user.is_superuser or self.request.user.rol == 'cardiologo'
        context['requiere_motivo_modificacion'] = bool(
            sesion and sesion.estado in ['CERRADA', 'FACTURADA']
        )
        context['puede_ver_desglose_admin'] = puede_ver_desglose_administrativo(self.request.user)

        # URL del botón cancelar: vuelve a la lista con mes/año filtrados
        fecha = registro.fecha_del_informe
        context['cancel_url'] = f"{reverse('liquidacion:registroestudios_list')}?{urlencode({'mes': fecha.month, 'año': fecha.year})}"

        return context

    def form_valid(self, form):
        sesion = self.object.sesion_contable
        user = self.request.user
        if sesion and not sesion.puede_registrar_practicas(self.request.user):
            messages.error(
                self.request,
                f"❌ La sesión de {sesion.mes}/{sesion.año} está en estado "
                f"{sesion.get_estado_display()}. No puedes editar prácticas."
            )
            return redirect(self.get_success_url())

        estudios_seleccionados = form.cleaned_data['estudio']
        if (not user.is_superuser) and user.rol != 'cardiologo':
            estudios_bloqueados = [
                est.nombre for est in estudios_seleccionados
                if es_estudio_cardiologico(est.tipo, est.nombre, est.codigo)
            ]
            if estudios_bloqueados:
                form.add_error(
                    'estudio',
                    'No tienes permisos para registrar estudios cardiológicos. '
                    f"Detectados: {', '.join(estudios_bloqueados)}"
                )
                return self.form_invalid(form)

        motivo_modificacion = (self.request.POST.get('motivo_modificacion') or '').strip()
        if sesion and sesion.estado in ['CERRADA', 'FACTURADA'] and not motivo_modificacion:
            form.add_error(
                None,
                'Debes indicar el motivo de modificación para corregir registros en sesiones cerradas o facturadas.'
            )
            return self.form_invalid(form)

        # Guardar objeto
        self.object = form.save(commit=False)
        if user.rol not in ROLES_LIQUIDAR_COMO_EXTRA_RESIDENCIA:
            self.object.liquidar_como_extra_residencia = False
        self.object.save()
        
        # v3.1 - Marzo 2026: NO usar save_m2m() con through model
        # Actualizar tabla intermedia RegistroEstudio con cantidades
        
        # Leer cantidades de estudios desde el POST
        cantidades_estudios = {}
        for key in self.request.POST:
            if key.startswith('cantidad_estudio_'):
                estudio_id = int(key.replace('cantidad_estudio_', ''))
                cantidad = int(self.request.POST[key])
                cantidades_estudios[estudio_id] = cantidad
        
        # Importar el modelo intermedio
        from liquidacion.models import RegistroEstudio, Estudios
        
        # Leer contextos de ubicación desde el POST
        contextos_estudios = {}
        for key in self.request.POST:
            if key.startswith('contexto_estudio_'):
                estudio_id = int(key.replace('contexto_estudio_', ''))
                contextos_estudios[estudio_id] = self.request.POST[key]

        # Limpiar relaciones existentes
        RegistroEstudio.objects.filter(registro=self.object).delete()

        # Obtener estudios seleccionados del formulario
        estudios_seleccionados = form.cleaned_data['estudio']

        # Crear nuevas relaciones con cantidades y contexto actualizados
        for estudio in estudios_seleccionados:
            cantidad = cantidades_estudios.get(estudio.id, 1)
            contextos_validos = contextos_disponibles_para_estudio(estudio.tipo, estudio.nombre, estudio.codigo)
            contexto = contextos_estudios.get(estudio.id, contextos_validos[0] if contextos_validos else 'SERVICIO')
            if contexto not in contextos_validos:
                contexto = contextos_validos[0] if contextos_validos else 'SERVICIO'
            RegistroEstudio.objects.create(
                registro=self.object,
                estudio=estudio,
                cantidad=cantidad,
                contexto=contexto,
            )
        
        # Recalcular cantidad de regiones con las cantidades especificadas
        total_regiones = 0
        for estudio in estudios_seleccionados:
            cantidad = cantidades_estudios.get(estudio.id, 1)
            total_regiones += (estudio.conteo_regiones_default * cantidad)
        
        self.object.cantidad_regiones = total_regiones

        # Fuente canónica para residencia+ECO general real: clasificación explícita post-M2M.
        if (
            user.rol in ROLES_LIQUIDAR_COMO_EXTRA_RESIDENCIA
            and self.object.liquidar_como_extra_residencia
        ):
            self.object.horario = 'EXTRA'
        elif user.rol in ROLES_RESIDENCIA:
            tiene_eco_general = any(
                es_eco_general_real_estudio(est)
                for est in estudios_seleccionados
            )
            if tiene_eco_general:
                nuevo_horario = clasificar_horario_residencia_por_proxy(
                    rol=user.rol,
                    fecha_registro=self.object.fecha_registro,
                    tiene_eco_general=True,
                    fecha_practica=self.object.fecha_del_informe,
                )
                self.object.horario = nuevo_horario or 'NA'
            else:
                self.object.horario = 'NA'
        
        # v3.1: Recalcular monto usando método unificado del modelo
        total_monto = self.object.calcular_monto()
        self.object.monto_calculado = total_monto
        
        # Auditoría: registrar quién y cuándo modificó
        self.object.modificado_por = self.request.user
        self.object.fecha_modificacion = now()
        self.object.motivo_modificacion = motivo_modificacion
        
        # También guardar campos de bonus urgencia que vienen del formulario
        self.object.save(update_fields=[
            'cantidad_regiones',
            'horario',
            'liquidar_como_extra_residencia',
            'monto_calculado',
            'paciente_internado',
            'fecha_hora_solicitud',
            'fecha_hora_informe',
            'modificado_por',
            'fecha_modificacion',
            'motivo_modificacion',
        ])
        _guardar_extra_residencia_en_sesion(self.request, form)
        
        # Mostrar desglose del cálculo actualizado
        desglose = self.object.get_desglose_monto()
        mensaje_desglose = (
            f"✅ Registro actualizado | "
            f"Estudios: {desglose['estudios']} | "
            f"Regiones: {desglose['regiones']} | "
            f"OS: {desglose['tipo_os']} | "
            f"Horario: {desglose['horario']} | "
            f"Monto: ${desglose['monto_final']}"
        )
        if desglose.get('bonus_urgencia'):
            mensaje_desglose += f" (incluye bonus urgencia {desglose['bonus_urgencia']})"
        
        messages.success(self.request, mensaje_desglose)
        return HttpResponseRedirect(self.get_success_url())

    def form_invalid(self, form):
        # Mostrar errores del formulario al usuario
        for field, errors in form.errors.items():
            for error in errors:
                if field == '__all__':
                    messages.error(self.request, f"Error: {error}")
                else:
                    field_label = form.fields.get(field).label if field in form.fields else field
                    messages.error(self.request, f"{field_label}: {error}")
        return super().form_invalid(form)

    def get_success_url(self):
        # Redirige a la lista con el mes y año del registro actualizado
        fecha = self.object.fecha_del_informe
        
        # Determinar el tipo de estudio para mantener la solapa activa
        # Solo hay 2 pestañas: 'ecografias' y 'otros' (RAD/TOM/RES)
        tipos = set(self.object.estudio.values_list('tipo', flat=True))
        
        if 'ECO' in tipos:
            tipo_estudio = 'ecografias'
        else:
            # Cualquier otro tipo (RAD, TOM, RES, etc.) va a 'otros'
            tipo_estudio = 'otros'
        
        query_string = urlencode({
            'mes': fecha.month, 
            'año': fecha.year,
            'tipo_estudio': tipo_estudio,
            'focus_registro': self.object.pk,
        })
        return f"{reverse('liquidacion:registroestudios_list')}?{query_string}"


class SolicitudRevisionHorarioRegistroCreateView(LoginRequiredMixin, SuccessMessageMixin, CreateView):
    """Alta médica de solicitud de revisión de horario (Fase A, sin impacto económico)."""

    model = SolicitudRevisionHorarioRegistro
    form_class = SolicitudRevisionHorarioRegistroForm
    template_name = 'liquidacion/revision_horario_solicitud_form.html'
    success_message = '✅ Solicitud de revisión enviada correctamente.'

    def dispatch(self, request, *args, **kwargs):
        if not _puede_acceder_panel_medico(request.user):
            messages.warning(request, 'No tienes permiso para acceder a esta sección.')
            return redirect('home')

        self.registro = get_object_or_404(
            RegistroEstudiosPorMedico.objects.select_related('sesion_contable', 'medico'),
            pk=kwargs['registro_pk'],
        )

        if self.registro.anulado:
            messages.error(request, 'No se puede solicitar revision sobre un registro anulado.')
            return redirect('liquidacion:registroestudios_list')

        if self.registro.medico_id != request.user.id:
            messages.error(request, '❌ No puedes solicitar revisión sobre registros de otro usuario.')
            return redirect('liquidacion:registroestudios_list')

        if self.registro.sesion_contable and self.registro.sesion_contable.estado in ['FACTURADA', 'PAGADA']:
            messages.error(
                request,
                '❌ No se puede solicitar revisión en sesiones FACTURADAS o PAGADAS.',
            )
            return redirect('liquidacion:registroestudios_list')

        if SolicitudRevisionHorarioRegistro.objects.filter(
            registro=self.registro,
            estado=SolicitudRevisionHorarioRegistro.ESTADO_PENDIENTE,
        ).exists():
            messages.warning(
                request,
                '⚠️ Ya existe una solicitud de revisión pendiente para este registro.',
            )
            return redirect('liquidacion:registroestudios_list')

        return super().dispatch(request, *args, **kwargs)

    def form_valid(self, form):
        form.instance.registro = self.registro
        form.instance.solicitado_por = self.request.user
        return super().form_valid(form)

    def get_success_url(self):
        return reverse('liquidacion:registroestudios_list')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['registro'] = self.registro
        return context

class RegistroEstudiosPorMedicoDeleteView(LoginRequiredMixin, DeleteView):
    model = RegistroEstudiosPorMedico
    template_name = 'liquidacion/registroestudios_confirm_delete_tailwind.html'

    def dispatch(self, request, *args, **kwargs):
        if not _puede_acceder_panel_medico(request.user):
            messages.warning(request, "No tienes permiso para acceder a esta sección.")
            return redirect('home')
        return super().dispatch(request, *args, **kwargs)

    def get_queryset(self):
        # Limita los registros a los del usuario logueado
        return RegistroEstudiosPorMedico.objects.filter(medico=self.request.user, anulado=False)

    def get_success_url(self):
        # Mantener el filtro de mes/año y tipo de estudio tras eliminar
        registro = self.object
        fecha = registro.fecha_del_informe
        
        # Determinar el tipo de estudio para mantener la solapa activa
        # Solo hay 2 pestañas: 'ecografias' y 'otros' (RAD/TOM/RES)
        tipos = set(registro.estudio.values_list('tipo', flat=True))
        
        if 'ECO' in tipos:
            tipo_estudio = 'ecografias'
        else:
            # Cualquier otro tipo (RAD, TOM, RES, etc.) va a 'otros'
            tipo_estudio = 'otros'
        
        query_string = urlencode({
            'mes': fecha.month,
            'año': fecha.year,
            'tipo_estudio': tipo_estudio
        })
        return f"{reverse('liquidacion:registroestudios_list')}?{query_string}"

    def post(self, request, *args, **kwargs):
        registro = self.get_object()
        self.object = registro

        sesion = registro.sesion_contable
        if sesion and not sesion.puede_registrar_practicas(request.user):
            messages.error(
                request,
                f"❌ La sesión de {sesion.mes}/{sesion.año} está en estado "
                f"{sesion.get_estado_display()}. No puedes eliminar prácticas."
            )
            return redirect(self.get_success_url())

        messages.success(request, "✅ Registro eliminado correctamente.")
        return super().post(request, *args, **kwargs)

# ============================================================
# [ANULADO - 16 de febrero 2026]
# Procedimientos de Intervensionismo - En Colegiales no se usa
# Los procedimientos se registran como Estudios normales
# Si necesitas datos históricos: ver liquidacion_backup_completo_2026-02-16.json
# ============================================================

# ========================================
# VISTA UNIFICADA - v3.0 (Feb 2026)
# ========================================

class LiquidacionPorMedicoPorMesListView(LoginRequiredMixin, UserPassesTestMixin, TemplateView):
    """
    Vista unificada de liquidación mensual - Todas las prácticas + Guardias
    
    PERMISOS: Solo administrativos, jefes de servicio y superusuarios
    
    Muestra en un solo portal:
    - Todas las prácticas (ECO + RAD + TOM + RES)
    - Guardias pasivas
    - Total general
    
    Eliminada lógica obsoleta:
    - DiaSinPacientes (no se usa en Colegiales)
    - Complemento por mínimo garantizado (no aplica)
    - Cálculo de regiones faltantes
    """
    template_name = 'liquidacion/liquidacion_por_medico_por_mes_tailwind.html'
    
    def test_func(self):
        """Solo administrativos, jefe de servicio, y superusuarios"""
        return (
            self.request.user.is_superuser or 
            self.request.user.rol in ['administrativo', 'jefe_servicio']
        )
    
    def handle_no_permission(self):
        messages.error(
            self.request,
            '❌ No tienes permisos para acceder a esta vista. '
            'Esta sección está disponible solo para personal administrativo y coordinadores.'
        )
        return redirect('home')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        form = FiltroMedicoMesForm(self.request.GET or None)
        context['form'] = form

        registros_por_medico = defaultdict(list)
        guardias_por_medico = defaultdict(list)

        if form.is_valid():
            medico = form.cleaned_data.get('medico')
            mes = form.cleaned_data.get('mes')
            año = form.cleaned_data.get('año')

            # Filtrar TODAS las prácticas (sin excluir ningún tipo)
            from liquidacion.models import RegistroEstudio
            registros = RegistroEstudiosPorMedico.objects.filter(anulado=False).select_related(
                'medico', 'sesion_contable'
            ).prefetch_related(
                Prefetch('estudio', queryset=Estudios.objects.all()),
                Prefetch('registroestudio_set', queryset=RegistroEstudio.objects.select_related('estudio')),
            ).distinct()

            # Filtrar guardias pasivas
            guardias = GuardiaPasiva.objects.select_related('medico')

            if medico:
                registros = registros.filter(medico=medico)
                guardias = guardias.filter(medico=medico)

            if mes and año:
                registros = registros.filter(fecha_del_informe__year=int(año), fecha_del_informe__month=int(mes))
                guardias = guardias.filter(fecha_guardia__year=int(año), fecha_guardia__month=int(mes))

            # Agrupar registros por médico
            for registro in registros.order_by('-fecha_del_informe'):
                registros_por_medico[registro.medico].append(registro)
            
            # Agrupar guardias por médico
            for guardia in guardias.order_by('-fecha_guardia'):
                guardias_por_medico[guardia.medico].append(guardia)

        # Preparar el contexto con datos por médico
        medico_data = []
        todos_medicos = set(registros_por_medico.keys()) | set(guardias_por_medico.keys())
        
        for medico in todos_medicos:
            registros = adjuntar_ultima_correccion_pacs(registros_por_medico.get(medico, []))
            guardias = guardias_por_medico.get(medico, [])
            
            # v3.1 - Marzo 2026: Enriquecer cada registro con cantidades de estudios
            from liquidacion.models import RegistroEstudio
            for registro in registros:
                # Agregar lista de estudios con cantidades al registro
                estudios_con_cantidades = []
                cantidades_por_tipo = defaultdict(int)  # {'RES': 4, 'ECO': 1}
                
                # Usa el prefetch cargado arriba — sin query adicional por registro
                for rel in registro.registroestudio_set.all():
                    estudios_con_cantidades.append({
                        'estudio': rel.estudio,
                        'cantidad': rel.cantidad,
                        'tipo': rel.estudio.tipo,
                        'contexto': rel.contexto,
                    })
                    cantidades_por_tipo[rel.estudio.tipo] += rel.cantidad
                
                # Agregar atributos temporales al objeto (no se guardan en BD)
                registro.estudios_con_cantidades = estudios_con_cantidades
                registro.cantidades_por_tipo = dict(cantidades_por_tipo)
            
            # Agrupar prácticas por tipo para mostrar subtotales (ya no se usa en v3.1)
            practicas_por_tipo = defaultdict(list)
            for registro in registros:
                for estudio in registro.estudio.all():
                    practicas_por_tipo[estudio.tipo].append(registro)
            
            total_regiones = sum(registro.cantidad_regiones for registro in registros)
            total_monto = sum(registro.monto_calculado for registro in registros)
            correcciones_pacs = [
                registro.correccion_pacs_info
                for registro in registros
                if registro.tiene_correccion_pacs
            ]
            total_impacto_pacs = sum(
                registro.impacto_correccion_pacs
                for registro in registros
                if registro.tiene_correccion_pacs
            )
            total_guardias = len(guardias)
            total_monto_guardias = sum(guardia.monto for guardia in guardias)
            
            medico_data.append({
                'medico': medico,
                'registros': registros,
                'practicas_por_tipo': dict(practicas_por_tipo),
                'guardias': guardias,
                'total_regiones': total_regiones,
                'total_monto': total_monto,
                'correcciones_pacs': correcciones_pacs,
                'correcciones_pacs_count': len(correcciones_pacs),
                'total_impacto_pacs': total_impacto_pacs,
                'total_guardias': total_guardias,
                'total_monto_guardias': total_monto_guardias,
                'total_general': total_monto + total_monto_guardias,
            })

        context['medico_data'] = medico_data
        
        if form.is_valid() and not medico_data:
            medico_seleccionado = form.cleaned_data.get('medico')
            if medico_seleccionado:
                context['mensaje_sin_registros'] = (
                    f"No se encontraron registros para {medico_seleccionado.get_full_name()} en el período consultado."
                )
            else:
                context['mensaje_sin_registros'] = "No se encontraron registros en el período consultado."

        return context

# [ANULADO - 16 de febrero 2026]
# Vista ProcedimientosPorMedicoPorMesListView eliminada
# En Colegiales, los procedimientos se registran como Estudios
# Ver liquidacion_backup_completo_2026-02-16.json para datos históricos

def generar_pdf_liquidacion(request):
    """
    Generar PDF de liquidación - v2.0
    Incluye: prácticas con montos, guardias pasivas, totales
    """
    from .services import generar_buffer_pdf_liquidacion
    buffer = generar_buffer_pdf_liquidacion()
    return FileResponse(buffer, as_attachment=True, filename="Liquidacion_Medicos_v2.pdf")

# [ANULADO - 16 de febrero 2026 y ELIMINADAS v3.0 - 17 feb 2026]
# Funciones exportar_excel_informes y exportar_excel_ecografias eliminadas
# Usar exportar_excel_liquidacion (vista unificada) en su lugar
# En Colegiales los procedimientos se registran como estudios
# Ver ANALISIS_LIQUIDACION_COLEGIALES.md para más detalles


# ========================================
# EXPORTACIÓN UNIFICADA - v3.0 (Feb 2026)
# ========================================

def _es_usuario_exportacion_liquidacion(user):
    return user.is_superuser or user.rol in ['administrativo', 'jefe_servicio']


def _generar_respuesta_exportacion_liquidacion(request, etiqueta_archivo):
    from .services import generar_buffer_excel_liquidacion

    medico_id = request.GET.get('medico')
    mes = request.GET.get('mes')
    año = request.GET.get('año')

    medico = None
    if medico_id:
        medico = get_object_or_404(User, id=medico_id)

    buffer, nombre_medico = generar_buffer_excel_liquidacion(medico=medico, mes=mes, año=año)

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = (
        f'attachment; filename="liquidacion_{etiqueta_archivo}_{nombre_medico}_{mes}_{año}.xlsx"'
    )
    response.write(buffer.read())
    return response


@login_required
def exportar_excel_liquidacion(request):
    """Exportación preliminar para seguimiento interno administrativo."""
    if not _es_usuario_exportacion_liquidacion(request.user):
        messages.error(request, '❌ No tienes permisos para exportar datos de liquidación.')
        return redirect('home')
    return _generar_respuesta_exportacion_liquidacion(request, etiqueta_archivo='PRELIMINAR')


@login_required
def exportar_excel_liquidacion_definitiva(request):
    """Exportación definitiva permitida solo con sesión FACTURADA/PAGADA."""
    if not _es_usuario_exportacion_liquidacion(request.user):
        messages.error(request, '❌ No tienes permisos para exportar datos de liquidación.')
        return redirect('home')

    mes = request.GET.get('mes')
    año = request.GET.get('año')
    if not mes or not año:
        messages.error(request, '❌ Debes indicar mes y año para exportación definitiva.')
        return redirect('liquidacion:sesiones_list')

    sesion = get_object_or_404(SesionContable, mes=mes, año=año)
    if sesion.estado not in ['FACTURADA', 'PAGADA']:
        messages.error(
            request,
            '❌ La exportación definitiva solo está habilitada para sesiones FACTURADAS o PAGADAS.',
        )
        return redirect('liquidacion:sesiones_list')

    return _generar_respuesta_exportacion_liquidacion(request, etiqueta_archivo='DEFINITIVA')

# A continuación, se agrega el formulario para carga masiva de estudios

User = get_user_model()

# --- Carga masiva desactivada temporalmente ---
from django.contrib.auth.mixins import UserPassesTestMixin

class CargaMasivaView(LoginRequiredMixin, UserPassesTestMixin, FormView):
    template_name = 'liquidacion/carga_formulario.html'
    form_class = CargaExcelForm
    success_url = reverse_lazy('carga-masiva')

    def test_func(self):
        return self.request.user.is_superuser or self.request.user.is_staff

    def post(self, request, *args, **kwargs):
        if 'datos_serializados' in request.POST:
            return self.confirmar_carga(request)
        return super().post(request, *args, **kwargs)

    def form_valid(self, form):
        archivo = form.cleaned_data['archivo_excel']
        df = pd.read_excel(archivo)
        registros_preview = []
        mapeo_estudios = {
            'US': {'tipo': 'ECO', 'nombre': 'ECO ABDOMINAL'},
            'CR': {'tipo': 'RAD', 'nombre': 'RX DE TÓRAX'},
            'DX': {'tipo': 'RAD', 'nombre': 'RX DE TÓRAX'},
            'CT': {'tipo': 'TOM', 'nombre': 'TC DE CEREBRO'},
            'MR': {'tipo': 'RES', 'nombre': 'RM CEREBRO C/ DIFUSIÓN'}
        }
        for _, fila in df.iterrows():
            try:
                nombre_medico = fila['Informe firmado por'].strip()
                medico = User.objects.filter(
                    first_name__in=nombre_medico.split(),
                    last_name__in=nombre_medico.split()
                ).first()
                if not medico:
                    continue
                dni = str(fila['Id. paciente'])[:8]
                nombre_completo = fila['Nombre del paciente'].strip()
                if ',' in nombre_completo:
                    partes = nombre_completo.split(',')
                    apellido = partes[0].strip()
                    nombre = partes[1].strip() if len(partes) > 1 else ''
                else:
                    partes = nombre_completo.split()
                    apellido = partes[0]
                    nombre = ' '.join(partes[1:]) if len(partes) > 1 else ''
                fecha = pd.to_datetime(fila['Fecha de firma final'], dayfirst=True).date()
                mod = fila['Mod.'].strip().upper()
                info_estudio = mapeo_estudios.get(mod)
                estudio = None
                if info_estudio:
                    estudio = Estudios.objects.filter(
                        tipo=info_estudio['tipo'],
                        nombre__iexact=info_estudio['nombre']
                    ).first()
                registros_preview.append({
                    'medico': medico.get_full_name(),
                    'nombre': nombre,
                    'apellido': apellido,
                    'dni': dni,
                    'fecha': str(fecha),
                    'mod': mod,
                    'estudio_base': info_estudio['nombre'] if info_estudio else 'No encontrado',
                    'estudio_tipo': info_estudio['tipo'] if info_estudio else '',
                })
            except Exception:
                continue
        context = self.get_context_data(form=form)
        context['registros'] = registros_preview
        context['registros_json'] = mark_safe(json.dumps(registros_preview, ensure_ascii=False))
        return self.render_to_response(context)

    def confirmar_carga(self, request):
        from liquidacion.models import RegistroEstudio
        try:
            registros_json = request.POST.get('datos_serializados')
            registros = json.loads(registros_json)
            cargados = 0
            errores = 0
            for item in registros:
                try:
                    medico = User.objects.filter(
                        first_name__in=item['medico'].split(),
                        last_name__in=item['medico'].split()
                    ).first()
                    if not medico:
                        continue
                    estudio = Estudios.objects.filter(
                        tipo=item['estudio_tipo'],
                        nombre__iexact=item['estudio_base']
                    ).first()
                    if not estudio:
                        continue
                    fecha = datetime.strptime(item['fecha'], '%Y-%m-%d').date()
                    sesion, _ = SesionContable.objects.get_or_create(
                        mes=fecha.month,
                        año=fecha.year,
                        defaults={'estado': 'ABIERTA'},
                    )
                    with transaction.atomic():
                        registro = RegistroEstudiosPorMedico.objects.create(
                            medico=medico,
                            nombre_paciente=item['nombre'],
                            apellido_paciente=item['apellido'],
                            dni_paciente=item['dni'],
                            fecha_del_informe=fecha,
                            sesion_contable=sesion,
                            cantidad_regiones=estudio.conteo_regiones_default,
                        )
                        RegistroEstudio.objects.create(
                            registro=registro,
                            estudio=estudio,
                            cantidad=1,
                            contexto='SERVICIO',
                        )
                        if medico.rol in ROLES_RESIDENCIA:
                            tiene_eco_general = es_eco_general_real_estudio(estudio)
                            if tiene_eco_general:
                                nuevo_horario = clasificar_horario_residencia_por_proxy(
                                    rol=medico.rol,
                                    fecha_registro=registro.fecha_registro,
                                    tiene_eco_general=True,
                                    fecha_practica=registro.fecha_del_informe,
                                )
                                registro.horario = nuevo_horario or 'NA'
                            else:
                                registro.horario = 'NA'
                        registro.monto_calculado = registro.calcular_monto()
                        registro.save(update_fields=['horario', 'monto_calculado'])
                    cargados += 1
                except Exception:
                    errores += 1
                    continue
            messages.success(request, f"✅ Se cargaron correctamente {cargados} registros. Errores: {errores}")
        except Exception as e:
            messages.error(request, f"❌ Error procesando la carga: {str(e)}")
        return redirect('carga-masiva')


# ============================================================================
# GESTIÓN DE SESIONES CONTABLES — Fase B (Mayo 2026)
# ============================================================================

def _resumen_recalculo_tarifas_pendiente_sesion(sesion):
    if sesion.estado in ['FACTURADA', 'PAGADA']:
        return {
            'disponible': False,
            'total_cambios': 0,
            'registros_cambian': 0,
            'guardias_cambian': 0,
            'mensaje': f'Bloqueado en {sesion.estado}',
        }

    fecha_desde = date(sesion.año, sesion.mes, 1)
    if sesion.mes == 12:
        fecha_hasta = date(sesion.año, 12, 31)
    else:
        fecha_hasta = date(sesion.año, sesion.mes + 1, 1) - timedelta(days=1)

    registros_base = RegistroEstudio.objects.filter(
        registro__sesion_contable=sesion,
        registro__fecha_del_informe__gte=fecha_desde,
        registro__fecha_del_informe__lte=fecha_hasta,
        estudio__grupo_tarifario__isnull=False,
    )

    grupo_ids = list(
        registros_base
        .values_list('estudio__grupo_tarifario_id', flat=True)
        .distinct()
    )
    ultima_recalculo_practicas = (
        HistorialRecalculoTarifaRegistro.objects
        .filter(sesion_contable=sesion)
        .order_by('-fecha_recalculo')
        .values_list('fecha_recalculo', flat=True)
        .first()
    )
    tarifas_practicas = TarifaGrupoTarifario.objects.none()
    if grupo_ids:
        tarifas_practicas = (
            TarifaGrupoTarifario.objects
            .filter(grupo_tarifario_id__in=grupo_ids, vigencia_desde__lte=fecha_hasta)
            .filter(Q(vigencia_hasta__isnull=True) | Q(vigencia_hasta__gte=fecha_desde))
        )
        if ultima_recalculo_practicas:
            tarifas_practicas = tarifas_practicas.filter(fecha_creacion__gt=ultima_recalculo_practicas)

    registros_cambian = 0
    if tarifas_practicas.exists():
        registros_cambian = (
            registros_base
            .filter(estudio__grupo_tarifario_id__in=tarifas_practicas.values('grupo_tarifario_id'))
            .values('registro_id')
            .distinct()
            .count()
        )

    guardias_base = GuardiaPasiva.objects.filter(
        sesion_contable=sesion,
        fecha_guardia__gte=fecha_desde,
        fecha_guardia__lte=fecha_hasta,
    )
    ultima_recalculo_guardias = (
        HistorialRecalculoTarifaGuardiaPasiva.objects
        .filter(sesion_contable=sesion)
        .order_by('-fecha_recalculo')
        .values_list('fecha_recalculo', flat=True)
        .first()
    )
    tarifas_guardias = (
        ConfiguracionGuardiaPasiva.objects
        .filter(vigente_desde__lte=fecha_hasta)
        .filter(Q(vigente_hasta__isnull=True) | Q(vigente_hasta__gte=fecha_desde))
    )
    if ultima_recalculo_guardias:
        tarifas_guardias = tarifas_guardias.filter(fecha_actualizacion__gt=ultima_recalculo_guardias)

    guardias_cambian = guardias_base.count() if tarifas_guardias.exists() else 0

    total_cambios = registros_cambian + guardias_cambian
    return {
        'disponible': total_cambios > 0,
        'total_cambios': total_cambios,
        'registros_cambian': registros_cambian,
        'guardias_cambian': guardias_cambian,
        'mensaje': (
            f'{total_cambios} ajuste(s) pendiente(s)'
            if total_cambios
            else 'Sin ajustes pendientes'
        ),
    }


class SesionContableListView(LoginRequiredMixin, UserPassesTestMixin, ListView):
    """
    Vista administrativa para gestionar el ciclo de vida de SesionContable.
    Permisos: administrativo, jefe_servicio, superuser.
    """
    model = SesionContable
    template_name = 'liquidacion/sesion_contable_list.html'
    context_object_name = 'sesiones'
    paginate_by = 2

    def test_func(self):
        return self.request.user.is_superuser or self.request.user.rol in ['administrativo', 'jefe_servicio']

    def handle_no_permission(self):
        messages.error(self.request, '❌ No tienes permisos para gestionar sesiones contables.')
        return redirect('home')

    def get_queryset(self):
        return SesionContable.objects.prefetch_related(
            'practicas',
            'guardias_pasivas',
            Prefetch(
                'historial_transiciones',
                queryset=HistorialSesionContable.objects.select_related('usuario').order_by('-fecha'),
                to_attr='historial_ordenado',
            ),
        ).order_by('-año', '-mes')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        user = self.request.user

        _MESES = ['', 'Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio',
                  'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre']
        _SIGUIENTE = {
            'ABIERTA': 'REVISION',
            'REVISION': 'CERRADA',
            'CERRADA': 'FACTURADA',
            'FACTURADA': 'PAGADA',
        }
        _ROLES_TRANSICION = {
            'ABIERTA':   ['administrativo', 'jefe_servicio'],
            'REVISION':  ['administrativo', 'jefe_servicio'],
            'CERRADA':   ['administrativo'],
            'FACTURADA': ['administrativo'],
        }

        sesiones_data = []
        for sesion in context['sesiones']:
            total_monto = (
                sesion.practicas.filter(anulado=False).aggregate(t=Sum('monto_calculado'))['t'] or 0
            )
            total_guardias_monto = (
                sesion.guardias_pasivas.aggregate(t=Sum('monto'))['t'] or 0
            )
            siguiente = _SIGUIENTE.get(sesion.estado)
            roles_ok = _ROLES_TRANSICION.get(sesion.estado, [])
            puede = user.is_superuser or user.rol in roles_ok
            gate_preview = {'bloqueantes': [], 'advertencias': []}
            if siguiente:
                gate_preview = evaluar_gate_consistencia_sesion(sesion, siguiente)
            control_eges = resumir_control_eges_sesion(sesion)
            requisito_rrhh = evaluar_requisito_rrhh_para_facturar(sesion)
            checklist_cierre = _enriquecer_checklist_cierre_visual(
                construir_checklist_cierre_sesion(
                    sesion,
                    user=user,
                    gate=gate_preview if siguiente else None,
                    control_eges=control_eges,
                    requisito_rrhh=requisito_rrhh,
                ),
                sesion,
            )
            recalculo_tarifas = _resumen_recalculo_tarifas_pendiente_sesion(sesion)

            historial_ordenado = getattr(sesion, 'historial_ordenado', [])

            sesiones_data.append({
                'sesion': sesion,
                'mes_nombre': _MESES[sesion.mes] if 1 <= sesion.mes <= 12 else str(sesion.mes),
                'count_practicas': sesion.practicas.filter(anulado=False).count(),
                'count_guardias': sesion.guardias_pasivas.count(),
                'total_monto_practicas': total_monto,
                'total_monto_guardias': total_guardias_monto,
                'total_general': total_monto + total_guardias_monto,
                'siguiente_estado': siguiente,
                'puede_transicionar': puede,
                'requiere_motivo': (sesion.estado, siguiente) in [
                    ('CERRADA', 'FACTURADA'),
                    ('FACTURADA', 'PAGADA'),
                ],
                'gate_bloqueantes_count': len(gate_preview['bloqueantes']),
                'gate_advertencias_count': len(gate_preview['advertencias']),
                'gate_bloqueantes_preview': gate_preview['bloqueantes'][:3],
                'gate_advertencias_preview': gate_preview['advertencias'][:3],
                'gate_bloqueantes_accionables': _enriquecer_issues_cierre(
                    [
                        item for item in gate_preview.get('items', [])
                        if item.get('estado') == 'bloqueante'
                    ],
                    sesion=sesion,
                    limite=3,
                ),
                'gate_advertencias_accionables': _enriquecer_issues_cierre(
                    [
                        item for item in gate_preview.get('items', [])
                        if item.get('estado') == 'advertencia'
                    ],
                    sesion=sesion,
                    limite=3,
                ),
                'control_eges': control_eges,
                'checklist_cierre': checklist_cierre,
                'requisito_rrhh': requisito_rrhh,
                'recalculo_tarifas': recalculo_tarifas,
                'ultima_preparacion_rrhh': requisito_rrhh['ultima_preparacion'],
                'historial_reciente': historial_ordenado[:5],
                'historial_count': len(historial_ordenado),
            })

        context['sesiones_data'] = sesiones_data
        return context


class SesionContableRecalculoTarifasView(LoginRequiredMixin, UserPassesTestMixin, TemplateView):
    """Preview y aplicacion controlada de recalculo por tarifas vigentes."""

    template_name = 'liquidacion/sesion_recalculo_tarifas.html'

    def test_func(self):
        return self.request.user.is_superuser or self.request.user.rol in ['administrativo', 'jefe_servicio']

    def handle_no_permission(self):
        messages.error(self.request, 'No tienes permisos para recalcular registros por tarifa.')
        return redirect('home')

    def dispatch(self, request, *args, **kwargs):
        self.sesion = get_object_or_404(SesionContable, pk=kwargs['pk'])
        if self.sesion.estado in ['FACTURADA', 'PAGADA']:
            messages.error(
                request,
                'No se puede recalcular por tarifas en sesiones FACTURADAS o PAGADAS.',
            )
            return redirect('liquidacion:sesiones_list')
        return super().dispatch(request, *args, **kwargs)

    def _rango_default(self):
        fecha_desde = date(self.sesion.año, self.sesion.mes, 1)
        if self.sesion.mes == 12:
            fecha_hasta = date(self.sesion.año, 12, 31)
        else:
            fecha_hasta = date(self.sesion.año, self.sesion.mes + 1, 1) - timedelta(days=1)
        return fecha_desde, fecha_hasta

    def _parse_date(self, value):
        try:
            return datetime.strptime(value, '%Y-%m-%d').date()
        except (TypeError, ValueError):
            return None

    def _build_preview(self, fecha_desde, fecha_hasta):
        registros = (
            RegistroEstudiosPorMedico.objects
            .filter(
                sesion_contable=self.sesion,
                fecha_del_informe__gte=fecha_desde,
                fecha_del_informe__lte=fecha_hasta,
                anulado=False,
            )
            .select_related('medico')
            .prefetch_related('registroestudio_set__estudio__grupo_tarifario')
            .order_by('fecha_del_informe', 'apellido_paciente', 'nombre_paciente', 'id')
        )

        preview = []
        guardias_preview = []
        total_actual = Decimal('0.00')
        total_nuevo = Decimal('0.00')
        total_guardias_actual = Decimal('0.00')
        total_guardias_nuevo = Decimal('0.00')
        for registro in registros:
            monto_actual = registro.monto_calculado or Decimal('0.00')
            monto_nuevo = registro.calcular_monto()
            diferencia = monto_nuevo - monto_actual
            total_actual += monto_actual
            total_nuevo += monto_nuevo
            estudios = ', '.join(
                rel.estudio.nombre
                for rel in registro.registroestudio_set.all()
            )
            preview.append({
                'registro': registro,
                'estudios': estudios,
                'monto_actual': monto_actual,
                'monto_nuevo': monto_nuevo,
                'diferencia': diferencia,
                'cambia': diferencia != Decimal('0.00'),
            })

        guardias = (
            GuardiaPasiva.objects
            .filter(
                sesion_contable=self.sesion,
                fecha_guardia__gte=fecha_desde,
                fecha_guardia__lte=fecha_hasta,
            )
            .select_related('medico')
            .order_by('fecha_guardia', 'medico__last_name', 'medico__first_name', 'id')
        )
        for guardia in guardias:
            monto_actual = guardia.monto or Decimal('0.00')
            monto_nuevo = ConfiguracionGuardiaPasiva.get_config(guardia.fecha_guardia).monto_vigente
            diferencia = monto_nuevo - monto_actual
            total_guardias_actual += monto_actual
            total_guardias_nuevo += monto_nuevo
            guardias_preview.append({
                'guardia': guardia,
                'monto_actual': monto_actual,
                'monto_nuevo': monto_nuevo,
                'diferencia': diferencia,
                'cambia': diferencia != Decimal('0.00'),
            })

        total_actual_general = total_actual + total_guardias_actual
        total_nuevo_general = total_nuevo + total_guardias_nuevo

        return {
            'items': preview,
            'guardias': guardias_preview,
            'total_registros': len(preview),
            'total_guardias': len(guardias_preview),
            'total_cambian_registros': sum(1 for item in preview if item['cambia']),
            'total_cambian_guardias': sum(1 for item in guardias_preview if item['cambia']),
            'total_cambian': (
                sum(1 for item in preview if item['cambia'])
                + sum(1 for item in guardias_preview if item['cambia'])
            ),
            'total_actual': total_actual_general,
            'total_nuevo': total_nuevo_general,
            'total_practicas_actual': total_actual,
            'total_practicas_nuevo': total_nuevo,
            'total_guardias_actual': total_guardias_actual,
            'total_guardias_nuevo': total_guardias_nuevo,
            'diferencia_total': total_nuevo_general - total_actual_general,
        }

    def _contexto(self, request, errors=None, preview=None, fecha_desde=None, fecha_hasta=None, motivo=''):
        default_desde, default_hasta = self._rango_default()
        return {
            'sesion': self.sesion,
            'fecha_desde': fecha_desde or default_desde,
            'fecha_hasta': fecha_hasta or default_hasta,
            'motivo': motivo or f'Recalculo por actualizacion de tarifas vigentes - {self.sesion.mes}/{self.sesion.año}',
            'preview': preview,
            'errors_preview': errors or [],
        }

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context.update(self._contexto(self.request))
        return context

    def post(self, request, *args, **kwargs):
        fecha_desde = self._parse_date(request.POST.get('fecha_desde'))
        fecha_hasta = self._parse_date(request.POST.get('fecha_hasta'))
        motivo = (request.POST.get('motivo') or '').strip()
        confirmar = request.POST.get('confirmar') == '1'
        errors = []

        if not fecha_desde:
            errors.append('Indica una fecha desde valida.')
        if not fecha_hasta:
            errors.append('Indica una fecha hasta valida.')
        if fecha_desde and fecha_hasta and fecha_hasta < fecha_desde:
            errors.append('La fecha hasta no puede ser anterior a fecha desde.')
        if not motivo:
            errors.append('Indica un motivo para auditar el recalculo.')

        preview = None
        if not errors:
            preview = self._build_preview(fecha_desde, fecha_hasta)
            if preview['total_registros'] == 0 and preview['total_guardias'] == 0:
                errors.append('No hay registros ni guardias pasivas en el rango seleccionado.')

        if errors or not confirmar:
            return self.render_to_response(
                self._contexto(
                    request,
                    errors=errors,
                    preview=None if errors else preview,
                    fecha_desde=fecha_desde,
                    fecha_hasta=fecha_hasta,
                    motivo=motivo,
                )
            )

        aplicados = 0
        guardias_aplicadas = 0
        sin_cambios = 0
        guardias_sin_cambios = 0
        with transaction.atomic():
            registros = (
                RegistroEstudiosPorMedico.objects
                .select_for_update()
                .filter(
                    sesion_contable=self.sesion,
                    fecha_del_informe__gte=fecha_desde,
                    fecha_del_informe__lte=fecha_hasta,
                    anulado=False,
                )
                .select_related('medico')
                .prefetch_related('registroestudio_set__estudio__grupo_tarifario')
                .order_by('id')
            )
            for registro in registros:
                monto_anterior = registro.monto_calculado or Decimal('0.00')
                monto_nuevo = registro.calcular_monto()
                if monto_nuevo == monto_anterior:
                    sin_cambios += 1
                    continue

                motivo_sistema = (
                    f'Recalculo por tarifas vigentes. '
                    f'Rango: {fecha_desde:%d/%m/%Y} - {fecha_hasta:%d/%m/%Y}. '
                    f'Monto: ${monto_anterior} -> ${monto_nuevo}. {motivo}'
                )
                registro.monto_calculado = monto_nuevo
                registro.modificado_por = request.user
                registro.fecha_modificacion = now()
                registro.motivo_modificacion = motivo_sistema
                registro.save(update_fields=[
                    'monto_calculado',
                    'modificado_por',
                    'fecha_modificacion',
                    'motivo_modificacion',
                ])
                HistorialRecalculoTarifaRegistro.objects.create(
                    sesion_contable=self.sesion,
                    registro=registro,
                    fecha_desde=fecha_desde,
                    fecha_hasta=fecha_hasta,
                    monto_anterior=monto_anterior,
                    monto_nuevo=monto_nuevo,
                    diferencia=monto_nuevo - monto_anterior,
                    motivo=motivo_sistema,
                    recalculado_por=request.user,
                )
                aplicados += 1

            guardias = (
                GuardiaPasiva.objects
                .select_for_update()
                .filter(
                    sesion_contable=self.sesion,
                    fecha_guardia__gte=fecha_desde,
                    fecha_guardia__lte=fecha_hasta,
                )
                .order_by('id')
            )
            for guardia in guardias:
                monto_anterior = guardia.monto or Decimal('0.00')
                monto_nuevo = ConfiguracionGuardiaPasiva.get_config(guardia.fecha_guardia).monto_vigente
                if monto_nuevo == monto_anterior:
                    guardias_sin_cambios += 1
                    continue

                motivo_sistema = (
                    f'Recalculo por tarifas vigentes de guardia pasiva. '
                    f'Rango: {fecha_desde:%d/%m/%Y} - {fecha_hasta:%d/%m/%Y}. '
                    f'Monto: ${monto_anterior} -> ${monto_nuevo}. {motivo}'
                )
                guardia.monto = monto_nuevo
                guardia.observaciones = (
                    f"{guardia.observaciones}\n{motivo_sistema}".strip()
                    if guardia.observaciones else motivo_sistema
                )
                guardia.save(update_fields=['monto', 'observaciones'])
                HistorialRecalculoTarifaGuardiaPasiva.objects.create(
                    sesion_contable=self.sesion,
                    guardia=guardia,
                    fecha_desde=fecha_desde,
                    fecha_hasta=fecha_hasta,
                    monto_anterior=monto_anterior,
                    monto_nuevo=monto_nuevo,
                    diferencia=monto_nuevo - monto_anterior,
                    motivo=motivo_sistema,
                    recalculado_por=request.user,
                )
                guardias_aplicadas += 1

        messages.success(
            request,
            (
                f'Recalculo por tarifas aplicado: {aplicados} registro(s) y '
                f'{guardias_aplicadas} guardia(s) actualizada(s). '
                f'Sin cambios: {sin_cambios} registro(s), {guardias_sin_cambios} guardia(s).'
            ),
        )
        return redirect('liquidacion:sesiones_list')


class PreparacionLiquidacionRRHHPreviewView(LoginRequiredMixin, UserPassesTestMixin, FormView):
    """Preview D1 de liquidacion residencia para RRHH. No envia emails."""

    template_name = 'liquidacion/preparacion_rrhh_preview.html'
    form_class = PreparacionLiquidacionRRHHForm

    def test_func(self):
        return self.request.user.is_superuser or self.request.user.rol in ['administrativo', 'jefe_servicio']

    def handle_no_permission(self):
        messages.error(self.request, 'No tienes permisos para preparar liquidacion RRHH.')
        return redirect('home')

    def dispatch(self, request, *args, **kwargs):
        self.sesion = get_object_or_404(SesionContable, pk=kwargs['pk'])
        if self.sesion.estado not in ['CERRADA', 'FACTURADA', 'PAGADA']:
            messages.error(
                request,
                'La preparacion RRHH solo esta disponible desde sesiones CERRADA, FACTURADA o PAGADA.',
            )
            return redirect('liquidacion:sesiones_list')
        self.snapshot = construir_snapshot_liquidacion_rrhh(self.sesion)
        self.snapshot_hash = calcular_hash_snapshot(self.snapshot)
        return super().dispatch(request, *args, **kwargs)

    def get_initial(self):
        initial = super().get_initial()
        initial.update({
            'asunto': asunto_default_rrhh(self.sesion),
            'cuerpo': cuerpo_default_rrhh(self.snapshot),
        })
        return initial

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        validaciones_accionables = _enriquecer_issues_cierre(
            self.snapshot['validaciones'].get('items', []),
            sesion=self.sesion,
        )
        context.update({
            'sesion': self.sesion,
            'snapshot': self.snapshot,
            'snapshot_hash': self.snapshot_hash,
            'proxima_version': proxima_version_preparacion_rrhh(self.sesion),
            'preparaciones_previas': (
                PreparacionLiquidacionRRHH.objects
                .filter(sesion_contable=self.sesion)
                .select_related('creado_por', 'actualizado_por')
                .order_by('-version')[:5]
            ),
            'validaciones_accionables_bloqueantes': [
                item for item in validaciones_accionables
                if item.get('estado') == 'bloqueante'
            ],
            'validaciones_accionables_advertencias': [
                item for item in validaciones_accionables
                if item.get('estado') == 'advertencia'
            ],
        })
        return context

    @transaction.atomic
    def form_valid(self, form):
        estado = (
            PreparacionLiquidacionRRHH.ESTADO_PREPARADO
            if 'guardar_preparado' in self.request.POST
            else PreparacionLiquidacionRRHH.ESTADO_BORRADOR
        )
        destinatarios = form.cleaned_data['destinatarios']
        bloqueantes = list(self.snapshot['validaciones']['bloqueantes'])

        if estado == PreparacionLiquidacionRRHH.ESTADO_PREPARADO and not destinatarios:
            bloqueantes.append('PREPARADO requiere al menos un destinatario.')

        if estado == PreparacionLiquidacionRRHH.ESTADO_PREPARADO and not self.snapshot['sesion']['requiere_rrhh']:
            bloqueantes.append('No hay practicas de residencia para preparar RRHH en esta sesion.')

        if estado == PreparacionLiquidacionRRHH.ESTADO_PREPARADO and bloqueantes:
            for issue in bloqueantes[:5]:
                messages.error(self.request, issue)
            return self.form_invalid(form)

        preparacion = PreparacionLiquidacionRRHH.objects.create(
            sesion_contable=self.sesion,
            version=proxima_version_preparacion_rrhh(self.sesion),
            estado=estado,
            destinatarios_json=destinatarios,
            cc_json=form.cleaned_data['cc'],
            asunto=form.cleaned_data['asunto'],
            cuerpo=form.cleaned_data['cuerpo'],
            resumen_json=self.snapshot,
            snapshot_hash=self.snapshot_hash,
            creado_por=self.request.user,
            actualizado_por=self.request.user,
        )
        messages.success(
            self.request,
            f'Preparacion RRHH v{preparacion.version} guardada como {preparacion.estado}.',
        )
        return redirect('liquidacion:preparacion_rrhh_preview', pk=self.sesion.pk)


@login_required
def sesion_contable_transicion(request, pk):
    """
    Avanza el estado de una SesionContable al siguiente paso del flujo.
    Solo acepta POST. Valida permisos por rol y estado actual.

    Flujo: ABIERTA → REVISION → CERRADA → FACTURADA → PAGADA
    - jefe_servicio puede avanzar hasta CERRADA
    - administrativo/superuser pueden avanzar hasta PAGADA
    """
    if request.method != 'POST':
        return redirect('liquidacion:sesiones_list')

    sesion = get_object_or_404(SesionContable, pk=pk)
    user = request.user

    FLUJO = {
        'ABIERTA':   {'siguiente': 'REVISION',   'roles': ['administrativo', 'jefe_servicio']},
        'REVISION':  {'siguiente': 'CERRADA',     'roles': ['administrativo', 'jefe_servicio']},
        'CERRADA':   {'siguiente': 'FACTURADA',   'roles': ['administrativo']},
        'FACTURADA': {'siguiente': 'PAGADA',      'roles': ['administrativo']},
    }

    config = FLUJO.get(sesion.estado)
    if not config:
        messages.error(request, '❌ Esta sesión ya está en estado final (PAGADA).')
        return redirect('liquidacion:sesiones_list')

    if not (user.is_superuser or user.rol in config['roles']):
        messages.error(request, '❌ No tienes permisos para esta transición.')
        return redirect('liquidacion:sesiones_list')

    motivo = (request.POST.get('motivo') or '').strip()
    transicion_financiera = (sesion.estado, config['siguiente']) in [
        ('CERRADA', 'FACTURADA'),
        ('FACTURADA', 'PAGADA'),
    ]
    if transicion_financiera and not motivo:
        messages.error(
            request,
            '❌ Debes indicar un motivo para transiciones financieras (FACTURADA/PAGADA).'
        )
        return redirect('liquidacion:sesiones_list')

    auditoria = evaluar_gate_consistencia_sesion(sesion, config['siguiente'])
    if auditoria['bloqueantes']:
        detalle = ' | '.join(auditoria['bloqueantes'][:3])
        if len(auditoria['bloqueantes']) > 3:
            detalle += f" | +{len(auditoria['bloqueantes']) - 3} inconsistencia(s)"
        messages.error(
            request,
            f"❌ No se puede pasar a {config['siguiente']}. Inconsistencias bloqueantes: {detalle}"
        )
        return redirect('liquidacion:sesiones_list')

    if auditoria['advertencias']:
        detalle = ' | '.join(auditoria['advertencias'][:2])
        if len(auditoria['advertencias']) > 2:
            detalle += f" | +{len(auditoria['advertencias']) - 2} advertencia(s)"
        messages.warning(
            request,
            f"⚠️ Advertencias de consistencia: {detalle}"
        )

    if config['siguiente'] == 'FACTURADA':
        requisito_rrhh = evaluar_requisito_rrhh_para_facturar(sesion)
        if not requisito_rrhh['ok']:
            messages.error(
                request,
                f"No se puede pasar a FACTURADA. {requisito_rrhh['mensaje']}",
            )
            return redirect('liquidacion:sesiones_list')

    from django.utils import timezone as tz

    estado_anterior = sesion.estado
    sesion.estado = config['siguiente']

    if config['siguiente'] == 'CERRADA':
        sesion.fecha_cierre = tz.now()
        sesion.cerrada_por = user
    elif config['siguiente'] == 'FACTURADA':
        sesion.fecha_facturacion = tz.now()
    elif config['siguiente'] == 'PAGADA':
        sesion.fecha_pago = tz.now()

    sesion.save()

    HistorialSesionContable.objects.create(
        sesion_contable=sesion,
        estado_anterior=estado_anterior,
        estado_nuevo=sesion.estado,
        usuario=user,
        motivo=motivo,
        origen=HistorialSesionContable.ORIGEN_WEB,
    )

    meses = ['', 'Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio',
             'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre']
    mes_nombre = meses[sesion.mes] if 1 <= sesion.mes <= 12 else str(sesion.mes)
    messages.success(
        request,
        f'✅ {mes_nombre} {sesion.año}: {estado_anterior} → {sesion.estado}'
    )
    return redirect('liquidacion:sesiones_list')
