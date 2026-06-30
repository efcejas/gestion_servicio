"""
services.py — Generación de documentos del módulo liquidacion.

Encapsula la lógica de construcción de PDF y Excel, separada del ciclo HTTP.
Las funciones retornan buffers (BytesIO) listos para ser enviados como descarga.
"""
import io
from datetime import datetime

from django.utils import timezone
from django.db.models import Prefetch, Q

from .grupo_tarifario_mapping import es_eco_general_real_estudio
from .models import (
    CorreccionPacsRegistro,
    Estudios,
    GuardiaPasiva,
    ReglaDescuentoResidencia,
    RegistroEstudiosPorMedico,
)


ROLES_RESIDENCIA = {
    'medico_residente',
    'jefe_residentes',
    'instructor_residentes',
}

CAMPO_REGLA_DESCUENTO_POR_ROL = {
    'medico_residente': 'aplica_medico_residente',
    'jefe_residentes': 'aplica_jefe_residentes',
    'instructor_residentes': 'aplica_instructor_residentes',
}


def adjuntar_ultima_correccion_pacs(registros):
    """Adjunta la ultima correccion PACS a cada registro, sin recalcular montos."""
    registros = list(registros)
    registro_ids = [registro.pk for registro in registros]
    correcciones_por_registro = {}

    if registro_ids:
        correcciones = (
            CorreccionPacsRegistro.objects
            .filter(registro_id__in=registro_ids)
            .select_related('corregido_por')
            .order_by('registro_id', '-fecha_correccion')
        )
        for correccion in correcciones:
            if correccion.registro_id not in correcciones_por_registro:
                correcciones_por_registro[correccion.registro_id] = correccion

    for registro in registros:
        correccion = correcciones_por_registro.get(registro.pk)
        registro.correccion_pacs_info = correccion
        registro.tiene_correccion_pacs = correccion is not None
        registro.impacto_correccion_pacs = (
            correccion.monto_nuevo - correccion.monto_anterior
            if correccion
            else 0
        )

    return registros


def _resultado_descuento_residencia(aplica, fuente, regla_id=None, motivo=''):
    return {
        'aplica': bool(aplica),
        'fuente': fuente,
        'regla_id': regla_id,
        'motivo': motivo,
    }


def _regla_vigente_para_estudio(estudio, fecha_ref):
    return (
        ReglaDescuentoResidencia.objects
        .filter(
            estudio=estudio,
            activo=True,
            vigencia_desde__lte=fecha_ref,
        )
        .filter(Q(vigencia_hasta__isnull=True) | Q(vigencia_hasta__gte=fecha_ref))
        .order_by('-vigencia_desde', '-id')
        .first()
    )


def _regla_vigente_para_grupo(grupo_tarifario, fecha_ref):
    if not grupo_tarifario:
        return None

    return (
        ReglaDescuentoResidencia.objects
        .filter(
            grupo_tarifario=grupo_tarifario,
            activo=True,
            vigencia_desde__lte=fecha_ref,
        )
        .filter(Q(vigencia_hasta__isnull=True) | Q(vigencia_hasta__gte=fecha_ref))
        .order_by('-vigencia_desde', '-id')
        .first()
    )


def _resultado_desde_regla(regla, rol, fuente):
    campo_rol = CAMPO_REGLA_DESCUENTO_POR_ROL[rol]
    aplica = getattr(regla, campo_rol)
    return _resultado_descuento_residencia(
        aplica=aplica,
        fuente=fuente,
        regla_id=regla.id,
        motivo=f"Regla explicita por {fuente}.",
    )


def estudio_aplica_descuento_residencia(estudio, rol, fecha=None):
    """Resuelve elegibilidad de descuento residencia sin modificar calculos.

    Precedencia: regla por estudio > regla por grupo tarifario > fallback legado.
    """
    if rol not in ROLES_RESIDENCIA:
        return _resultado_descuento_residencia(
            aplica=False,
            fuente='rol_no_residencia',
            motivo='El rol no pertenece a residencia.',
        )

    fecha_ref = fecha or timezone.localdate()
    regla_estudio = _regla_vigente_para_estudio(estudio, fecha_ref)
    if regla_estudio:
        return _resultado_desde_regla(regla_estudio, rol, fuente='estudio')

    regla_grupo = _regla_vigente_para_grupo(getattr(estudio, 'grupo_tarifario', None), fecha_ref)
    if regla_grupo:
        return _resultado_desde_regla(regla_grupo, rol, fuente='grupo')

    aplica_fallback = es_eco_general_real_estudio(estudio)
    return _resultado_descuento_residencia(
        aplica=aplica_fallback,
        fuente='fallback_legado',
        motivo='Fallback legado ECO general real.' if aplica_fallback else 'Fallback legado sin descuento.',
    )


def clasificar_horario_residencia_por_proxy(rol, fecha_registro, tiene_eco_general):
    """
    Clasifica horario INTRA/EXTRA para residencia usando fecha_registro como proxy.

    Retorna:
        'INTRA', 'EXTRA' o None (cuando no aplica la regla)
    """
    if rol not in ROLES_RESIDENCIA:
        return None
    if not tiene_eco_general:
        return None
    if not fecha_registro:
        return None

    fecha_local = timezone.localtime(fecha_registro)

    # Feriados institucionales centralizados en control_guardias.
    from control_guardias.models import Feriado

    if fecha_local.weekday() >= 5:
        return 'EXTRA'
    if Feriado.objects.filter(fecha=fecha_local.date()).exists():
        return 'EXTRA'

    return 'INTRA' if 8 <= fecha_local.hour < 17 else 'EXTRA'


def generar_buffer_pdf_liquidacion():
    """
    Genera el PDF de liquidación completa (todos los médicos).

    Retorna:
        BytesIO posicionado al inicio, listo para su descarga.
    """
    from django.contrib.auth import get_user_model
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

    User = get_user_model()

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    Story = []
    styles = getSampleStyleSheet()

    titulo = Paragraph("<b>Liquidación de Estudios por Médico - v2.0</b>", styles["Title"])
    fecha_generacion = Paragraph(
        f"<b>Fecha de generación:</b> {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}",
        styles["Normal"],
    )
    Story.append(titulo)
    Story.append(fecha_generacion)
    Story.append(Spacer(1, 10))

    medicos = User.objects.filter(es_medico=True)
    for medico in medicos:
        encabezado_medico = Paragraph(
            f"<b>Médico:</b> {medico.get_full_name()}", styles["Heading2"]
        )
        Story.append(encabezado_medico)
        Story.append(Spacer(1, 5))

        registros = RegistroEstudiosPorMedico.objects.filter(
            medico=medico
        ).prefetch_related("estudio").order_by("-fecha_registro")

        if registros.exists():
            data = [["Fecha", "Paciente", "Estudios", "Regiones", "OS", "Horario", "Monto"]]
            total_regiones = 0
            total_monto = 0

            for registro in registros:
                estudios_lista = registro.estudio.all()
                estudios_texto = ", ".join(e.nombre for e in estudios_lista) if estudios_lista.exists() else "N/A"
                regiones = registro.cantidad_regiones
                total_regiones += regiones
                total_monto += registro.monto_calculado

                data.append([
                    registro.fecha_del_informe.strftime('%d/%m/%Y') if registro.fecha_del_informe else "N/A",
                    f"{registro.apellido_paciente.upper()} {registro.nombre_paciente.upper()}",
                    estudios_texto,
                    str(regiones),
                    registro.get_tipo_obra_social_display(),
                    registro.get_horario_display(),
                    f"${registro.monto_calculado:,.2f}"
                ])

            data.append(["", "", "TOTAL PRÁCTICAS", str(total_regiones), "", "", f"${total_monto:,.2f}"])

            tabla = Table(data, colWidths=[60, 100, 120, 50, 60, 50, 70])
            tabla.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#003366")),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('ALIGN', (3, 1), (3, -1), 'CENTER'),
                ('ALIGN', (6, 1), (6, -1), 'RIGHT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 6),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor("#e0e0e0")),
                ('LEFTPADDING', (0, 0), (-1, -1), 3),
                ('RIGHTPADDING', (0, 0), (-1, -1), 3),
                ('TOPPADDING', (0, 0), (-1, -1), 3),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
            ]))
            Story.append(tabla)
        else:
            Story.append(Paragraph(
                "<i>No hay prácticas registradas para este médico.</i>", styles["Normal"]))

        Story.append(Spacer(1, 10))

        guardias = GuardiaPasiva.objects.filter(medico=medico).order_by('-fecha_guardia')
        if guardias.exists():
            encabezado_guardias = Paragraph("<b>Guardias Pasivas</b>", styles["Heading3"])
            Story.append(encabezado_guardias)
            Story.append(Spacer(1, 3))

            data_guardias = [["Fecha", "Tipo Guardia", "Monto", "Observaciones"]]
            total_guardias = 0

            for guardia in guardias:
                total_guardias += guardia.monto
                data_guardias.append([
                    guardia.fecha_guardia.strftime('%d/%m/%Y'),
                    guardia.get_tipo_guardia_display(),
                    f"${guardia.monto:,.2f}",
                    guardia.observaciones or ""
                ])

            data_guardias.append(["", "TOTAL GUARDIAS", f"${total_guardias:,.2f}", ""])

            tabla_guardias = Table(data_guardias, colWidths=[70, 100, 80, 260])
            tabla_guardias.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#70AD47")),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('ALIGN', (2, 1), (2, -1), 'RIGHT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 6),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor("#E2EFDA")),
                ('LEFTPADDING', (0, 0), (-1, -1), 3),
                ('RIGHTPADDING', (0, 0), (-1, -1), 3),
                ('TOPPADDING', (0, 0), (-1, -1), 3),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
            ]))
            Story.append(tabla_guardias)
            Story.append(Spacer(1, 10))

            if registros.exists():
                total_general = total_monto + total_guardias
                total_general_text = Paragraph(
                    f"<b>TOTAL GENERAL (Prácticas + Guardias): ${total_general:,.2f}</b>",
                    styles["Heading3"]
                )
                Story.append(total_general_text)

        Story.append(Spacer(1, 15))

    doc.build(Story)
    buffer.seek(0)
    return buffer


def generar_buffer_excel_liquidacion(medico=None, mes=None, año=None):
    """
    Genera el Excel de liquidación.

    Parámetros:
        medico : instancia de User (o None para todos los médicos)
        mes    : str o int (1-12) o None
        año    : str o int (4 dígitos) o None

    Retorna:
        (BytesIO, str): buffer posicionado al inicio y nombre del médico
                        para usar en el Content-Disposition del filename.
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    medico_id = medico.id if medico else None
    nombre_medico = f"{medico.first_name}_{medico.last_name}" if medico else "todos_los_medicos"

    registros = RegistroEstudiosPorMedico.objects.prefetch_related(
        Prefetch('estudio', queryset=Estudios.objects.all())
    ).distinct()

    if medico_id:
        registros = registros.filter(medico_id=medico_id)
    if mes and año:
        registros = registros.filter(fecha_del_informe__year=int(año), fecha_del_informe__month=int(mes))

    guardias = GuardiaPasiva.objects.none()
    if medico_id and mes and año:
        guardias = GuardiaPasiva.objects.filter(
            medico_id=medico_id,
            fecha_guardia__year=int(año),
            fecha_guardia__month=int(mes)
        ).order_by('fecha_guardia')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Liquidación Completa"

    registros = adjuntar_ultima_correccion_pacs(registros.order_by('-fecha_del_informe'))

    headers_practicas = [
        "Fecha", "Paciente", "DNI", "Estudios",
        "Tipo", "Regiones", "Obra Social", "Horario", "Monto", "Bonus",
        "Ajuste PACS", "Tipo ajuste PACS", "Horario anterior", "Horario nuevo",
        "Monto anterior PACS", "Monto nuevo PACS", "Hora PACS", "Observacion ajuste PACS",
    ]
    ws.append(headers_practicas)

    for cell in ws[1]:
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

    total_regiones = 0
    total_monto_practicas = 0

    for registro in registros:
        estudios_lista = registro.estudio.all()
        estudios_nombres = ", ".join(e.nombre for e in estudios_lista) if estudios_lista.exists() else "N/A"
        tipos_estudios = ", ".join(e.get_tipo_display() for e in estudios_lista) if estudios_lista.exists() else "N/A"
        bonus_icon = "⚡ SÍ" if registro.paciente_internado else ""
        correccion = getattr(registro, 'correccion_pacs_info', None)

        ws.append([
            registro.fecha_del_informe.strftime("%d/%m/%Y"),
            f"{registro.apellido_paciente.upper()} {registro.nombre_paciente.upper()}",
            registro.dni_paciente,
            estudios_nombres,
            tipos_estudios,
            registro.cantidad_regiones,
            registro.get_tipo_obra_social_display(),
            registro.get_horario_display(),
            float(registro.monto_calculado),
            bonus_icon,
            "SI" if correccion else "NO",
            correccion.get_tipo_correccion_display() if correccion else "",
            correccion.horario_anterior if correccion and correccion.horario_anterior else "",
            correccion.horario_nuevo if correccion and correccion.horario_nuevo else "",
            float(correccion.monto_anterior) if correccion else "",
            float(correccion.monto_nuevo) if correccion else "",
            correccion.hora_pacs.strftime("%H:%M") if correccion and correccion.hora_pacs else "",
            correccion.observacion if correccion else "",
        ])

        total_regiones += registro.cantidad_regiones
        total_monto_practicas += registro.monto_calculado

    ws.append([])
    totales_practicas_row = ws.max_row + 1
    ws.append(["", "", "", "", "SUBTOTAL PRÁCTICAS", total_regiones, "", "", float(total_monto_practicas), "", "", "", "", "", "", "", "", ""])

    for cell in ws[totales_practicas_row]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

    ws.cell(row=totales_practicas_row, column=9).number_format = '$#,##0.00'

    total_monto_guardias = 0
    header_row = None
    totales_guardias_row = None

    if guardias.exists():
        ws.append([])
        ws.append([])

        headers_guardias = ["Fecha", "Tipo de Guardia", "Monto", "Observaciones"]
        header_row = ws.max_row + 1
        ws.append(headers_guardias)

        for col_num, cell in enumerate(ws[header_row], 1):
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")

        for guardia in guardias:
            ws.append([
                guardia.fecha_guardia.strftime("%d/%m/%Y"),
                guardia.get_tipo_guardia_display(),
                float(guardia.monto),
                guardia.observaciones or ""
            ])
            total_monto_guardias += guardia.monto

        ws.append([])
        totales_guardias_row = ws.max_row + 1
        ws.append(["", "SUBTOTAL GUARDIAS", float(total_monto_guardias), ""])

        for cell in ws[totales_guardias_row]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

        ws.cell(row=totales_guardias_row, column=3).number_format = '$#,##0.00'

    ws.append([])
    ws.append([])
    total_general_row = ws.max_row + 1
    ws.append(["", "", "", "", "TOTAL GENERAL", "", "", "", float(total_monto_practicas + total_monto_guardias), "", "", "", "", "", "", "", "", ""])

    for cell in ws[total_general_row]:
        cell.font = Font(bold=True, size=12)
        cell.fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")

    ws.cell(row=total_general_row, column=9).number_format = '$#,##0.00'

    for row in range(2, totales_practicas_row):
        ws.cell(row=row, column=9).number_format = '$#,##0.00'
        ws.cell(row=row, column=15).number_format = '$#,##0.00'
        ws.cell(row=row, column=16).number_format = '$#,##0.00'

    if guardias.exists() and header_row and totales_guardias_row:
        guardias_start = header_row + 1
        guardias_end = totales_guardias_row - 1
        for row in range(guardias_start, guardias_end + 1):
            ws.cell(row=row, column=3).number_format = '$#,##0.00'

    for column_cells in ws.columns:
        length = 0
        column_letter = column_cells[0].column_letter
        for cell in column_cells:
            try:
                if cell.value:
                    cell_length = len(str(cell.value))
                    if cell_length > length:
                        length = cell_length
            except Exception:
                pass
        ws.column_dimensions[column_letter].width = length + 3

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer, nombre_medico
