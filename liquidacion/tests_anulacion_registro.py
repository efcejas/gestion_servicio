from datetime import date
from decimal import Decimal
from io import BytesIO

from django.contrib.auth import get_user_model
from django.test import TestCase
from django.urls import reverse
from openpyxl import load_workbook

from eges_import.models import ImportBatch

from .models import (
    Estudios,
    PreparacionLiquidacionRRHH,
    RegistroEstudio,
    RegistroEstudiosPorMedico,
    SesionContable,
)
from .services import generar_buffer_excel_liquidacion
from .services_eges import construir_preview_cruce_liquidacion_eges
from .services_rrhh import (
    calcular_hash_snapshot,
    construir_snapshot_liquidacion_rrhh,
    evaluar_requisito_rrhh_para_facturar,
)


User = get_user_model()


class AnulacionRegistroEstudioTest(TestCase):
    def setUp(self):
        self.admin = User.objects.create_user(
            username='admin_anulacion',
            password='x',
            rol='administrativo',
            perfil_completo=True,
        )
        self.residente = User.objects.create_user(
            username='res_anulacion',
            password='x',
            rol='medico_residente',
            first_name='Ana',
            last_name='Residente',
            perfil_completo=True,
        )
        self.sesion = SesionContable.objects.create(mes=7, año=2026, estado='REVISION')
        self.estudio = Estudios.objects.create(
            nombre='ECOGRAFIA COMPLETA DE ABDOMEN',
            tipo='ECO',
            conteo_regiones=1,
            conteo_regiones_default=1,
            precio_cober=Decimal('1000.00'),
            precio_otras_os=Decimal('1000.00'),
            activo=True,
        )
        self.registro = RegistroEstudiosPorMedico.objects.create(
            sesion_contable=self.sesion,
            medico=self.residente,
            nombre_paciente='Elena',
            apellido_paciente='Paciente Anulado',
            dni_paciente='12345678',
            fecha_del_informe=date(2026, 7, 10),
            tipo_obra_social='OTRAS_OS',
            horario='EXTRA',
            cantidad_regiones=1,
            monto_calculado=Decimal('1000.00'),
        )
        RegistroEstudio.objects.create(
            registro=self.registro,
            estudio=self.estudio,
            cantidad=1,
            contexto='SERVICIO',
        )
        RegistroEstudiosPorMedico.objects.filter(pk=self.registro.pk).update(
            monto_calculado=Decimal('1000.00'),
        )
        self.registro.refresh_from_db()

    def _anular(self, user=None, motivo='No corresponde segun control administrativo.'):
        self.client.force_login(user or self.admin)
        return self.client.post(
            reverse('liquidacion:registroestudios_anular', args=[self.registro.pk]),
            {'motivo': motivo},
        )

    def test_anulacion_conserva_monto_y_registra_trazabilidad(self):
        response = self._anular()

        self.assertEqual(response.status_code, 302)
        self.registro.refresh_from_db()
        self.assertTrue(self.registro.anulado)
        self.assertEqual(self.registro.monto_calculado, Decimal('1000.00'))
        self.assertEqual(self.registro.anulado_por, self.admin)
        self.assertIsNotNone(self.registro.fecha_anulacion)
        self.assertEqual(
            self.registro.motivo_anulacion,
            'No corresponde segun control administrativo.',
        )

    def test_residente_no_puede_anular(self):
        response = self._anular(user=self.residente)

        self.assertEqual(response.status_code, 302)
        self.registro.refresh_from_db()
        self.assertFalse(self.registro.anulado)

    def test_no_permite_anular_en_sesion_facturada(self):
        self.sesion.estado = 'FACTURADA'
        self.sesion.save(update_fields=['estado'])

        self._anular()

        self.registro.refresh_from_db()
        self.assertFalse(self.registro.anulado)

    def test_excel_personal_excluye_registro_anulado(self):
        self._anular()
        self.client.force_login(self.residente)

        response = self.client.get(
            reverse('liquidacion:exportar_excel_mis_registros'),
            {'mes': 7, 'año': 2026},
        )

        self.assertEqual(response.status_code, 200)
        workbook = load_workbook(BytesIO(response.content), data_only=False)
        practicas = workbook['Practicas']
        pacientes = [practicas.cell(row=row, column=3).value for row in range(2, practicas.max_row + 1)]
        self.assertNotIn('Paciente Anulado, Elena', pacientes)
        self.assertEqual(workbook['Resumen']['B3'].value, 0)

    def test_excel_administrativo_excluye_registro_anulado(self):
        self._anular()

        buffer, _ = generar_buffer_excel_liquidacion(
            medico=self.residente,
            mes=7,
            año=2026,
        )

        workbook = load_workbook(buffer, data_only=False)
        values = [cell.value for row in workbook.active.iter_rows() for cell in row]
        self.assertNotIn('PACIENTE ANULADO ELENA', values)
        self.assertNotIn(1000.0, values)
    def test_cruce_eges_excluye_registro_anulado(self):
        self._anular()
        batch = ImportBatch.objects.create(
            usuario=self.admin,
            archivo_nombre='Turnos-Julio-ECO.xls',
        )

        preview = construir_preview_cruce_liquidacion_eges(self.sesion, batch)

        self.assertEqual(preview['resumen']['total'], 0)
        self.assertEqual(preview['resultados'], [])

    def test_snapshot_rrhh_excluye_registro_anulado(self):
        self._anular()
        self.sesion.estado = 'CERRADA'
        self.sesion.save(update_fields=['estado'])

        snapshot = construir_snapshot_liquidacion_rrhh(self.sesion)

        self.assertEqual(snapshot['profesionales'], [])
        self.assertEqual(snapshot['totales']['cantidad_practicas'], 0)
        self.assertEqual(snapshot['totales']['monto_practicas'], '0.00')

    def test_lista_profesional_muestra_traza_pero_no_suma_monto(self):
        self._anular()
        self.client.force_login(self.residente)

        response = self.client.get(
            reverse('liquidacion:registroestudios_list'),
            {'mes': 7, 'año': 2026},
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'ANULADO')
        self.assertEqual(response.context['total_practicas'], 0)
        self.assertEqual(response.context['total_monto_practicas'], Decimal('0.00'))

    def test_anulacion_deja_preparacion_rrhh_previa_desactualizada(self):
        registro_activo = RegistroEstudiosPorMedico.objects.create(
            sesion_contable=self.sesion,
            medico=self.residente,
            nombre_paciente='Paciente',
            apellido_paciente='Activo',
            dni_paciente='87654321',
            fecha_del_informe=date(2026, 7, 11),
            tipo_obra_social='OTRAS_OS',
            horario='EXTRA',
            cantidad_regiones=1,
            monto_calculado=Decimal('1000.00'),
        )
        RegistroEstudio.objects.create(
            registro=registro_activo,
            estudio=self.estudio,
            cantidad=1,
            contexto='SERVICIO',
        )
        RegistroEstudiosPorMedico.objects.filter(pk=registro_activo.pk).update(
            monto_calculado=Decimal('1000.00'),
        )
        self.sesion.estado = 'CERRADA'
        self.sesion.save(update_fields=['estado'])
        snapshot = construir_snapshot_liquidacion_rrhh(self.sesion)
        PreparacionLiquidacionRRHH.objects.create(
            sesion_contable=self.sesion,
            version=1,
            estado=PreparacionLiquidacionRRHH.ESTADO_PREPARADO,
            destinatarios_json=['rrhh@example.com'],
            asunto='Liquidacion julio',
            cuerpo='Resumen',
            resumen_json=snapshot,
            snapshot_hash=calcular_hash_snapshot(snapshot),
            creado_por=self.admin,
            actualizado_por=self.admin,
        )

        self._anular()
        requisito = evaluar_requisito_rrhh_para_facturar(self.sesion)

        self.assertFalse(requisito['ok'])
        self.assertIn('desactualizada', requisito['mensaje'])
